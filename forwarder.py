import asyncio
import io
import os
import re
import signal
from collections import Counter, OrderedDict
from datetime import datetime, timedelta, timezone

from telebot.async_telebot import AsyncTeleBot
from telebot.types import InputFile, ReactionTypeEmoji

from telethon import TelegramClient, events, utils
from telethon.errors import (AuthKeyDuplicatedError, AuthKeyUnregisteredError,
                             SessionRevokedError, UserDeactivatedBanError)
from telethon.sessions import StringSession
from telethon.tl.functions.messages import SendReactionRequest
from telethon.tl.types import MessageMediaWebPage, ReactionEmoji

# ---------------------------------------------------------------------------
# Existing configuration (unchanged)
# ---------------------------------------------------------------------------

# Read token directly from Railway variables
BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')

ADMIN_ID = 7578145913
KEYWORDS = ['You received']

FORWARD_RULES = {
    # MH X LARRY GROUP 2 (channel) -> CHIME PICCASO
    -1003894781195: [-5350880041],
    # Chime Rev & out no-7 -> CHIME GAFFER
    -1002335630148: [-5580596463],
    # MH x LARRY VENMO -> GAFFER VENMO
    #
    # GAFFER VENMO only, as of 2026-08-18. It used to fan out to PICCASO VENMO
    # as well, which meant one payment moved two sets of books - the venmo side
    # now matches the chime side, where each source feeds exactly one target.
    -5339749243: [-5100231154],
}

# Sources where ONLY bot messages may be forwarded, never humans. Messages from
# people in these chats are ignored even if they match a keyword.
BOT_ONLY_SOURCES = {
    -1002335630148,
}

# ---------------------------------------------------------------------------
# Telethon configuration (new)
#
# Telethon is used for READING ONLY. It logs in as your user account so it can
# see messages posted by other bots, which the Bot API never delivers. Nothing
# is ever sent from the user account - every outgoing message still goes out
# through `bot` (your bot token) below.
# ---------------------------------------------------------------------------

API_ID = os.getenv('TELEGRAM_API_ID')
API_HASH = os.getenv('TELEGRAM_API_HASH')

# Preferred for deployment: a StringSession produced by telethon_login.py.
# Falls back to a local .session file for running on your Mac.
TELETHON_SESSION = os.getenv('TELETHON_SESSION')
TELETHON_SESSION_NAME = os.getenv('TELETHON_SESSION_NAME', 'user_session')

# Which bot's messages to relay. Usernames or numeric ids, comma separated.
# Matching is case insensitive and a leading @ is optional.
# Empty (the default) accepts ANY bot in the source chats, which is what you
# want while more than one notification bot is in play. Humans are never
# accepted on this path. Set it to pin a specific bot, e.g.
# SOURCE_BOTS=Hasan_Transection_Bot
SOURCE_BOTS = os.getenv('SOURCE_BOTS', '')

# Which chats Telethon watches. Defaults to the source chats of FORWARD_RULES.
TELETHON_SOURCE_CHATS = os.getenv('TELETHON_SOURCE_CHATS', '')

TELETHON_ENABLED = bool(API_ID and API_HASH)

# The one exception to "reading only" above, and it is a narrow one: the cashout
# escalation has to reach people PRIVATELY, and a bot cannot do that. A bot may
# only message someone who has already pressed Start on it, and it can never
# address a person by @username at all - only by numeric id. The user account
# has neither limit. So when the bot cannot deliver an escalation DM, the user
# account sends it instead (and reacts, where a channel denies the bot).
# Set USERBOT_SEND=0 to keep the account strictly read-only, accepting that only
# people who have started the bot will get the private warning.
USERBOT_SEND = os.getenv('USERBOT_SEND', '1') != '0'

# Reacting is not sending, and lumping the two together costs the one thing that
# must not be lost. USERBOT_SEND exists to stop the account posting MESSAGES on
# your behalf. The ❤ is a single mark on a message that is already there, and it
# is the ONLY durable record that a cashout was actioned - the open requests
# live in memory and a redeploy wipes them.
#
# The bot token frequently cannot react on somebody else's message in a group,
# so the user account is not a nicety here, it is the route that usually works.
# Tying it to USERBOT_SEND meant turning off userbot messaging silently turned
# off the ❤ as well, and a paid cashout then reads as still outstanding.
USERBOT_REACT = os.getenv('USERBOT_REACT', '1') != '0'

# Railway overlaps deploys: the replacement container boots while the outgoing
# one is still alive. For Telethon that is fatal - two IPs on one auth key and
# Telegram destroys the key permanently, which silently stops every forward
# until someone logs in again. Holding the new container back past the
# changeover keeps the two from ever being connected at the same moment. Costs
# one quiet minute after a deploy.
ON_RAILWAY = bool(os.getenv('RAILWAY_ENVIRONMENT') or os.getenv('RAILWAY_SERVICE_ID')
                  or os.getenv('RAILWAY_PROJECT_ID'))
TELETHON_START_DELAY = int(os.getenv('TELETHON_START_DELAY', '45' if ON_RAILWAY else '0'))

# The Bot API poll is held back for exactly the same reason, which the getUpdates
# 409 was long written off as "harmless". It is not. Telegram hands each update
# to exactly ONE poller, so two containers do not duplicate - they SPLIT. The
# outgoing container is the one holding the open cashout requests in memory, so
# a /out that happens to land on the incoming container finds nothing pending
# and is ignored as ordinary traffic. A real cashout goes unanswered and nothing
# reports it.
#
# Waiting costs nothing here: Telegram queues updates server-side for 24 hours
# and delivers the backlog the moment polling starts. Unlike the userbot, which
# is genuinely deaf while held back, this only defers.
BOT_START_DELAY = int(os.getenv('BOT_START_DELAY', str(TELETHON_START_DELAY)))

# On connect, look back over this window and forward anything that was missed
# while the listener was down. 0 disables the sweep. The cap is a guard against
# a misconfiguration dumping hours of history into a group.
CATCHUP_LOOKBACK_MINUTES = int(os.getenv('CATCHUP_LOOKBACK_MINUTES', '180'))
CATCHUP_MAX = int(os.getenv('CATCHUP_MAX', '25'))
CATCHUP_SCAN_LIMIT = int(os.getenv('CATCHUP_SCAN_LIMIT', '300'))

# The target is read back further than the source. A target group carries more
# traffic than its source - every forward, plus /add and /out confirmations,
# milestones and idle prompts - so an equal limit can run out of room before it
# has covered the same window, and anything past that point would look
# undelivered when it is sitting right there.
CATCHUP_TARGET_SCAN_LIMIT = int(os.getenv('CATCHUP_TARGET_SCAN_LIMIT',
                                          str(CATCHUP_SCAN_LIMIT * 3)))

# ---------------------------------------------------------------------------
# Runtime state (unchanged)
# ---------------------------------------------------------------------------

forwarded_count = 0
today_count = 0
start_time = datetime.now()
last_messages = []

userbot_status = 'disabled'

# Set once run_userbot() has a client, so the shutdown handler can release it.
_active_client = None

bot = AsyncTeleBot(BOT_TOKEN)


async def notify_admin(msg):
    try:
        await send_group(ADMIN_ID, msg)
    except Exception as e:
        print(f"Notify Admin Error: {e}", flush=True)


class _ConflictWatcher:
    """Notices when a second process is polling this same bot token.

    It has to hook telebot's exception_handler rather than wrap bot.polling():
    with none_stop=True, polling() catches the 409 internally, logs it and
    carries on, so nothing ever propagates out for a try/except to see.

    A 409 is worth shouting about because it is never only a polling problem.
    Every process polling this token is also running run_userbot() with the
    same TELETHON_SESSION from a different IP, and that is what makes Telegram
    destroy the auth key and stop all forwarding.

    It is also not the harmless thing it was once written off as. Telegram gives
    each update to exactly ONE poller, so two containers SPLIT the updates
    rather than duplicating them - and the open cashout requests live in the
    memory of one of them. A /out delivered to the wrong container finds nothing
    pending and is dropped as ordinary traffic."""

    # run_bot() now waits BOT_START_DELAY before its first poll, so the deploy
    # changeover should be over before this can fire at all. A conflict is
    # therefore no longer the expected cost of a deploy, and the threshold is
    # low: it means the hold was too short, or there really are two services.
    ALERT_AFTER = int(os.getenv('BOT_CONFLICT_ALERTS', '5'))

    def __init__(self):
        self.conflicts = 0
        self.warned = False

    async def handle(self, exception):
        text = str(exception)
        if '409' not in text and 'terminated by other getUpdates' not in text:
            return False        # not ours: let telebot log it as usual

        self.conflicts += 1
        print(f"⚠️ [CONFLICT x{self.conflicts}] Another process is polling this "
              f"token.", flush=True)

        if self.conflicts >= self.ALERT_AFTER and not self.warned:
            self.warned = True
            await notify_admin(
                "🚨 TWO BOT INSTANCES ARE RUNNING on the same token.\n\n"
                "Telegram keeps cutting one of them off (getUpdates 409), and it "
                "splits incoming messages between them - a /out can land on the "
                "instance that never saw the request, and be ignored.\n\n"
                "Both are also using the same TELETHON_SESSION from different IPs, "
                "which destroys the session key and stops ALL forwarding.\n\n"
                f"Polling waits {BOT_START_DELAY}s on start-up specifically to let a "
                "deploy changeover finish, so this is either a second Railway "
                "service on this repo, or that hold is too short - raise "
                "BOT_START_DELAY. Replacing TELETHON_SESSION while this is true "
                "will just burn the new one too.")
        return True             # handled: suppress the repeated stack-trace log


_conflict_watcher = _ConflictWatcher()
bot.exception_handler = _conflict_watcher


def get_uptime():
    uptime = datetime.now() - start_time
    hours = int(uptime.total_seconds() // 3600)
    minutes = int((uptime.total_seconds() % 3600) // 60)
    return f"{hours}h {minutes}m"


# ---------------------------------------------------------------------------
# Shared forwarding pipeline
#
# Both input paths (Bot API for humans, Telethon for bots) funnel into
# process_incoming() so the keyword filter, routing, counters, logging and
# error handling exist in exactly one place.
# ---------------------------------------------------------------------------

def _id_variants(chat_id):
    """Telegram writes supergroup ids as -100<internal> in some places and
    -<internal> in others. Return the alternate spelling of a chat id so a rule
    matches regardless of which form the config uses."""
    text = str(chat_id)
    if text.startswith('-100'):
        return [int('-' + text[4:])]
    if text.startswith('-'):
        return [int('-100' + text[1:])]
    return []


def resolve_rule_key(chat_id):
    """Exact match first so existing behaviour is untouched, then fall back to
    the alternate id spelling."""
    if chat_id in FORWARD_RULES:
        return chat_id
    for variant in _id_variants(chat_id):
        if variant in FORWARD_RULES:
            return variant
    return None


# Nepal time. A fixed offset rather than a zoneinfo lookup: Nepal has no DST,
# and Railway's container has no tzdata installed.
LOCAL_TZ = timezone(timedelta(hours=5, minutes=45))

# Matches the timestamp the notification bot embeds, e.g. "03:35 AM - 30 Jul 2026"
_STAMP_RE = re.compile(
    r'\b\d{1,2}:\d{2}\s*[AaPp][Mm]\s*-\s*\d{1,2}\s+[A-Za-z]{3}\s+\d{4}')

# The notification bot's own clock runs roughly 14 minutes fast, so the time it
# writes into the message is wrong. Rewrite it to when the message was actually
# sent, in Nepal time. Set FIX_TIMESTAMPS=0 to forward the text untouched.
FIX_TIMESTAMPS = os.getenv('FIX_TIMESTAMPS', '1') != '0'


def correct_timestamp(text, sent_at):
    """Replace the bot's embedded timestamp with the real send time. Text with
    no recognisable timestamp is returned unchanged."""
    if not FIX_TIMESTAMPS or sent_at is None:
        return text
    stamp = sent_at.astimezone(LOCAL_TZ).strftime('%I:%M %p - %d %b %Y')
    return _STAMP_RE.sub(stamp, text, count=1)


_seen_messages = OrderedDict()


def _is_duplicate(key):
    """Telethon replays recent messages after a reconnect. Drop anything we
    have already forwarded."""
    if key in _seen_messages:
        return True
    _seen_messages[key] = True
    while len(_seen_messages) > 500:
        _seen_messages.popitem(last=False)
    return False


# ---------------------------------------------------------------------------
# Target group ledger
#
# Each target group keeps its OWN running totals, controlled only by Ethan and
# Larry. A forwarded payment adds to Total In; /add and /out adjust either side.
# The two total lines of a forwarded notification are rewritten to show this
# ledger, so the target group shows your books rather than the source group's.
#
# The bot never does arithmetic on the source group's figures - it reads them
# once to open the ledger, and from then on the ledger is its own.
# ---------------------------------------------------------------------------

MILESTONE_IN_STEP = int(os.getenv('MILESTONE_IN_STEP', '5000'))
MILESTONE_OUT_STEP = int(os.getenv('MILESTONE_OUT_STEP', '1000'))
MILESTONE_MENTIONS = os.getenv('MILESTONE_MENTIONS', '@ethannxxxx @Larryyxx')

# Only these accounts may move the numbers. Numeric ids, so a username change
# can neither break nor hijack it. @ethannxxxx and @Larryyxx.
LEDGER_ADMINS = {7418675217, 7578145913}

# Every chat that receives forwards.
TARGET_CHATS = {t for targets in FORWARD_RULES.values() for t in targets}

# Groups the bot still keeps for, though nothing is routed to them any more.
# PICCASO VENMO stopped being fed on 2026-08-18 and is still a live group with
# the same people in it, so what it stops getting is FORWARDS - not its books
# and not its rules. A group put back into FORWARD_RULES is covered by
# TARGET_CHATS again automatically, so nothing here needs undoing.
FORMER_TARGETS = {-5306739731}                      # PICCASO VENMO

# What send_group() scrubs on the way out - see strip_identities(). Redaction
# follows who is READING a group, not which table its id is in today, and the
# audience of a group that stopped being fed has not changed at all.
REDACTED_CHATS = TARGET_CHATS | FORMER_TARGETS

# Whose books the bot keeps: /add, /out and /set, and the recovery that reads
# them back after a deploy. The two belong together and must never be split -
# a group where the commands work but recovery does not would show figures the
# next redeploy silently reverts, which is the one thing the ledger rules exist
# to prevent. Payments arriving is a separate question from books being kept.
LEDGER_CHATS = TARGET_CHATS | FORMER_TARGETS

BOT_ID = int(BOT_TOKEN.split(':')[0]) if BOT_TOKEN and ':' in BOT_TOKEN else None

_TOTAL_IN_RE = re.compile(r'Total\s*In\s*:?\s*([\d,]+(?:\.\d+)?)', re.I)
_TOTAL_OUT_RE = re.compile(r'Total\s*Out\s*:?\s*([\d,]+(?:\.\d+)?)', re.I)

# Same patterns, split so the number can be substituted in place.
_TOTAL_IN_SUB = re.compile(r'(Total\s*In\s*:?\s*)([\d,]+(?:\.\d+)?)', re.I)
_TOTAL_OUT_SUB = re.compile(r'(Total\s*Out\s*:?\s*)([\d,]+(?:\.\d+)?)', re.I)

_RECEIVED_RE = re.compile(r'You\s+received\s+\$?\s*([\d,]+(?:\.\d+)?)', re.I)
# Amount for /add and /out, with an optional minus written either way round:
# "/add -100" and "/add-100" both take 100 back off Total In. The joined form
# needs its own handler as well - Telegram reads the command as the word after
# the slash, so "/add-100" is the command "add-100" and never reaches a
# commands=['add'] handler at all. This only parses it.
_CMD_AMOUNT_RE = re.compile(r'^/\w+(?:@\w+)?\s*(-?)\s*\$?\s*([\d,]+(?:\.\d+)?)')

# The command word itself, tolerating the joined form: "/add-100" -> "add".
_CMD_NAME_RE = re.compile(r'^/(\w+)')

# Which spellings need routing past telebot's command matcher.
_JOINED_LEDGER_RE = re.compile(r'^/(?:add|out)(?:@\w+)?-', re.I)
_SET_RE = re.compile(r'^/set(?:@\w+)?\s+(in|out)\s+\$?\s*([\d,]+(?:\.\d+)?)', re.I)

# target chat id -> {'in': float, 'out': float}
_ledger = {}


def _to_float(raw):
    try:
        return float(raw.replace(',', ''))
    except (AttributeError, ValueError):
        return None


def parse_received_amount(text):
    """The dollar figure out of 'You received $15.0 from Gabriel W.'"""
    m = _RECEIVED_RE.search(text)
    return _to_float(m.group(1)) if m else None


def rewrite_totals(text, total_in, total_out):
    """Swap the two total lines for the ledger's figures, preserving the rest of
    the line exactly. Text without a totals block is returned untouched."""
    if not _TOTAL_IN_SUB.search(text):
        return text
    text = _TOTAL_IN_SUB.sub(lambda m: f"{m.group(1)}{total_in:,.2f}", text, count=1)
    text = _TOTAL_OUT_SUB.sub(lambda m: f"{m.group(1)}{total_out:,.2f}", text, count=1)
    return text


def parse_totals(text):
    """Pull (total_in, total_out) out of a notification. Either may be None."""
    def num(pattern):
        m = pattern.search(text)
        if not m:
            return None
        try:
            return float(m.group(1).replace(',', ''))
        except ValueError:
            return None
    return num(_TOTAL_IN_RE), num(_TOTAL_OUT_RE)


def crossed_milestone(previous, current, step):
    """Highest step boundary newly crossed, or None.

    A decrease returns None: the bot's totals reset at the start of a new cycle,
    and a reset must not replay alerts for ground already covered."""
    if step <= 0 or previous is None or current is None or current <= previous:
        return None
    if int(current // step) > int(previous // step):
        return int(current // step) * step
    return None


def _money(value):
    return f"{value:,.2f}$" if value is not None else "n/a"


def _step_label(step):
    return f"{step // 1000}K" if step >= 1000 and step % 1000 == 0 else f"{step:,}"


def in_milestone_text(level, total_in, total_out):
    return (f"🎉 MILESTONE HIT — ${level:,} IN! 🎉\n\n"
            f"➕ Total In : {_money(total_in)}\n"
            f"➖ Total Out: {_money(total_out)}\n\n"
            "Outstanding work, team! 🔥\n\n"
            "Every single deposit stacked up to this — that is\n"
            "consistency, not luck. The engine is running and\n"
            "the numbers speak for themselves.\n\n"
            f"Onward to the next {_step_label(MILESTONE_IN_STEP)}! 💪\n\n"
            f"{MILESTONE_MENTIONS}\n\n"
            "-ETHAN")


def out_milestone_text(level, total_in, total_out):
    return (f"⚠️ OUT ALERT — ${level:,} CROSSED ⚠️\n\n"
            f"➕ Total In : {_money(total_in)}\n"
            f"➖ Total Out: {_money(total_out)}\n\n"
            "The out is climbing and needs a look before it\n"
            "runs further ahead. Worth a quick review of what\n"
            "is going out and why.\n\n"
            f"Please check in: {MILESTONE_MENTIONS}\n\n"
            "-ETHAN")


async def check_milestones(target, before, after):
    """Post an alert to one target group when its ledger crosses a threshold.

    `before` and `after` are (total_in, total_out) for that group's ledger."""
    alerts = []
    level = crossed_milestone(before[0], after[0], MILESTONE_IN_STEP)
    if level:
        alerts.append(('IN', level, in_milestone_text(level, after[0], after[1])))
    level = crossed_milestone(before[1], after[1], MILESTONE_OUT_STEP)
    if level:
        alerts.append(('OUT', level, out_milestone_text(level, after[0], after[1])))

    for kind, level, body in alerts:
        print(f"🏁 [MILESTONE {kind}] ${level:,} crossed in {target}", flush=True)
        try:
            await send_group(target, body)
            print(f"   🎯 {kind} milestone sent to {target}", flush=True)
        except Exception as e:
            print(f"   ❌ milestone send failed for {target}: {e}", flush=True)


def ledger_snapshot(target):
    """Current figures for a target, or zeroes if it has no ledger yet."""
    led = _ledger.get(target)
    return (led['in'], led['out']) if led else (0.0, 0.0)


def ledger_preview_payment(target, text):
    """What a forwarded payment WOULD do, without committing anything.

    Returns (before, after) as (in, out) tuples, or None when there is nothing
    to show - a message with no totals block and no ledger yet."""
    if target not in _ledger:
        # Opening balance is the source group's current figures. The payment in
        # this very message is already inside that figure, so it is not added
        # again - otherwise the first message would be counted twice.
        src_in, src_out = parse_totals(text)
        if src_in is None and src_out is None:
            return None
        opening = (src_in or 0.0, src_out or 0.0)
        return opening, opening

    before = ledger_snapshot(target)
    amount = parse_received_amount(text) or 0.0
    return before, (before[0] + amount, before[1])


def ledger_commit(target, after, note=''):
    """Store new figures.

    Called ONLY once the message carrying them is on the wire. The ledger is
    always exactly what the newest message in the group shows, so a failed send
    can never leave the books ahead of what anyone can see."""
    opening = target not in _ledger
    _ledger[target] = {'in': after[0], 'out': after[1]}
    print(f"📒 [LEDGER] {target} {'opened at' if opening else 'now'} "
          f"in={after[0]:,.2f} out={after[1]:,.2f}{note}", flush=True)


# ---------------------------------------------------------------------------
# Idle payment watchdog
#
# When a target group stops receiving payments, ask the group whether there is a
# problem. Repeats every interval while it stays quiet, with different wording
# each time, and resets the moment a payment arrives.
# ---------------------------------------------------------------------------

IDLE_ALERT_MINUTES = int(os.getenv('IDLE_ALERT_MINUTES', '90'))
IDLE_ALERT_CHATS = {-5580596463, -5350880041}      # CHIME GAFFER, CHIME PICCASO

IDLE_NAMES = {-5580596463: 'CHIME GAFFER', -5350880041: 'CHIME PICCASO'}
# Short names so the groups can be targeted from a private chat.
IDLE_ALIASES = {'gaffer': -5580596463, 'piccaso': -5350880041}

# Rotated so a repeat never reads identically. The duration line differs every
# time in any case, so even a full cycle cannot produce a duplicate message.
_IDLE_BODIES = [
    ("⏳ No payments here for {duration}.",
     "Is everything okay on the payment side, or are you\n"
     "facing any issues? Let us know if there is any!"),
    ("⏳ Still quiet — {duration} without a payment.",
     "Checking in again. Is everything running fine on\n"
     "your side, or is something holding payments up?\n"
     "Let us know if there is any issue!"),
    ("⏳ {duration} with no payments.",
     "Any update on the payments? If something is stuck\n"
     "or not working, tell us and we will sort it out.\n"
     "Let us know if there is any issue!"),
    ("⏳ {duration} quiet now.",
     "Nothing has come through in a while. Is there a\n"
     "problem at your end? A quick word either way would\n"
     "help — let us know if there is any issue!"),
]

# target -> {'last': datetime|None, 'since': datetime, 'sent': int}
_idle_state = {}


def _idle_slot(target):
    return _idle_state.setdefault(target, {'last': None,
                                           'since': datetime.now(timezone.utc),
                                           'sent': 0,
                                           'paused': False})


def _humanise(minutes):
    hours, mins = divmod(int(minutes), 60)
    parts = []
    if hours:
        parts.append(f"{hours} hour" + ("s" if hours != 1 else ""))
    if mins:
        parts.append(f"{mins} minute" + ("s" if mins != 1 else ""))
    return " ".join(parts) or "0 minutes"


async def note_payment(target):
    """A payment landed: restart the clock and the wording rotation.

    A payment also lifts a pause - if money is moving again, whatever was wrong
    has cleared, and a pause nobody remembers would silently kill the watchdog."""
    if target not in IDLE_ALERT_CHATS or chat_paused(target):
        return
    slot = _idle_slot(target)
    now = datetime.now(timezone.utc)
    slot['last'] = now
    slot['since'] = now
    slot['sent'] = 0

    if slot['paused']:
        # Silently re-armed. Nothing is posted: the group never sees chatter
        # about the watchdog's own state.
        slot['paused'] = False
        print(f"🔔 [IDLE] {target} auto-resumed on payment", flush=True)


def set_idle_paused(target, paused):
    slot = _idle_slot(target)
    slot['paused'] = paused
    if not paused:
        # Fresh clock, so resuming after a long pause cannot fire instantly.
        slot['since'] = datetime.now(timezone.utc)
        slot['sent'] = 0


def idle_alert_text(target, quiet_minutes, index):
    """Just the elapsed time and the question - no figures. The ledger belongs
    in payment messages, not in a prompt asking whether something is wrong."""
    header, question = _IDLE_BODIES[index % len(_IDLE_BODIES)]
    return (f"{header.format(duration=_humanise(quiet_minutes))}\n\n"
            f"{question}\n\n"
            "-ETHAN")


async def idle_watchdog():
    if IDLE_ALERT_MINUTES <= 0:
        print("ℹ️ [IDLE] watchdog disabled (IDLE_ALERT_MINUTES=0)", flush=True)
        return

    for target in IDLE_ALERT_CHATS:
        _idle_slot(target)          # clock starts now, not at epoch
    print(f"⏳ [IDLE] watching {len(IDLE_ALERT_CHATS)} groups, "
          f"prompt every {_humanise(IDLE_ALERT_MINUTES)} of silence", flush=True)

    while True:
        await asyncio.sleep(60)
        now = datetime.now(timezone.utc)
        for target in sorted(IDLE_ALERT_CHATS):
            if chat_paused(target):
                continue            # out of service: no prompt, no clock
            slot = _idle_slot(target)
            if slot['paused']:
                continue
            quiet = (now - slot['since']).total_seconds() / 60.0
            due = IDLE_ALERT_MINUTES * (slot['sent'] + 1)
            if quiet < due:
                continue
            try:
                await send_group(target, idle_alert_text(target, due, slot['sent']))
                slot['sent'] += 1
                print(f"⏳ [IDLE] {target} quiet {_humanise(due)}, "
                      f"prompt #{slot['sent']} sent", flush=True)
            except Exception as e:
                print(f"❌ [IDLE] prompt failed for {target}: {e}", flush=True)


# ---------------------------------------------------------------------------
# Cashout requests
#
# Payments travel outward, source -> target. Cashout requests travel the other
# way: the chime groups ask for money to go OUT, and that request has to be seen
# and actioned by the crew on the other side.
#
#   CHIME PICCASO --"CASHOUT REQUEST"--> MH X LARRY GROUP 2
#   CHIME GAFFER  --"CASHOUT REQUEST"--> Chime Rev & out no-7
#
# The request is posted with the crew tagged. If nobody answers it inside the
# timeout it is re-tagged in the group AND each of them is warned privately.
# When somebody finally replies with /out, that reply is sent back to the chime
# group the request came from and the forwarded request is hearted.
#
# This is a SEPARATE table from FORWARD_RULES on purpose. Putting these pairs in
# there would drag the chime groups into TARGET_CHATS and make the ledger, the
# milestones and the idle watchdog all fire on the wrong side of the flow.
# ---------------------------------------------------------------------------

CASHOUT_KEYWORD = os.getenv('CASHOUT_KEYWORD', 'CASHOUT REQUEST')

def _parse_cashout_routes(raw):
    """asked_in:handled_in[:out_goes_to], comma separated.

    The third field is optional and defaults to the group that asked, which is
    the ordinary case: a cashout is requested in a chime group and the /out
    answering it comes back to that same group. Give it explicitly to send the
    /out somewhere else.

    Config rather than code so a group can be changed from Railway without a
    redeploy. A malformed entry is dropped with a warning rather than taking
    the whole bot down - one bad route must not stop the others working."""
    routes = {}
    for item in raw.split(','):
        item = item.strip()
        if not item:
            continue
        parts = [p.strip() for p in item.split(':')]
        if len(parts) not in (2, 3):
            print(f"⚠️ [CASHOUT] ignoring route {item!r} - expected "
                  "asked_in:handled_in or asked_in:handled_in:out_goes_to", flush=True)
            continue
        try:
            source, handling = int(parts[0]), int(parts[1])
            out_to = int(parts[2]) if len(parts) == 3 else source
        except ValueError:
            print(f"⚠️ [CASHOUT] ignoring route {item!r} - chat ids must be numbers",
                  flush=True)
            continue
        routes[source] = {'handling': handling, 'out_to': out_to}
    return routes


CASHOUT_ROUTES = _parse_cashout_routes(os.getenv(
    'CASHOUT_ROUTES',
    '-5350880041:-1003894781195,'        # CHIME PICCASO -> MH X LARRY GROUP 2
    '-5580596463:-1002335630148'))       # CHIME GAFFER  -> Chime Rev & out no-7

# Reverse lookup, for recognising a message posted in a handling group.
_CASHOUT_HANDLERS = {route['handling']: source
                     for source, route in CASHOUT_ROUTES.items()}

CHAT_NAMES = {
    -5350880041: 'CHIME PICCASO',
    -5580596463: 'CHIME GAFFER',
    -1003894781195: 'MH X LARRY GROUP 2',
    -1002335630148: 'Chime Rev & out no-7',
}

# Labels for any group added later, so the logs and the admin alerts stay
# readable instead of falling back to raw ids. CHAT_NAMES=-100123=SOME GROUP,...
for _item in os.getenv('CHAT_NAMES', '').split(','):
    if '=' in _item:
        _cid, _, _label = _item.partition('=')
        try:
            CHAT_NAMES[int(_cid.strip())] = _label.strip()
        except ValueError:
            print(f"⚠️ [CASHOUT] ignoring chat name {_item!r} - id must be a number",
                  flush=True)

# Tagged on the forwarded request and again on every reminder.
CASHOUT_MENTIONS = os.getenv('CASHOUT_MENTIONS',
                             '@Maynuddin23 @MHSUPPORTZONE @maynuddin233')

# Anyone here speaking in the handling group counts as "they responded" and puts
# the reminder clock back to zero. The request itself stays OPEN until a real
# /out lands - an acknowledgement is not a cashout.
CASHOUT_RESPONDERS = {h.strip().lower().lstrip('@') for h in
                      os.getenv('CASHOUT_RESPONDERS',
                                'Maynuddin23,MHSUPPORTZONE,maynuddin233').split(',')
                      if h.strip()}

def _handle_list(raw):
    return [h.strip().lstrip('@') for h in raw.split(',') if h.strip()]


def _parse_group_crew(raw):
    """chat_id=handle handle,chat_id=handle - crew for ONE group only.

    Keyed by the HANDLING group, because that is the side crew work on: the
    group a request is posted into and the group the /out comes back from.

    A malformed entry is dropped with a warning rather than taking the bot
    down with it - one unreadable id must not cost the other routes their
    crew, exactly as _parse_cashout_routes decided."""
    groups = {}
    for item in raw.split(','):
        item = item.strip()
        if not item:
            continue
        cid, _, handles = item.partition('=')
        try:
            chat_id = int(cid.strip())
        except ValueError:
            print(f"⚠️ [CASHOUT] ignoring group crew {item!r} - expected "
                  "chat_id=handle", flush=True)
            continue
        names = [h.strip().lstrip('@') for h in handles.split() if h.strip()]
        if names:
            groups.setdefault(chat_id, []).extend(names)
    return groups


# Two audiences, two different messages.
#
# Ethan and Larry are told WHERE a request came from and where it is stuck,
# because chasing it is their job and that means knowing which groups are
# involved. Their ids are seeded in _user_ids below, so it reaches them from the
# bot itself rather than from the user account.
CASHOUT_ADMIN_HANDLES = _handle_list(
    os.getenv('CASHOUT_ADMIN_DM_HANDLES', 'ethannxxxx,larryyxx'))

# The crew get the same alert with none of the routing. Which groups a request
# travelled between is not theirs to see - they need the request itself and
# what to do about it, nothing more.
CASHOUT_CREW_HANDLES = _handle_list(
    os.getenv('CASHOUT_CREW_DM_HANDLES', 'Maynuddin23,MHSUPPORTZONE,maynuddin233'))

# Crew who work ONE handling group rather than both.
#
# The three above are crew everywhere. Somebody taken on for one side is not
# crew on the other, and that difference is not cosmetic: tagging them in the
# other group asks them to pay a cashout that was never theirs, and counting a
# reaction of theirs there would stop the chase for the people who actually
# have to do it.
#
# Everything else about them is identical to the three above - tagged on the
# request and on every reminder, their reaction acknowledges it, their /out is
# relayed, booked and hearted, they get the last-resort DM, and their name is
# stripped out of anything bound for a chime group.
CASHOUT_GROUP_CREW = _parse_group_crew(os.getenv(
    'CASHOUT_GROUP_CREW',
    '-1002335630148=NPR_CA'))        # prutok sha - Chime Rev & out no-7 only

# Told how a request is progressing rather than that it is stuck: picked up,
# then completed. Larry sends the screenshot afterwards, so he is the one who
# needs to know the moment there is something to send.
CASHOUT_PROGRESS_HANDLES = _handle_list(
    os.getenv('CASHOUT_PROGRESS_DM_HANDLES', 'larryyxx'))

# How long a request may sit before the first chase, and the master switch for
# the whole ladder: 0 disables chasing altogether.
CASHOUT_TIMEOUT_MINUTES = int(os.getenv('CASHOUT_TIMEOUT_MINUTES', '5'))


def _minute_ladder(raw, fallback):
    """A comma-separated list of minutes, sorted and de-duplicated.

    Anything unreadable falls back rather than emptying the ladder. A typo in a
    Railway variable must not be the thing that quietly stops cashouts being
    chased - that failure looks exactly like a quiet day."""
    try:
        rungs = sorted({int(m.strip()) for m in raw.split(',') if m.strip()})
    except ValueError:
        print(f"⚠️ [CASHOUT] unreadable CASHOUT_NUDGE_MINUTES {raw!r} - "
              f"using {fallback}", flush=True)
        return list(fallback)
    return [m for m in rungs if m > 0] or list(fallback)


# The group ladder, in minutes SINCE THE REQUEST WAS POSTED - absolute marks,
# not gaps between reminders. An unanswered request is re-tagged in the handling
# group at 5, 8 and 12 minutes, and then the group falls silent.
#
# Acknowledging it does not move these. A reaction used to silence the group and
# start a private ladder instead; it no longer does, because what is being
# chased is the /out, and "seen" is not "paid".
CASHOUT_NUDGE_MINUTES = _minute_ladder(
    os.getenv('CASHOUT_NUDGE_MINUTES', '5,8,12'), [5, 8, 12])

# A request nobody has acknowledged by this point earns the crew one private
# word - once, and only ever for silence. Reacting to it stops this from firing
# at all; see CASHOUT_SEEN_DM_MINUTES. 0 removes the private step.
#
# It may fall BETWEEN two group rungs, and that is fine: it does not end the
# group ladder, and the ladder does not end it.
CASHOUT_CREW_DM_MINUTES = int(os.getenv('CASHOUT_CREW_DM_MINUTES', '10'))

# Once the crew HAVE acknowledged it, the group is left alone and this is how
# long they get before Larry is told it is still not processed. Measured from
# the acknowledgement, not from when the request opened.
CASHOUT_SEEN_DM_MINUTES = int(os.getenv('CASHOUT_SEEN_DM_MINUTES', '6'))

# A hard cap on group rounds, on top of the ladder above; 0 lets the ladder
# alone decide. Stopping is not giving up: the request stays OPEN, so a late
# /out is still forwarded, booked and hearted.
CASHOUT_MAX_NUDGES = int(os.getenv('CASHOUT_MAX_NUDGES', '3'))

# How long to hold the "answered without a /out" alert before sending it.
#
# The crew routinely post the message first and EDIT the /out onto it a few
# seconds later. Firing the moment the first version lands turns the ordinary
# way they work into an alert, every single time. Holding it means the edit
# arrives first and there is nothing to report. 0 sends immediately.
CASHOUT_ISSUE_DELAY_SECONDS = int(os.getenv('CASHOUT_ISSUE_DELAY_SECONDS', '60'))

# The same request arriving twice inside this many seconds is one request
# delivered twice, not two cashouts, and only the first is posted. Beyond the
# window an identical request is taken at face value - asking for the same
# amount to the same cashtag twice in a day is ordinary. 0 disables the guard.
CASHOUT_DEDUP_SECONDS = int(os.getenv('CASHOUT_DEDUP_SECONDS', '120'))

CASHOUT_HEART = os.getenv('CASHOUT_HEART', '❤')

# ---------------------------------------------------------------------------
# Mention watch
# ---------------------------------------------------------------------------
#
# The groups are MUTED. An @ in one of them therefore reaches nobody until
# somebody happens to scroll back through it, which is exactly the silent
# failure this bot exists to remove elsewhere. A private message is the only
# thing that actually arrives.

# Watched in the four groups on the cashout routes - both chime groups and both
# handling groups. The VENMO targets are deliberately out: nobody is chased
# there. Override with a comma-separated list of ids.
def _default_mention_chats():
    chats = []
    for _source, _route in CASHOUT_ROUTES.items():
        for _cid in (_source, _route['handling']):
            if _cid not in chats:
                chats.append(_cid)
    return chats


def _with_variants(chat_ids):
    """Both spellings of every id, so a match cannot depend on which form
    Telegram happened to use - see _id_variants."""
    both = set()
    for chat_id in chat_ids:
        both.add(chat_id)
        both.update(_id_variants(chat_id))
    return both


_raw_mention_chats = os.getenv('MENTION_CHATS', '').strip()
MENTION_CHATS = _with_variants(
    [int(c.strip()) for c in _raw_mention_chats.split(',') if c.strip()]
    if _raw_mention_chats else _default_mention_chats())

# Whose mentions to watch for, and who hears about them.
MENTION_WATCH_HANDLES = _handle_list(
    os.getenv('MENTION_WATCH_HANDLES', 'ethannxxxx,larryyxx'))
MENTION_ALERT_HANDLES = _handle_list(
    os.getenv('MENTION_ALERT_HANDLES', 'ethannxxxx,larryyxx'))

# "@handle" as a whole word. The lookbehind stops an email tail or a doubled @
# from counting, and \b stops @larryyxx matching inside @larryyxxx.
_MENTION_RE = (re.compile(
    r'(?<![\w@])@(' + '|'.join(re.escape(h) for h in MENTION_WATCH_HANDLES) + r')\b',
    re.I) if MENTION_WATCH_HANDLES else None)

# ---------------------------------------------------------------------------
# Taking a group out of service
#
# A PAUSE, not an unwiring. Deleting a group from FORWARD_RULES and
# CASHOUT_ROUTES would stop it too, but putting it back means rebuilding those
# tables by hand and hoping the rebuild is exact - and they are what the
# report, the mention watch, the idle watchdog and the ledger all derive
# themselves from. The routes stay exactly as they are. One gate makes the bot
# deaf and mute in the listed chats, and taking an id back out of the list is
# the whole of turning it on again, with every rule it had before still
# written down where it always was.
#
# Paused means paused: nothing forwarded, no cashout opened, answered, hearted
# or chased, no ledger moved, no milestone, no idle prompt, no mention DM, no
# retraction, no report, and no reply to a command typed there.
#
# It is also SILENT, and that is the one place it differs from the emergency
# stop. The stop reports what it swallowed, because it is an incident and
# somebody has to pick the money up by hand. A pause is a decision that those
# groups are not the bot's business for now, so it reads as absent: nothing in
# the groups, nothing to the crew, and nothing to Ethan and Larry beyond the
# console and the line /status leads with.
# ---------------------------------------------------------------------------

# Worked from a private chat with /group off and /group on, and the state
# SURVIVES A REDEPLOY - the same way the emergency stop does, by living in a
# message the bot posted to Ethan and Larry rather than in memory. See
# set_groups_paused() and recover_group_switch() below.
#
# This default is only what is true before the first marker exists. Once
# somebody has worked the switch, the marker is the authority and this line
# stops mattering - it is kept as the state the bot boots into while the
# private chat is being read back, which is deliberately the SILENT one.
PAUSED_CHATS_DEFAULT = os.getenv('PAUSED_CHATS', '-1003894781195,-5350880041')


def _parse_chat_list(raw):
    ids = []
    for item in raw.split(','):
        item = item.strip()
        if not item:
            continue
        try:
            ids.append(int(item))
        except ValueError:
            print(f"⚠️ [PAUSED] ignoring {item!r} - chat ids must be numbers",
                  flush=True)
    return ids


# As configured, for anything a person reads; with both id spellings, for
# anything the code tests membership against - see _id_variants.
PAUSED_CHAT_IDS = _parse_chat_list(PAUSED_CHATS_DEFAULT)
PAUSED_CHATS = _with_variants(PAUSED_CHAT_IDS)


def set_paused_chats(chat_ids):
    """Replace the out-of-service set, keeping both spellings in step.

    One function rather than two assignments: the readable list and the
    membership set have to move together, and a gate testing a set that no
    longer matches the list somebody was shown is the worst of both."""
    global PAUSED_CHAT_IDS, PAUSED_CHATS
    PAUSED_CHAT_IDS = sorted(set(chat_ids), key=lambda c: (chat_name(c), c))
    PAUSED_CHATS = _with_variants(PAUSED_CHAT_IDS)


def _parse_resume_marks(raw):
    """chat_id@2026-08-18T14:30 - when a group came back, UTC.

    Set this in the same change that takes an id out of PAUSED_CHATS. It is
    what stops the boot sweep handing a resumed group the history it
    deliberately missed; see catch_up(). Self-expiring - once the resume is
    further back than CATCHUP_LOOKBACK_MINUTES it does nothing at all, and the
    entry can be deleted whenever anyone next passes it."""
    marks = {}
    for item in raw.split(','):
        item = item.strip()
        if not item:
            continue
        cid, _, stamp = item.partition('@')
        try:
            chat_id = int(cid.strip())
            when = datetime.fromisoformat(stamp.strip())
        except ValueError:
            print(f"⚠️ [PAUSED] ignoring resume mark {item!r} - expected "
                  "chat_id@YYYY-MM-DDTHH:MM", flush=True)
            continue
        if when.tzinfo is None:
            when = when.replace(tzinfo=timezone.utc)
        marks[chat_id] = when
        for variant in _id_variants(chat_id):
            marks[variant] = when
    return marks


RESUMED_CHATS = _parse_resume_marks(os.getenv('RESUMED_CHATS', ''))


def note_resumed(chat_ids, when):
    """Remember that these chats came back at this moment."""
    for chat_id in chat_ids:
        RESUMED_CHATS[chat_id] = when
        for variant in _id_variants(chat_id):
            RESUMED_CHATS[variant] = when


def _live_resume_marks():
    """The resume marks that still matter, one entry per chat.

    A mark older than the boot sweep's own window cannot change what that
    sweep does, so it is dropped rather than carried in the marker for ever.
    Variants are collapsed out: this is what gets written down, and writing
    each chat twice would make the record harder to read than the thing it
    records."""
    horizon = datetime.now(timezone.utc) - timedelta(minutes=CATCHUP_LOOKBACK_MINUTES)
    live = {}
    for chat_id, when in RESUMED_CHATS.items():
        if when < horizon:
            continue
        if any(variant in live for variant in _id_variants(chat_id)):
            continue
        live[chat_id] = when
    return live


def chat_paused(chat_id):
    """Is this exact chat out of service?"""
    return chat_id in PAUSED_CHATS


def resumed_at(chat_id):
    """When this chat came back, or None."""
    return RESUMED_CHATS.get(chat_id)


def route_paused(chat_id):
    """Is this chat, or the group at the other end of its route, out of service?

    Half a route running is worse than none of it. A request forwarded into a
    group nobody is reading is a cashout nobody will pay; a payment booked in
    a chime group whose source has been silenced leaves the books telling a
    story the groups cannot see. So pausing either end stops the pair."""
    if chat_paused(chat_id):
        return True

    source = _canonical(chat_id, CASHOUT_ROUTES)
    if source is not None:
        route = CASHOUT_ROUTES[source]
        return chat_paused(route['handling']) or chat_paused(route['out_to'])

    handling = _canonical(chat_id, _CASHOUT_HANDLERS)
    if handling is not None:
        origin = _CASHOUT_HANDLERS[handling]
        return chat_paused(origin) or chat_paused(CASHOUT_ROUTES[origin]['out_to'])

    rule_key = resolve_rule_key(chat_id)
    if rule_key is not None:
        return any(chat_paused(target) for target in FORWARD_RULES[rule_key])
    return False


def route_members(chat_id):
    """Every chat that goes out of service together with this one.

    A route is ONE unit, and naming either end takes the whole of it: the
    chime group that asks, the handling group that pays it, wherever the /out
    is sent if that is somewhere else, and the payment direction between the
    same two groups. Leaving half a route running is the exact shape of
    trouble route_paused() exists to prevent, and it would be far too easy to
    create from a phone by naming one group and assuming the other followed."""
    members = {chat_id}

    source = _canonical(chat_id, CASHOUT_ROUTES)
    handling = _canonical(chat_id, _CASHOUT_HANDLERS)
    if handling is not None:
        source = _CASHOUT_HANDLERS[handling]
    if source is not None:
        route = CASHOUT_ROUTES[source]
        members.update((source, route['handling'], route['out_to']))

    for member in list(members):
        rule_key = resolve_rule_key(member)
        if rule_key is not None:
            members.add(rule_key)
            members.update(FORWARD_RULES[rule_key])
    return members


# What a group may be called when working the switch by hand. Short, because
# these get typed on a phone, and every group answers to something obvious.
GROUP_ALIASES = {
    'piccaso': -5350880041,
    'gaffer': -5580596463,
    'larry2': -1003894781195,
    'mhlarry': -1003894781195,
    'rev': -1002335630148,
    'chimerev': -1002335630148,
}

for _item in os.getenv('GROUP_ALIASES', '').split(','):
    if '=' in _item:
        _alias, _, _cid = _item.partition('=')
        try:
            GROUP_ALIASES[_alias.strip().lower()] = int(_cid.strip())
        except ValueError:
            print(f"⚠️ [PAUSED] ignoring alias {_item!r} - id must be a number",
                  flush=True)


def resolve_group(argument):
    """A name or a raw id typed by a person -> the chats it stands for.

    Returns None for anything not recognised. A raw id is accepted because the
    aliases cannot cover a group added in Railway, and being unable to work
    the switch is worse than typing a number."""
    argument = (argument or '').strip().lower().lstrip('@')
    if not argument:
        return None
    named = GROUP_ALIASES.get(argument)
    if named is None:
        try:
            named = int(argument)
        except ValueError:
            return None
        if (_canonical(named, CASHOUT_ROUTES) is None
                and _canonical(named, _CASHOUT_HANDLERS) is None
                and resolve_rule_key(named) is None
                and named not in TARGET_CHATS):
            return None                 # a number, but not a group we work
    return route_members(named)


def _known_groups():
    """Every chat on a route this bot works, in a stable, readable order."""
    known = set()
    for chat in list(CASHOUT_ROUTES) + list(FORWARD_RULES):
        known.update(route_members(chat))
    return sorted(known, key=lambda c: (chat_name(c), c))


def group_switch_state_text(state_ids=None):
    """What is in and out of service, for a person reading it on a phone.

    `state_ids` is passed explicitly while the switch is being worked, because
    the message has to describe the state it is the record OF - which, on the
    way back into service, is not yet the state the bot is in."""
    paused = (set(_with_variants(state_ids)) if state_ids is not None
              else PAUSED_CHATS)
    known = _known_groups()
    out = [chat_name(c) for c in known if c in paused]
    live = [chat_name(c) for c in known if c not in paused]

    lines = []
    if out:
        lines.append('⏸️ OUT OF SERVICE\n' + '\n'.join(f'  • {name}' for name in out))
    else:
        lines.append('▶️ Every group is in service.')
    if live:
        lines.append('▶️ RUNNING\n' + '\n'.join(f'  • {name}' for name in live))
    return '\n\n'.join(lines)


# The durable record, and the notification, in one message - exactly as the
# emergency stop does it. Railway wipes the disk on every deploy, so a switch
# held only in memory would put two groups back into service on the next push
# with nobody having asked for it.
GROUP_PAUSE_MARK = '⏸️ GROUPS OUT OF SERVICE'
GROUP_RESUME_MARK = '▶️ GROUPS BACK IN SERVICE'

# Machine-readable lines at the foot of that message. The marker is read back
# by a bot that cannot remember what it wrote, so the state has to be IN the
# text rather than implied by it.
_GROUP_STATE_RE = re.compile(r'^STATE[ \t]*(.*)$', re.M)
_GROUP_RESUMED_RE = re.compile(r'^RESUMED[ \t]*(.*)$', re.M)


def group_switch_marker(pausing, changed, state_ids, who=None):
    """The message that is both the record and the notification."""
    stamp = datetime.now(LOCAL_TZ).strftime('%I:%M %p - %d %b %Y')
    names = ', '.join(chat_name(c) for c in changed) or 'nothing'
    state = ','.join(str(c) for c in sorted(state_ids))
    resumed = ','.join(f"{c}@{when.strftime('%Y-%m-%dT%H:%M')}"
                       for c, when in sorted(_live_resume_marks().items()))

    if pausing:
        head = (f"{GROUP_PAUSE_MARK}\n\n{names}\n\n"
                "Nothing is being forwarded, chased, booked or answered in "
                "them, and nothing about them is being reported - not to the "
                "crew, and not to you. Cashout requests already open there are "
                "kept, not cancelled.\n\n"
                "Send /group on to put them back.")
    else:
        head = (f"{GROUP_RESUME_MARK}\n\n{names}\n\n"
                "Working again exactly as before. The boot sweep will not go "
                "back over the time they were out of service, so nothing from "
                "that window is about to be forwarded or booked.")

    return (f"{head}\n\n"
            f"{group_switch_state_text(state_ids)}\n\n"
            f"Changed at {stamp}{f' by {who}' if who else ''}.\n"
            f"STATE {state}\n"
            f"RESUMED {resumed}")


async def set_groups_paused(chats, pausing, who=None):
    """Take groups out of service, or put them back. Returns who was not told.

    The order differs by direction and both directions fail safe, exactly as
    the emergency stop does. TAKING OUT sets the state FIRST: a marker that
    fails to post must never leave a group being worked when somebody has just
    said to stop. PUTTING BACK posts the marker first and only then lifts the
    pause, so a group is never live without a record saying it should be.

    The resume mark is taken at the same moment as the change, and it is what
    stops the next deploy's sweep replaying the window nobody was working."""
    changed = sorted(chats, key=lambda c: (chat_name(c), c))
    now = datetime.now(timezone.utc)
    after = (set(PAUSED_CHAT_IDS) | set(changed)) if pausing else (
        set(PAUSED_CHAT_IDS) - set(changed))

    if pausing:
        # At once. The point of the switch is that it takes effect when it is
        # pressed, and a marker that fails to post must not leave a group
        # being worked after somebody has said to stop.
        set_paused_chats(after)
    else:
        # Recorded before the marker is written, so the marker carries it and
        # the next deploy's sweep knows where the missed window ends.
        note_resumed(changed, now)

    failed = await dm_handles(CASHOUT_ADMIN_HANDLES,
                              group_switch_marker(pausing, changed, after, who))
    await warn_unreachable(failed)

    if not pausing:
        # The record exists. Only now do the groups go live again.
        set_paused_chats(after)

    print(f"{'⏸️' if pausing else '▶️'} [PAUSED] "
          f"{', '.join(chat_name(c) for c in changed)} "
          f"{'taken out of service' if pausing else 'back in service'}"
          f"{f' by {who}' if who else ''}"
          f"{f' - could not reach {failed}' if failed else ''}", flush=True)
    return failed


# ---------------------------------------------------------------------------
# Payment retraction
# ---------------------------------------------------------------------------
#
# A payment is forwarded and booked within seconds of landing, so by the time
# anyone can see that it should not count, the money is already on the target's
# books and the copy is already sitting in the group. Reacting to the ORIGINAL
# in the source group undoes both: the forwarded copy is deleted, and the amount
# comes back off that target's Total In.
#
# ANY reaction retracts - the user's explicit choice. There is deliberately no
# confirmation step, so a stray tap on a payment in one of these groups really
# does take the money off the books.
RETRACT_SOURCES = _with_variants(
    [int(c.strip()) for c in
     os.getenv('RETRACT_SOURCES', '-1002335630148').split(',') if c.strip()])

# (rule key, source message id) -> [(target, message id there, amount booked)]
#
# In memory, so a redeploy forgets it and a later reaction finds nothing. That
# is an acceptable limit rather than a hidden one: Telegram will not let a bot
# delete a group message more than 48 hours old either, so this was never going
# to work retrospectively.
_delivered = OrderedDict()
RETRACT_MEMORY = int(os.getenv('RETRACT_MEMORY', '400'))

# How deep to read the target group when the record above is gone. A redeploy
# wipes it, and deploys are frequent, so the memory alone made this work only
# for payments forwarded since the last restart - which is not what anyone
# means by "undo that one".
RETRACT_SCAN_LIMIT = int(os.getenv('RETRACT_SCAN_LIMIT', '300'))


def remember_delivery(source_key, target, message_id, amount):
    """Note what a forward put where, so a reaction can undo exactly that.

    The trailing None is the Telethon entity, which only the history fallback
    has. This id came from our own send, so the bot token can address it
    directly and no entity is needed."""
    if source_key is None or message_id is None:
        return
    _delivered.setdefault(source_key, []).append((target, message_id, amount, None))
    while len(_delivered) > RETRACT_MEMORY:
        _delivered.popitem(last=False)

# @username -> numeric id. A bot can only DM an id, so the ones we know are
# seeded here and the rest are learned the first time that person speaks in a
# watched chat. @Larryyxx and @ethannxxxx, matching LEDGER_ADMINS.
_user_ids = {'larryyxx': 7418675217, 'ethannxxxx': ADMIN_ID}

# Set once the userbot knows who it is, so it never DMs its own account.
_userbot_username = ''

# /out anywhere in the message, as a word. "/out", "/out 500", "ok /out done".
_OUT_CMD_RE = re.compile(r'(?:^|\s)/out\b', re.I)

# The figure inside that reply, wherever it sits. The ledger's own _CMD_AMOUNT_RE
# is anchored to the start of the message, which is right for a typed command but
# would miss "ok /out 500" - and this text was typed by someone else, in another
# group, with no reason to keep to that shape.
_OUT_AMOUNT_RE = re.compile(r'/out(?:@\w+)?\s+\$?\s*([\d,]+(?:\.\d+)?)', re.I)

# handling chat id -> [request, ...], oldest first
_pending_cashouts = {}

# handling chat id -> {fingerprint: when it was claimed}
#
# _pending_cashouts is only written AFTER the request has been posted, and
# posting is an await on a network call. asyncio runs the next ready task inside
# that await, so a second copy of the same request arriving in the window
# compares itself against a list that is still empty, passes, and posts again -
# two identical messages in the handling group, two escalation ladders and two
# things to answer for one cashout. That is a single-container bug: it needs no
# second Railway process, only two copies close enough together.
#
# The claim is taken BEFORE the send, with no await between deciding and
# recording, so nothing can interleave. This is what _pending_cashouts cannot
# be: a record of a request that is still in flight.
#
# It is held for the length of the send and no longer - once the request is on
# _pending_cashouts, that list answers the question and the claim is dropped.
# Giving the claim its own window would mean two clocks that can disagree, and
# a claim outliving the request it stood for would refuse the genuine second
# cashout that arrives right after the first is paid.
_cashout_claims = {}


def _claim_cashout(handling, fingerprint, now):
    """Reserve this request text, or report that somebody already holds it.

    Contains no await on purpose. The look and the claim have to be one
    indivisible step, or this is just the same race one level down."""
    if not CASHOUT_DEDUP_SECONDS:
        return True
    claims = _cashout_claims.setdefault(handling, {})
    for held, claimed in list(claims.items()):
        # Only reachable if a send neither returned nor raised for the whole
        # window. Nothing should sit here that long; this is the safety net that
        # stops one wedged send from muting a route for good.
        if (now - claimed).total_seconds() > CASHOUT_DEDUP_SECONDS:
            del claims[held]
    if fingerprint in claims:
        return False
    claims[fingerprint] = now
    return True


def _release_cashout(handling, fingerprint):
    """Give the claim back when the post never happened.

    A claim held for a message that failed to send would swallow the next
    genuine attempt, and the group would then be told nothing at all - a silent
    miss, which is worse than the duplicate this guards against."""
    claims = _cashout_claims.get(handling)
    if claims:
        claims.pop(fingerprint, None)
        if not claims:
            _cashout_claims.pop(handling, None)


# ---------------------------------------------------------------------------
# Emergency stop
#
# One switch that takes the whole cashout flow out of service: nothing
# forwarded, nothing chased, no ledger moved. For when something is going wrong
# and the right move is to stop the bot touching real money until a person has
# looked at it.
#
# It has to survive a redeploy, and that is the hard part. Railway wipes the
# disk and boots a replacement container on every push, so a switch held only in
# memory would turn itself back ON at the worst possible moment - the incident
# is very often the reason somebody is about to push. So the state lives in a
# message, the same way the ledger does, and is read back on boot.
#
# That message goes in the PRIVATE chat with Ethan and Larry, never in a group.
# Which means the marker has to be readable from a private chat, and a bot
# cannot read its own history at all - only the userbot can. It works because
# the userbot IS one of their accounts, so the bot's DM to them is a
# conversation the userbot can open and read back. See recover_cashout_switch().
#
# Silent in the groups, but never silent to a person: anything real that arrives
# while the stop is on is reported to Ethan and Larry rather than dropped.
# ---------------------------------------------------------------------------

CASHOUT_STOP_MARK = '⏸️ CASHOUT FORWARDING STOPPED'
CASHOUT_RESUME_MARK = '▶️ CASHOUT FORWARDING RESUMED'

_cashout_stopped = False
_cashout_stopped_by = None


def cashout_stopped():
    return _cashout_stopped


async def set_cashout_stopped(stopping, who=None):
    """Take the cashout flow out of service, or put it back.

    The order differs by direction, and both directions fail safe. STOPPING
    sets the flag FIRST: the point of the switch is that it takes effect at
    once, and a marker that fails to post must never leave the flow running.
    RESUMING posts the marker first and only then clears the flag, so the flow
    is never live without a record saying so.

    Returns the chats the marker could not be posted to. An empty list means the
    state will survive the next redeploy."""
    global _cashout_stopped, _cashout_stopped_by

    if stopping:
        _cashout_stopped = True
        _cashout_stopped_by = who

    stamp = datetime.now(LOCAL_TZ).strftime('%I:%M %p - %d %b %Y')
    if stopping:
        body = (f"{CASHOUT_STOP_MARK}\n\n"
                "Cashout requests are NOT being forwarded and nothing is being "
                "chased. Nothing has been posted in any group.\n\n"
                f"Stopped at {stamp}"
                f"{f' by {who}' if who else ''}.\n\n"
                "Send /cashout on to start it again.")
    else:
        body = (f"{CASHOUT_RESUME_MARK}\n\n"
                "Cashout requests are being forwarded again as normal.\n\n"
                f"Resumed at {stamp}"
                f"{f' by {who}' if who else ''}.")

    # Private only. Nothing about the switch is ever posted in a group - not the
    # stop, not the resume. This message is also the durable record, which is
    # why it is sent even when nobody needs telling.
    failed = await dm_handles(CASHOUT_ADMIN_HANDLES, body)
    await warn_unreachable(failed)

    if not stopping:
        # The record exists. Only now does the flow go live again.
        _cashout_stopped = False
        _cashout_stopped_by = None

    print(f"{'⏸️' if stopping else '▶️'} [SWITCH] cashout forwarding "
          f"{'STOPPED' if stopping else 'RESUMED'}"
          f"{f' by {who}' if who else ''}"
          f"{f' - could not reach {failed}' if failed else ''}", flush=True)
    return failed


async def _note_while_stopped(chat_id, text, user_id, username):
    """Report anything real that the stop swallowed.

    The switch must not become a silent hole. A cashout request that arrived
    while the flow was out of service, or a /out answering one, is real money
    that nobody would otherwise hear about - which is the failure this whole bot
    exists to prevent. Ordinary chatter in those groups is not reported."""
    text = text or ''
    source = _canonical(chat_id, CASHOUT_ROUTES)
    if source is not None and CASHOUT_KEYWORD.lower() in text.lower():
        what = f"A CASHOUT REQUEST arrived in {chat_name(source)}"
    elif _canonical(chat_id, _CASHOUT_HANDLERS) is not None and _OUT_CMD_RE.search(text):
        what = f"A /out was sent in {chat_name(chat_id)}"
    else:
        return                              # not something the flow would act on

    print(f"⏸️ [SWITCH] ignored while stopped: {what}", flush=True)
    await warn_unreachable(await dm_handles(
        CASHOUT_ADMIN_HANDLES,
        f"⏸️ {what}, but cashout forwarding is STOPPED.\n\n"
        f"{' '.join(text.split())[:300]}\n\n"
        "It has NOT been forwarded, booked or chased, and nobody else has been "
        "told. Handle it by hand, or send /cashout on to start the flow again."))


def is_caption_out(text):
    """Does this media caption carry a /out?

    Both input paths drop captioned media before anything else, and that is
    right for the payment side: a notification is always plain text, so reading
    a screenshot as one could only ever invent a payment that never happened.

    The single exception is a /out. The crew habitually reply to a request with
    the Cash App screenshot proving they sent it and write the /out as its
    caption, and that message is the answer to a request the bot is still
    chasing - dropping it leaves the request open, the ❤ unplaced, the
    reminders running and the money unbooked.

    Only /out passes. A captioned CASHOUT REQUEST still opens nothing, and the
    caption never reaches process_incoming."""
    return bool(text) and bool(_OUT_CMD_RE.search(text))


def chat_name(chat_id):
    return CHAT_NAMES.get(chat_id, str(chat_id))


def _canonical(chat_id, table):
    """Match a chat id against a config table under either id spelling."""
    if chat_id in table:
        return chat_id
    for variant in _id_variants(chat_id):
        if variant in table:
            return variant
    return None


def remember_user(sender):
    """Learn @username -> id from anyone who speaks in a watched chat.

    This is what makes the escalation DM reachable by the bot: the crew are
    configured by username, but the Bot API can only send to an id."""
    username = (getattr(sender, 'username', '') or '').lower()
    user_id = getattr(sender, 'id', None)
    if username and user_id and _user_ids.get(username) != user_id:
        _user_ids[username] = user_id
        print(f"👤 [CASHOUT] learned @{username} = {user_id}", flush=True)


def group_crew(handling):
    """The crew who work this handling group only - [] for anywhere else."""
    if handling is None:
        return []
    key = _canonical(handling, CASHOUT_GROUP_CREW)
    return list(CASHOUT_GROUP_CREW.get(key, [])) if key is not None else []


def crew_handles(handling):
    """Who is DMed about a request waiting in this group: everyone's crew,
    then this group's own. Order matters only in that it keeps the existing
    three where they have always been."""
    known = {h.lower() for h in CASHOUT_CREW_HANDLES}
    return list(CASHOUT_CREW_HANDLES) + [h for h in group_crew(handling)
                                         if h.lower() not in known]


def crew_mentions(handling):
    """The tag line posted into this group, on the request and every reminder.

    Falls back to exactly CASHOUT_MENTIONS where a group has nobody of its
    own, so the groups that had no group crew read byte for byte as before."""
    tagged = {m.lower().lstrip('@') for m in CASHOUT_MENTIONS.split()}
    extra = ' '.join('@' + h for h in group_crew(handling)
                     if h.lower() not in tagged)
    return f"{CASHOUT_MENTIONS} {extra}".strip() if extra else CASHOUT_MENTIONS


def _is_responder(user_id, username, handling=None):
    """Whose engagement counts as the CREW picking a request up.

    Ethan and Larry deliberately do NOT count, though they did until
    2026-08-09 (`user_id in LEDGER_ADMINS`). Acknowledgement now stops the
    group chase outright, so an admin reacting would silence the very
    reminders meant to reach the crew - and an admin looking at a request is
    not the crew acting on it.

    `handling` says WHERE they spoke, which is what makes a one-group crew
    member count there and nowhere else. It defaults to None so a caller with
    no group in hand still gets the everywhere-crew answer."""
    name = (username or '').lower()
    if name in CASHOUT_RESPONDERS:
        return True
    return name in {h.lower() for h in group_crew(handling)}


# A Telegram @username: 5-32 characters, starting with a letter. The lookbehind
# is what keeps "bob@larryyxx" from matching, exactly as the mention watch does.
_HANDLE_RE = re.compile(r'(?<![\w@])@[A-Za-z][A-Za-z0-9_]{3,31}')


def _crew_identifiers():
    """Every handle the config knows the crew by, lowercased and without the @.

    Read from the live config rather than hardcoded, so adding somebody to
    CASHOUT_CREW_HANDLES, CASHOUT_MENTIONS or CASHOUT_GROUP_CREW covers this
    too and there is no second list to forget.

    EVERY group's crew, not just one group's. Redaction is about what may
    leave, and a /out written in one handling group is relayed into a chime
    group - a name is a name wherever it was typed."""
    words = set()
    for handle in list(CASHOUT_CREW_HANDLES) + list(CASHOUT_RESPONDERS):
        words.add(str(handle).lower().lstrip('@'))
    for handle in CASHOUT_MENTIONS.split():
        words.add(handle.strip().lower().lstrip('@'))
    for handles in CASHOUT_GROUP_CREW.values():
        for handle in handles:
            words.add(str(handle).lower().lstrip('@'))
    return {w for w in words if w}


def strip_identities(text, username=None, full_name=None):
    """Take every crew handle and name out of something bound for a chime group.

    The /out is relayed verbatim, which is right for the figure and wrong for
    the people: a crew member typing "sent by @Maynuddin23" on their own
    screenshot would put that straight into the group that asked. The chime
    groups are never told who handles their cashouts.

    ANY @mention goes, not only the ones in the config - somebody new on the
    crew is exactly the case a fixed list would miss. Bare names go too, since
    a handle written without its @ is still a name.

    Only what is RELAYED is cleaned. book_cashout_out() reads the original, so
    the amount can never be redacted out from under the ledger."""
    if not text:
        return text

    cleaned = _HANDLE_RE.sub('', text)

    names = set(_crew_identifiers())
    if username:
        names.add(str(username).lower().lstrip('@'))
    # Longest first, so "maynuddin233" is not left as a stray "3" by "maynuddin23".
    phrases = sorted(names, key=len, reverse=True)
    if full_name and full_name.strip():
        phrases.insert(0, full_name.strip().lower())

    for phrase in phrases:
        if phrase:
            cleaned = re.sub(r'(?<!\w)' + re.escape(phrase) + r'(?!\w)', '',
                             cleaned, flags=re.I)

    cleaned = re.sub(r'[ \t]{2,}', ' ', cleaned)
    cleaned = re.sub(r'[ \t]+(\r?\n)', r'\1', cleaned)
    cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
    cleaned = cleaned.strip()

    # Redacting everything away would leave nothing to send, and an empty
    # message is a failed send - which leaves the request open and the cashout
    # unanswered. The instruction itself is what has to survive.
    return cleaned or '/out'


# A cashtag, never an amount: $jenny-buhr and $Hawkins-Floral-Decor match,
# $500 does not, because the character after the $ must be a letter.
_CASHTAG_RE = re.compile(r'\$[A-Za-z][\w.\-]*')


def clean_out_for_relay(text, username=None, full_name=None):
    """What the chime group is shown when the /out named somebody.

    A /out with nothing to hide is relayed **verbatim**, exactly as it always
    has been - the crew's own wording is theirs and the group that asked is
    used to seeing it.

    When it does name somebody, subtracting the words leaves a sentence with
    holes in it - "/out 25 - sent by" - which reads like a bug and still hints
    that a name was removed. So the message is rebuilt instead, from the only
    two things the group that asked actually needs: the figure and the cashtag.
    The screenshot still travels; only the words around it are replaced."""
    cleaned = strip_identities(text, username, full_name)
    if cleaned == text:
        return text                     # nothing was hidden - leave it alone

    amount = _OUT_AMOUNT_RE.search(text or '')
    cashtag = _CASHTAG_RE.search(cleaned)

    rebuilt = f"/out {amount.group(1)}" if amount else '/out'
    if cashtag:
        rebuilt += f"\n{cashtag.group(0)}"
    print(f"🧾 [CASHOUT] the /out named somebody - relaying it as "
          f"{rebuilt.splitlines()[0]!r} only", flush=True)
    return rebuilt


async def send_group(chat_id, text, **kwargs):
    """The one door out. Everything the bot says goes through here.

    A crew handle must never reach a TARGET group - CHIME PICCASO, CHIME
    GAFFER or either VENMO group. Those groups are told figures and never who
    moved them, and that has to hold for every message, not only the /out
    relay: a forwarded payment, a milestone, a correction, anything added
    later. Guarding the door rather than each caller is what makes "never"
    true, including for code that does not exist yet.

    Everywhere else is a pass-through. It has to be: the handling groups are
    tagged with those exact handles on every request, which is the whole point
    of CASHOUT_MENTIONS, and the DMs to Ethan and Larry deliberately name the
    crew member who is stuck."""
    # A paused group hears nothing at all. The feature gates upstream should
    # mean nothing ever reaches this line, but "should" is not what a live
    # group deserves: one check at the only door out is what makes the silence
    # true for code that does not exist yet, exactly as the redaction below
    # makes "no crew name" true.
    if chat_paused(chat_id):
        print(f"⏸️ [PAUSED] not posting into {chat_name(chat_id)}: "
              f"{' '.join((text or '').split())[:80]}", flush=True)
        return None

    if chat_id in REDACTED_CHATS:
        text = strip_identities(text)
    return await bot.send_message(chat_id, text, **kwargs)


async def dm_handles(handles, text, receipts=None):
    """Private warning to each of these handles. Returns those it could not reach.

    Bot first, because a message from the bot is the one that fits the rest of
    the system. The user account is the fallback for anyone the bot cannot reach
    - see USERBOT_SEND.

    Pass `receipts` a list to collect where each copy landed, which is what makes
    a message retractable later - see delete_crew_notice(). Which client sent it
    travels with the id, because an id only means anything alongside the chat it
    was sent through."""
    unreachable = []
    for handle in handles:
        key = handle.lower()
        sent = False

        user_id = _user_ids.get(key)
        if user_id:
            try:
                delivered = await send_group(user_id, text)
                sent = True
                if receipts is not None and getattr(delivered, 'message_id', None):
                    receipts.append(('bot', user_id, delivered.message_id))
                print(f"✉️ [CASHOUT] DM sent to @{handle} via bot", flush=True)
            except Exception as e:
                # Almost always "bot can't initiate conversation with a user".
                print(f"⚠️ [CASHOUT] bot DM to @{handle} failed: {e}", flush=True)

        if not sent and USERBOT_SEND and _active_client is not None:
            try:
                delivered = await _active_client.send_message(handle, text)
                sent = True
                if receipts is not None and getattr(delivered, 'id', None):
                    receipts.append(('user', handle, delivered.id))
                # Messaging the logged-in account itself is allowed on purpose.
                # Telegram files it under Saved Messages, which that person does
                # see - and if the session belongs to someone ON this list, the
                # alternative is the person most likely to act on it being the
                # only one who never hears about it.
                where = ' (Saved Messages)' if key == _userbot_username else ''
                print(f"✉️ [CASHOUT] DM sent to @{handle} via user account{where}",
                      flush=True)
            except Exception as e:
                print(f"⚠️ [CASHOUT] user-account DM to @{handle} failed: {e}", flush=True)

        if not sent:
            unreachable.append(handle)

    return unreachable


async def warn_unreachable(unreachable):
    """One alert covering everyone who heard nothing, however many rounds it took."""
    if not unreachable:
        return
    names = ', '.join('@' + h for h in unreachable)
    print(f"❌ [CASHOUT] nobody could be DMed at: {names}", flush=True)
    await notify_admin(
        f"⚠️ Could not privately warn {names} about an overdue cashout.\n\n"
        "Ask them to open the bot and press Start, which lets the bot DM "
        "them directly. Until then only the group mention reaches them.")


# Telegram's various ways of saying "that message is not there any more".
_GONE_MARKERS = ('message to react not found', 'message_id_invalid',
                 'message not found', 'message to reply not found',
                 'message to delete not found')


def _looks_deleted(error):
    text = str(error).lower()
    return any(marker in text for marker in _GONE_MARKERS)


async def heart_request(chat_id, message_id):
    """❤ a cashout request, marking it actioned.

    Falls back to the user account because the bot may not be able to react on
    someone else's message in that chat, and the account already can. A message
    that has been deleted is not an error and gets no retry - there is nothing
    left to mark.

    Both routes failing is NOT a cosmetic miss. The ❤ is the durable record that
    a request was actioned - the only one, since the open requests themselves
    live in memory and a redeploy wipes them. A cashout that was paid but never
    marked reads as still outstanding to anyone scrolling the group. So when
    neither route can place it, Ethan is told, with the error, rather than it
    ending as one line in a log nobody is reading."""
    bot_error = user_error = None
    try:
        await bot.set_message_reaction(chat_id, message_id,
                                       [ReactionTypeEmoji(CASHOUT_HEART)])
        print(f"❤️ [CASHOUT] hearted {message_id} in {chat_name(chat_id)}", flush=True)
        return True
    except Exception as e:
        if _looks_deleted(e):
            print(f"🗑️ [CASHOUT] request {message_id} in {chat_name(chat_id)} is "
                  "gone - nothing to heart, leaving it.", flush=True)
            return False
        bot_error = e
        print(f"⚠️ [CASHOUT] bot could not react in {chat_name(chat_id)}: {e}",
              flush=True)

    if USERBOT_REACT and _active_client is not None:
        try:
            entity = await _resolve_one(_active_client, chat_id)
            if entity is not None:
                await _active_client(SendReactionRequest(
                    peer=entity, msg_id=message_id,
                    reaction=[ReactionEmoji(emoticon=CASHOUT_HEART)]))
                print(f"❤️ [CASHOUT] hearted {message_id} in {chat_name(chat_id)} "
                      "via user account", flush=True)
                return True
            user_error = 'the user account could not resolve that chat'
        except Exception as e:
            user_error = e
            print(f"❌ [CASHOUT] user-account reaction failed in "
                  f"{chat_name(chat_id)}: {e}", flush=True)
    else:
        user_error = ('no user account available - USERBOT_REACT is off'
                      if not USERBOT_REACT else
                      'the user account is not connected yet')

    await notify_admin(
        f"⚠️ A cashout in {chat_name(chat_id)} was completed but could NOT be "
        f"marked with the {CASHOUT_HEART}.\n\n"
        f"Request message: {message_id}\n"
        f"Bot: {bot_error}\n"
        f"User account: {user_error}\n\n"
        "The money side is done - the /out was sent and the totals are booked. "
        "Only the mark is missing, so that request will look outstanding in the "
        "group. Placing it by hand settles it.\n\n"
        "If this keeps happening the bot is most likely not an administrator in "
        "that group, or the group restricts which reactions may be used.")
    return False


async def open_cashout_request(source, text, sent_at, origin_msg_id):
    """Post a cashout request into its handling group and start the clock.

    `origin_msg_id` is the request as it sits in the chime group. That is the
    message that gets the ❤ once the cashout is actioned, so the group that
    asked can see at a glance which of its requests have been dealt with."""
    route = CASHOUT_ROUTES[source]
    handling = route['handling']

    # _is_duplicate keys off the message id, which catches the same message
    # reaching both input paths. It cannot catch the same REQUEST arriving as
    # two different messages - a double send, or two copies carrying different
    # ids - and that lands as two identical posts in the handling group, two
    # sets of reminders and two things to answer for one cashout.
    fingerprint = ' '.join((text or '').split()).lower()
    now = datetime.now(timezone.utc)
    if CASHOUT_DEDUP_SECONDS:
        for open_request in _pending_cashouts.get(handling, []):
            if open_request.get('fingerprint') != fingerprint:
                continue
            age = (now - open_request['opened']).total_seconds()
            if age <= CASHOUT_DEDUP_SECONDS:
                print(f"🔁 [CASHOUT] identical request already open in "
                      f"{chat_name(handling)} ({age:.0f}s ago) - not posting a "
                      "second copy", flush=True)
                return

    # The check above can only see requests that finished posting. This covers
    # the one still in flight - see _cashout_claims.
    if not _claim_cashout(handling, fingerprint, now):
        print(f"🔁 [CASHOUT] identical request already being posted to "
              f"{chat_name(handling)} - not posting a second copy", flush=True)
        return

    body = f"{correct_timestamp(text, sent_at)}\n\n{crew_mentions(handling)}"
    try:
        sent = await send_group(handling, body)
    except Exception as e:
        _release_cashout(handling, fingerprint)
        print(f"❌ [CASHOUT] could not post {chat_name(source)} request to "
              f"{chat_name(handling)}: {e}", flush=True)
        await notify_admin(
            f"🚨 A CASHOUT REQUEST from {chat_name(source)} could not be posted to "
            f"{chat_name(handling)}:\n{e}\n\n"
            "Nobody has been told about it. Check the bot is still a member there "
            "and allowed to post.")
        return

    if sent is None:
        # The door refused it - the handling group is out of service. The gate
        # in observe_cashout() means this should be unreachable; it is handled
        # rather than left to raise, because a request record with no message
        # id behind it would strand the claim and take the listener down with
        # it. Silent, like everything else about a paused group.
        _release_cashout(handling, fingerprint)
        return

    request = {
        'origin': source,
        'fingerprint': fingerprint,        # what the duplicate guard compares
        # Resolved once, at open time, so re-pointing a route in Railway cannot
        # strand a request that is already in flight.
        'out_to': route['out_to'],
        'text': text,
        'origin_msg_id': origin_msg_id,    # in the chime group - gets the ❤
        'message_id': sent.message_id,     # in the handling group - reminders reply to it
        'opened': now,                     # the ladder is measured from here
        # Moved by a reply OR a reaction. Records that somebody engaged; it no
        # longer moves the ladder, which runs on absolute time from 'opened'.
        'last_seen': now,
        'seen': False,                     # somebody has acknowledged it
        'seen_at': None,                   # ...and when, which picks Larry's mark
        'seen_by': None,                   # ...and who, for his notice
        'progress_told': False,            # Larry has been told it was picked up
        'larry_told': False,               # ...and that it is STILL not processed
        # Where the crew's 17-minute chase landed, so it can be taken back once
        # the /out arrives - see delete_crew_notice().
        'crew_notice': [],
        # ...and the same for the loud half of the ladder: every group reminder
        # posted for THIS request, so they go with it - see
        # delete_group_notice(). Per request, not per group: two cashouts open
        # in one handling group each have their own, and paying one must not
        # clear the reminders still chasing the other.
        'group_notice': [],
        # Each private escalation fires at most ONCE per request. Repeating them
        # every round is what made the chase unreadable, so the group post is
        # the recurring part and a DM is a one-off "this one needs you".
        'crew_told': False,
        'admin_told': False,
        'issue_told': False,               # answered with something that is not a /out
        'issue_due': None,                 # ...and when to say so, if the edit never comes
        'issue_who': None,                 # ...and who it was
        'mismatch_told': False,            # paid a different figure than it asked for
        'nudges': 0,
        'exhausted': False,
    }
    _pending_cashouts.setdefault(handling, []).append(request)
    # Handover. The request is now on the books, and the check at the top of
    # this function can see it, so the claim has done its job. Releasing here
    # rather than on a clock of its own keeps _pending_cashouts the single
    # answer to "is this a duplicate" - two windows that could disagree is how
    # the guard would start refusing genuine second cashouts.
    _release_cashout(handling, fingerprint)
    print(f"📤 [CASHOUT] {chat_name(source)} -> {chat_name(handling)} "
          f"(msg {sent.message_id}), waiting {CASHOUT_TIMEOUT_MINUTES}m for /out",
          flush=True)

    # Larry follows a cashout from here: submitted, picked up, completed.
    missed = await dm_handles(CASHOUT_PROGRESS_HANDLES,
                              cashout_submitted_text(request, handling))
    await warn_unreachable(missed)


async def book_cashout_out(origin, text):
    """Move the chime group's Total Out by the figure in the /out reply.

    The forwarded /out is posted BY the bot, so the ordinary /out handler never
    sees it - that one is reachable only by Ethan and Larry, and stays that way.
    The money still has to leave the books, so it is booked here instead, in
    exactly the shape /out itself posts.

    That shape is not cosmetic. recover_ledgers() rebuilds the books after a
    deploy by reading back the newest bot message carrying BOTH totals, so
    booking silently - without posting the figures - would let the next redeploy
    roll this cashout straight back off the books."""
    match = _OUT_AMOUNT_RE.search(text)
    amount = _to_float(match.group(1)) if match else None
    if amount is None or amount <= 0:
        # A bare "/out" with no figure. The instruction still gets forwarded and
        # hearted; there is simply nothing to book.
        print(f"📒 [CASHOUT] no amount in the /out - {chat_name(origin)} totals "
              "left alone", flush=True)
        return None

    before = ledger_snapshot(origin)
    after = (before[0], before[1] + amount)
    try:
        await send_group(origin,
                               f"📤 Out = -{amount:,.2f}$\n\n"
                               f"📊 Group Total:\n"
                               f"➕ Total In : {after[0]:,.2f}$\n"
                               f"➖ Total Out: {after[1]:,.2f}$")
    except Exception as e:
        # Nothing committed - the same rule the rest of the ledger keeps: the
        # books never move ahead of what the group can actually see.
        print(f"❌ [CASHOUT] Total Out NOT moved in {chat_name(origin)}, "
              f"send failed: {e}", flush=True)
        return None

    ledger_commit(origin, after, f" (cashout /out {amount:,.2f})")
    await check_milestones(origin, before, after)
    return amount


def describe_user(username, full_name):
    """Both the @handle and the display name, whichever of them Telegram gave us."""
    if username and full_name:
        return f"@{username} ({full_name})"
    if username:
        return f"@{username}"
    if full_name:
        return full_name
    return "someone tagged on the request"


def cashout_submitted_text(request, handling):
    return (f"📨 CASHOUT REQUEST HAS BEEN SUBMITTED\n\n"
            f"{_cashout_preview(request)}\n\n"
            f"From: {chat_name(request['origin'])}\n"
            f"Sent to: {chat_name(handling)}, crew tagged.\n\n"
            "Waiting on a reaction, then the /out.")


def cashout_issue_text(request, handling, who, user_id):
    """For Ethan and Larry: a request was taken on, then answered with anything
    except the /out. Names WHO, because the next step is asking them."""
    return ("⚠️ SOMETHING IS WRONG WITH A CASHOUT\n\n"
            f"{_cashout_preview(request)}\n\n"
            f"From: {chat_name(request['origin'])}\n"
            f"Waiting in: {chat_name(handling)}\n\n"
            f"{who} acknowledged it, then sent a message or a screenshot "
            "instead of the /out.\n"
            f"Their id: {user_id}\n\n"
            "Please check the group's messages.")


async def flag_cashout_issue(handling, request, user_id, username, full_name=None):
    """A crew member who already acknowledged a request posts something that is
    not a /out.

    Different in kind from silence: somebody is engaged with this request and
    still not sending the /out, which usually means they have hit a problem.
    Ethan and Larry hear about it; the crew are NOT told - they are the ones
    being asked about, and they already have the request in front of them.

    This ARMS the alert rather than sending it. The crew routinely post the
    message first and edit the /out onto it seconds later, so firing on the
    first version turns the ordinary way they work into an alert every time.
    Holding it for CASHOUT_ISSUE_DELAY_SECONDS lets the edit land first - and
    an edit that carries a /out settles the request, which drops it from the
    queue and means this never fires at all.

    Armed once per request. A second message during the hold does NOT push the
    deadline out, or somebody typing steadily would defer it forever."""
    if request['issue_told'] or request['issue_due']:
        return
    who = describe_user(username, full_name)
    request['issue_who'] = (who, user_id)
    if CASHOUT_ISSUE_DELAY_SECONDS <= 0:
        await send_cashout_issue(handling, request)
        return
    request['issue_due'] = (datetime.now(timezone.utc)
                            + timedelta(seconds=CASHOUT_ISSUE_DELAY_SECONDS))
    print(f"⏳ [CASHOUT] {who} answered an acknowledged request in "
          f"{chat_name(handling)} without a /out - holding "
          f"{CASHOUT_ISSUE_DELAY_SECONDS}s for an edit", flush=True)


async def send_cashout_issue(handling, request):
    """Send the held alert. Once per request, like every other one here."""
    if request['issue_told']:
        return
    request['issue_told'] = True
    request['issue_due'] = None
    who, user_id = request['issue_who']
    print(f"⚠️ [CASHOUT] no /out edit arrived in {chat_name(handling)} - "
          f"telling Ethan and Larry about {who}", flush=True)
    missed = await dm_handles(CASHOUT_ADMIN_HANDLES,
                              cashout_issue_text(request, handling, who, user_id))
    await warn_unreachable(missed)


async def flush_pending_issue(handling, request):
    """Send a held alert once its hold is up.

    Runs on every watchdog pass, BEFORE the exhausted guard - the chase being
    over does not mean a problem raised during it should be swallowed. A /out
    arriving in the meantime removes the request from the queue entirely, so
    this is never reached for one that got answered."""
    if not request['issue_due'] or request['issue_told']:
        return
    if datetime.now(timezone.utc) < request['issue_due']:
        return
    await send_cashout_issue(handling, request)


def cashout_picked_up_text(request, who):
    return (f"👀 The cashout process is being handled by {who}.\n\n"
            f"{_cashout_preview(request)}\n\n"
            "Waiting on the /out now.")


def cashout_completed_text(request, out_to, amount):
    figure = f" OF {amount:,.2f}$" if amount else ""
    return (f"✅ CASHOUT{figure} TO {chat_name(out_to)} COMPLETED SUCCESSFULLY\n\n"
            f"{_cashout_preview(request)}\n\n"
            f"Please send the required screenshot to {chat_name(out_to)}.")


async def note_cashout_seen(chat_id, message_id, user_id, username, full_name=None):
    """A tagged responder reacted to the forwarded request.

    A reaction is an acknowledgement, so it carries exactly the standing a reply
    does: the handling group STOPS being chased about it. It does NOT close the
    request - only a real /out does that - so the request stays open and Larry
    is told, once, that it is acknowledged and still not processed.

    Larry is told once, the first time a request is picked up. Reacting is tied
    to one specific message, so unlike a plain reply in the group there is no
    doubt about WHICH request has been taken on."""
    handling = _canonical(chat_id, _CASHOUT_HANDLERS)
    if handling is None:
        return False
    if not _is_responder(user_id, username, handling):
        return False               # a reaction from anyone else means nothing
    for request in _pending_cashouts.get(handling, []):
        if request['message_id'] != message_id:
            continue
        request['last_seen'] = datetime.now(timezone.utc)
        _mark_acknowledged(request, describe_user(username, full_name))
        print(f"👀 [CASHOUT] @{username or user_id} reacted in {chat_name(handling)} "
              "- group chasing stops, Larry gets the word instead", flush=True)

        if not request['progress_told']:
            request['progress_told'] = True     # once per request, not per reaction
            missed = await dm_handles(
                CASHOUT_PROGRESS_HANDLES,
                cashout_picked_up_text(request, describe_user(username, full_name)))
            await warn_unreachable(missed)
        return True
    return False


# The figure a cashout request is ASKING for. "Amount : 25" is the shape the
# notification bot uses; the loose "$25" is the fallback for one typed by hand.
# Neither can match a cashtag, because a digit must follow the $.
_REQUEST_AMOUNT_RE = re.compile(r'amount\s*[:\-]?\s*\$?\s*([\d,]+(?:\.\d+)?)', re.I)
_LOOSE_AMOUNT_RE = re.compile(r'\$\s*([\d,]+(?:\.\d+)?)')


def request_amount(text):
    """What a cashout request is asking for, or None if it cannot be read."""
    match = _REQUEST_AMOUNT_RE.search(text or '') or _LOOSE_AMOUNT_RE.search(text or '')
    return _to_float(match.group(1)) if match else None


def out_amount(text):
    """What a /out actually paid, or None if it cannot be read."""
    paid = _OUT_AMOUNT_RE.search(text or '')
    return _to_float(paid.group(1)) if paid else None


def cashout_shortfall_text(request, handling, asked, paid):
    """For Larry: the figure that went is not the figure that was asked for."""
    gap = asked - paid
    word = 'SHORT BY' if gap > 0 else 'OVER BY'
    return (f"⚠️ CASHOUT PAID {word} {abs(gap):,.2f}$\n\n"
            f"From: {chat_name(request['origin'])}\n"
            f"Paid in: {chat_name(handling)}\n\n"
            f"{_cashout_preview(request)}\n\n"
            f"Asked for: {asked:,.2f}$\n"
            f"Paid: {paid:,.2f}$\n\n"
            f"The /out has been relayed and {paid:,.2f}$ booked to "
            f"{chat_name(request.get('out_to', request['origin']))} — the books "
            "follow what was actually sent, not what was asked for.")


async def flag_amount_mismatch(handling, request, text):
    """The /out does not pay what the request asked for.

    Immediate, not on any timer: an underpayment is only cheap to fix while
    somebody still remembers the cashout, and nothing else in the flow would
    ever mention it. The request is settled and hearted either way - the money
    HAS moved, and pretending otherwise would strand a real payment.

    Larry only, and once. Reading the figure off the SCREENSHOT is not possible
    here; this compares what was asked with what the /out says was paid, which
    is the same error caught one step earlier."""
    asked = request_amount(request.get('text'))
    paid = out_amount(text)
    if asked is None or paid is None or asked == paid:
        return                      # nothing to compare, or nothing wrong
    if request.get('mismatch_told'):
        return
    request['mismatch_told'] = True
    print(f"⚠️ [CASHOUT] {chat_name(handling)} paid {paid:,.2f} against a request "
          f"for {asked:,.2f} - telling Larry", flush=True)
    await warn_unreachable(await dm_handles(
        CASHOUT_PROGRESS_HANDLES,
        cashout_shortfall_text(request, handling, asked, paid)))


def _match_request(queue, reply_to, text=None):
    """Which open request a /out is answering.

    An explicit reply is unambiguous, so it wins.

    Otherwise the AMOUNT decides. Taking the oldest request blindly was wrong in
    the one case that matters: with two requests open, a /out paying the second
    settled and hearted the FIRST. The person who asked was left with an
    unmarked request still being chased, and somebody else's - for a different
    figure entirely - was marked done without being paid. Seen live in
    Chime Rev & out no-7 on 2026-08-04.

    The oldest is still the fallback, for a /out with no figure or a figure
    matching nothing, because settling the wrong request is recoverable and
    dropping a real /out strands a cashout. Among several asking for the same
    amount the oldest wins, which is the order they are worked in."""
    if reply_to:
        for request in queue:
            if request['message_id'] == reply_to:
                return request

    paid = _OUT_AMOUNT_RE.search(text or '')
    if paid:
        wanted = _to_float(paid.group(1))
        if wanted is not None:
            for request in queue:
                if request_amount(request.get('text')) == wanted:
                    return request

    return queue[0]


# ---------------------------------------------------------------------------
# Albums
#
# The crew habitually answer with SEVERAL screenshots at once - the payment and
# the history behind it - and Telegram sends an album as separate messages that
# merely share a media_group_id. Only ONE of them carries the caption, so
# copying "the message the /out was on" relayed one picture and silently
# dropped the rest. Seen live in Chime Rev on 2026-08-09: two screenshots sent,
# one arrived in CHIME GAFFER.
#
# The parts arrive either side of the captioned one, so both directions have to
# work: siblings already seen are flushed when the /out lands, and siblings
# arriving afterwards are copied on sight.

_album_parts = OrderedDict()          # (chat, group_id) -> {'ids', 'to', 'sent'}
ALBUM_MEMORY = 40                     # albums remembered at once
ALBUM_RELAY_SECONDS = int(os.getenv('ALBUM_RELAY_SECONDS', '180'))


def _album_slot(handling, group_id):
    key = (handling, str(group_id))
    slot = _album_parts.get(key)
    if slot is None:
        slot = {'ids': [], 'to': None, 'sent': set(),
                'at': datetime.now(timezone.utc)}
        _album_parts[key] = slot
        while len(_album_parts) > ALBUM_MEMORY:
            _album_parts.popitem(last=False)
    return slot


async def _copy_album_part(handling, out_to, message_id):
    """One extra picture, on its own, carrying no words.

    caption='' is deliberate rather than left alone: Telegram allows only one
    caption per album so a sibling normally has none, but copy_message keeps
    whatever is there, and anything a crew member typed could carry a name into
    a group that must never see one."""
    try:
        await bot.copy_message(out_to, handling, message_id, caption='')
        print(f"🖼️ [CASHOUT] extra screenshot copied to {chat_name(out_to)}",
              flush=True)
        return True
    except Exception as e:
        print(f"⚠️ [CASHOUT] could not copy an extra screenshot to "
              f"{chat_name(out_to)}: {e}", flush=True)
        return False


async def note_album_part(handling, group_id, message_id):
    """A picture that belongs to an album, seen in a handling group.

    If the /out has already been dealt with, this one is late and goes across
    immediately. Otherwise it is remembered until the /out arrives - which is
    the ordinary case, since the caption sits on the first of the album."""
    if not group_id or message_id is None:
        return
    slot = _album_slot(handling, group_id)
    if message_id not in slot['ids']:
        slot['ids'].append(message_id)

    if slot['to'] is None or message_id in slot['sent']:
        return
    if (datetime.now(timezone.utc) - slot['at']).total_seconds() > ALBUM_RELAY_SECONDS:
        return
    slot['sent'].add(message_id)
    await _copy_album_part(handling, slot['to'], message_id)


async def relay_album_siblings(handling, group_id, out_to, captioned_id):
    """The /out has gone across. Send the rest of its album after it.

    Registered rather than done once and forgotten: the parts that have not
    arrived yet are the common case, and note_album_part() carries them over as
    they land."""
    if not group_id:
        return
    slot = _album_slot(handling, group_id)
    slot['to'] = out_to
    slot['at'] = datetime.now(timezone.utc)
    slot['sent'].add(captioned_id)

    for message_id in list(slot['ids']):
        if message_id in slot['sent']:
            continue
        slot['sent'].add(message_id)
        await _copy_album_part(handling, out_to, message_id)


async def _deliver_out(out_to, handling, text, username=None, full_name=None,
                       message_id=None, has_media=False, media_group_id=None):
    """Put the /out in front of the group that asked. True if it landed.

    Shared by the ordinary completion and the manual one below, so the two
    cannot drift: the screenshot, the redaction and the fall-back to plain text
    have to behave identically whichever route settles the cashout.

    The screenshot is the proof the money went, and the group that asked is the
    group that wants to see it. copy_message, NEVER forward_message - a forward
    carries a "Forwarded from" header naming the account it came from and links
    back to the handling group, and the chime groups are never told who handles
    their cashouts or that the handling group exists at all."""
    relay = clean_out_for_relay(text, username, full_name)

    if has_media and message_id is not None:
        try:
            await bot.copy_message(out_to, handling, message_id, caption=relay)
            print(f"🖼️ [CASHOUT] screenshot + /out copied to {chat_name(out_to)}",
                  flush=True)
            # One picture is often only part of the proof. Send the rest of the
            # album after it - AFTER the /out has landed, so an extra screenshot
            # failing can never cost the instruction itself.
            await relay_album_siblings(handling, media_group_id, out_to, message_id)
            return True
        except Exception as e:
            # Losing the picture is a nuisance; losing the /out strands the
            # cashout. Fall through and send the instruction on its own.
            print(f"⚠️ [CASHOUT] could not copy the screenshot to "
                  f"{chat_name(out_to)}: {e} - sending the /out by itself",
                  flush=True)

    try:
        await send_group(out_to, relay)
        return True
    except Exception as e:
        print(f"❌ [CASHOUT] /out could not be sent to {chat_name(out_to)}: {e}",
              flush=True)
        return False


async def force_complete_cashout(handling, text, user_id, username,
                                 full_name=None, message_id=None, has_media=False,
                                 media_group_id=None):
    """Ethan or Larry finishing a cashout by hand, with nothing open here.

    The standing rule is that a /out with no request pending is ordinary traffic
    and is left completely alone. That rule is right for the crew and wrong for
    Ethan and Larry, because open requests live in MEMORY: a redeploy wipes
    them, and a redeploy is exactly the moment somebody needs to finish a
    cashout the bot has forgotten. Until now their /out did nothing at all - it
    sat in the group looking actioned to everyone reading it, while the group
    that asked was told nothing and its books never moved.

    Ethan and Larry only. The crew's /out with nothing open stays ordinary
    traffic, which is what keeps this from firing on chatter.

    There is nothing to ❤: the request this answers is not in memory, so its
    message id is not either. The money side is what matters here, and it is
    done exactly as the ordinary completion does it."""
    source = _CASHOUT_HANDLERS.get(handling)
    if source is None:
        return
    out_to = CASHOUT_ROUTES.get(source, {}).get('out_to', source)

    print(f"🔧 [CASHOUT] no request open in {chat_name(handling)} - completing "
          f"by hand for {username or user_id}", flush=True)

    if not await _deliver_out(out_to, handling, text, username, full_name,
                              message_id, has_media, media_group_id):
        return                          # nothing landed, so nothing is booked

    booked = await book_cashout_out(out_to, text)

    await warn_unreachable(await dm_handles(
        CASHOUT_ADMIN_HANDLES,
        f"🔧 A /out was completed by hand in {chat_name(handling)}.\n\n"
        f"Nothing was open there — most likely the request was lost in a "
        f"redeploy — so it was taken at face value.\n\n"
        f"Sent to: {chat_name(out_to)}\n"
        f"Booked: {f'{booked:,.2f}$' if booked else 'nothing (no figure in the /out)'}\n\n"
        "No ❤ was placed: the original request is not in memory, so there is "
        "nothing to mark. Worth adding by hand if it is still there."))


async def handle_cashout_reply(handling, text, user_id, username, reply_to,
                               full_name=None, message_id=None, has_media=False,
                               media_group_id=None):
    """A message in a handling group.

    Ordinary traffic is never touched - including an ordinary /out - unless a
    request we forwarded is genuinely outstanding. The one exception is a /out
    from Ethan or Larry, which completes a cashout even with nothing open; see
    force_complete_cashout().

    `message_id` and `has_media` are what let the screenshot travel with the
    /out - see _deliver_out()."""
    text = text or ''
    queue = _pending_cashouts.get(handling)
    if not queue:
        if user_id in LEDGER_ADMINS and _OUT_CMD_RE.search(text):
            await force_complete_cashout(handling, text, user_id, username,
                                         full_name, message_id, has_media,
                                         media_group_id)
        return
    responder = _is_responder(user_id, username, handling)
    target = _match_request(queue, reply_to, text)

    # Read BEFORE the loop below sets it. The first thing a crew member says is
    # an acknowledgement; a later one, with still no /out, is not.
    was_acknowledged = target['seen']

    if responder:
        # Somebody is on it. The handling group stops being chased about these;
        # they stay open until a /out actually arrives.
        now = datetime.now(timezone.utc)
        who = describe_user(username, full_name)
        for request in queue:
            request['last_seen'] = now
            _mark_acknowledged(request, who)
        print(f"💬 [CASHOUT] @{username or user_id} replied in "
              f"{chat_name(handling)} - group chasing stops", flush=True)

    if not _OUT_CMD_RE.search(text):
        if responder and was_acknowledged:
            await flag_cashout_issue(handling, target, user_id, username, full_name)
        return

    request = target
    # Normally the group that asked, but a route may send the /out elsewhere.
    out_to = request.get('out_to', request['origin'])

    # The screenshot is the proof the money went, and the group that asked is
    # the group that wants to see it.
    #
    # copy_message, NEVER forward_message. A forward carries a "Forwarded from"
    # header naming the account it came from, and links back to the handling
    # group - the chime groups are never told who handles their cashouts, or
    # that the handling group exists at all. A copy arrives as an ordinary
    # message from the bot: the picture, the /out as its caption, and no trace
    # of the sender, their username or the chat it was taken from.
    # Left open deliberately if this fails: the request is not settled until the
    # /out has actually landed somewhere anyone can see it. The ORIGINAL text is
    # kept for book_cashout_out() below, so the redaction inside _deliver_out()
    # can never change the figure that reaches the ledger.
    if not await _deliver_out(out_to, handling, text, username, full_name,
                              message_id, has_media, media_group_id):
        return

    queue.remove(request)
    if not queue:
        _pending_cashouts.pop(handling, None)

    # Booked where the /out was posted, never anywhere else. recover_ledgers()
    # rebuilds a group's books from the totals in that group's own messages, so
    # moving one group's ledger while posting the figures into another would
    # corrupt both.
    booked = await book_cashout_out(out_to, text)

    waited = (datetime.now(timezone.utc) - request['opened']).total_seconds() / 60.0
    print(f"✅ [CASHOUT] /out sent to {chat_name(out_to)} after "
          f"{_humanise(waited)} ({request['nudges']} reminder(s))", flush=True)

    # The ❤ goes on the original request in the chime group, not on the copy in
    # the handling group: it is the group that asked which needs to see, at a
    # glance, which of its requests have been dealt with.
    await heart_request(request['origin'], request['origin_msg_id'])

    # ...and the chase is unsaid, now that it is wrong - privately to the crew,
    # and in the handling group where the same words are still tagging them.
    # After the booking, deliberately: the books moving is what makes it wrong,
    # and deleting first would leave the instruction gone if the booking failed.
    await delete_crew_notice(request)
    await delete_group_notice(handling, request)

    # Paying 150 against a request for 200 settles the request like any other
    # /out - the money really did move - but it is not what was asked for, and
    # nothing else in the flow would ever say so.
    await flag_amount_mismatch(handling, request, text)

    # Last, once the request is marked done in the chime group: Larry is told it
    # completed and which group the screenshot goes to. Sent even if the ❤ could
    # not be applied - the money has moved either way, and he still has to send
    # the proof.
    missed = await dm_handles(CASHOUT_PROGRESS_HANDLES,
                              cashout_completed_text(request, out_to, booked))
    await warn_unreachable(missed)


async def observe_cashout(chat_id, text, message_id, sent_at,
                          user_id=None, username=None, reply_to=None,
                          full_name=None, has_media=False, is_edit=False,
                          media_group_id=None):
    """Single entry point for both input paths.

    Called for every message in a watched chat that the gates upstream let
    through. Decides whether it opens a cashout request, answers one, flags a
    problem with one, or is none of our business.

    `has_media` is what lets a screenshot with NO caption count: posting one
    instead of the /out is itself the signal that something has gone wrong."""
    if not text and not has_media:
        return
    if BOT_ID is not None and user_id == BOT_ID:
        return                              # never act on our own posts

    # A paused group is not the bot's business at all. This sits ABOVE the
    # emergency stop deliberately: the stop reports what it swallows, and a
    # group that has been taken out of service must not be generating those
    # reports. Logged only when it was something the flow would have acted on,
    # so the console does not fill with ordinary chatter.
    if route_paused(chat_id):
        if text and (CASHOUT_KEYWORD.lower() in text.lower()
                     or _OUT_CMD_RE.search(text)):
            print(f"⏸️ [PAUSED] {chat_name(chat_id)} is out of service - that "
                  "cashout traffic was left alone.", flush=True)
        return

    # The emergency stop sits here, above every branch, so ONE check covers
    # opening a request, answering one and flagging a problem with one. Anything
    # real that this swallows is reported rather than dropped.
    if _cashout_stopped:
        await _note_while_stopped(chat_id, text, user_id, username)
        return

    source = _canonical(chat_id, CASHOUT_ROUTES)
    if source is not None:
        if not text or CASHOUT_KEYWORD.lower() not in text.lower():
            return
        if _is_duplicate(('cashout', source, message_id)):
            return
        await open_cashout_request(source, text, sent_at, message_id)
        return

    handling = _canonical(chat_id, _CASHOUT_HANDLERS)
    if handling is not None:
        # An edit is a NEW event for a message we have already seen, so it
        # cannot share the original's key or it would be dropped as a replay -
        # which is exactly how the /out edited onto a screenshot went missing.
        # The edited TEXT is part of the key, so the same edit arriving down
        # both input paths is still one event, while a second, different edit
        # (fixing the figure) gets its turn.
        key = (('cashout-edit', handling, message_id,
                ' '.join((text or '').split())) if is_edit
               else ('cashout-reply', handling, message_id))
        if _is_duplicate(key):
            return

        # A /out is acted on ONCE per message, however often it is edited
        # afterwards. Editing the cashtag on a /out that has already been
        # relayed and booked must not book it twice - and with the request
        # settled and gone, Ethan's or Larry's edit would reach
        # force_complete_cashout(), which takes a /out at face value.
        if _OUT_CMD_RE.search(text or '') and _is_duplicate(
                ('out-done', handling, message_id)):
            return

        # Remembered before anything else: an album part that is not the
        # captioned one carries no /out and would otherwise be dropped here as
        # ordinary media, which is exactly how the extra screenshots went
        # missing.
        if has_media:
            await note_album_part(handling, media_group_id, message_id)

        await handle_cashout_reply(handling, text, user_id, username, reply_to,
                                   full_name, message_id=message_id,
                                   has_media=has_media,
                                   media_group_id=media_group_id)


async def _confirm_gone(client, chat, msg_id):
    """Has that message really been deleted?

    Only needed when Telegram did not say which chat a deletion happened in.
    Anything other than a definite "not there" leaves the request alone."""
    try:
        entity = await _resolve_one(client, chat)
        if entity is None:
            return False
        found = await client.get_messages(entity, ids=[msg_id])
        return not (found and found[0] is not None)
    except Exception as e:
        print(f"⚠️ [CASHOUT] could not confirm deletion in {chat_name(chat)}: {e} "
              "- leaving it open.", flush=True)
        return False


async def close_deleted_cashouts(client, chat_id, deleted_ids):
    """Drop any open request that has been deleted from EITHER end.

    A request lives as two messages: the original in the chime group and the
    copy the bot posted in the handling group. Deleting either one means it has
    been dealt with - withdrawn, or handled and tidied away. Chasing it after
    that is nagging about something nobody can see, and if the handling-group
    copy is the one that went, the reminders cannot even be posted: they reply
    to it. Treat it as settled: out of the queue, no reminder, no reaction.

    Telegram only says WHICH chat a deletion happened in when it was a channel.
    With no chat to go on, a bare message id could belong to any watched chat,
    and dropping the wrong request would silently strand a real cashout - the
    /out answering it would find nothing pending. So in that case the message
    is looked up before anything is dropped."""
    if not deleted_ids:
        return
    gone = set(deleted_ids)

    known = None
    if chat_id is not None:
        known = _canonical(chat_id, CASHOUT_ROUTES) or _canonical(chat_id, _CASHOUT_HANDLERS)
        if known is None:
            return                  # a deletion in some chat we do not care about

    for handling, queue in list(_pending_cashouts.items()):
        for request in list(queue):
            ends = ((request['origin'], request['origin_msg_id'], 'request'),
                    (handling, request['message_id'], 'forwarded copy'))
            for chat, msg_id, label in ends:
                if msg_id not in gone:
                    continue
                if known is not None and chat != known:
                    continue
                if known is None and not await _confirm_gone(client, chat, msg_id):
                    continue
                queue.remove(request)
                print(f"🗑️ [CASHOUT] {label} in {chat_name(chat)} was deleted after "
                      f"{request['nudges']} reminder(s) - treating it as handled. "
                      "No chasing, no reaction.", flush=True)
                break
        if not queue:
            _pending_cashouts.pop(handling, None)


def cashout_nudge_text(waited_minutes, handling=None):
    """Deliberately unsigned.

    This is the one message the bot posts into a handling group, and nothing of
    Ethan's belongs there. The milestone and idle messages ARE signed, because
    those only ever go to the chime groups.

    Sent whether or not anybody has acknowledged the request. A reaction says
    somebody has seen it, not that the money has gone, so it no longer buys
    silence in the group - only a /out stops this."""
    return (f"⏰ OUT REQUEST HAS CROSSED {_humanise(waited_minutes)} TIMEFRAME\n\n"
            "This cashout request is still waiting on a /out.\n\n"
            f"{crew_mentions(handling)}")


def _cashout_preview(request):
    return ' '.join((request['text'] or '').split())[:200]


def cashout_admin_dm_text(request, handling, waited_minutes):
    """For Ethan and Larry: the full picture, routing included."""
    return (f"⏰ OUT REQUEST HAS CROSSED {_humanise(waited_minutes)} TIMEFRAME\n\n"
            f"From: {chat_name(request['origin'])}\n"
            f"Waiting in: {chat_name(handling)}\n\n"
            f"{_cashout_preview(request)}\n\n"
            + ("Acknowledged in the group, but still no /out.\n"
               if request.get('seen') else "Nobody has sent a /out for it yet.\n")
            + f"Tagged in the group: {crew_mentions(handling)}")


def cashout_acknowledged_text(request, handling, waited_minutes):
    """For Larry: they have put their hand up, and the money still has not moved.

    Routing included, like every other message to Ethan and Larry - which groups
    a request travelled between is theirs. It says plainly that the group has
    stopped being chased, because from here nobody else is going to be prompted:
    this notice IS the chase."""
    who = request.get('seen_by') or 'Somebody'
    return (f"👀 ACKNOWLEDGED BUT STILL NOT PROCESSED — {_humanise(waited_minutes)}\n\n"
            f"From: {chat_name(request['origin'])}\n"
            f"Waiting in: {chat_name(handling)}\n\n"
            f"{_cashout_preview(request)}\n\n"
            f"{who} acknowledged it in the group, but no /out has been sent.\n"
            "The group is no longer being reminded about it, and the crew get no "
            "further private chase. It stays open - a late /out still counts.")


def cashout_crew_dm_text(request, waited_minutes):
    """For the crew: the request and what to do about it.

    No group names and no 'waiting in' line - which groups a request moved
    between is not something they need, and the handling group is where they are
    already being tagged anyway."""
    return (f"⏰ OUT REQUEST HAS CROSSED {_humanise(waited_minutes)} TIMEFRAME\n\n"
            f"{_cashout_preview(request)}\n\n"
            "This is still waiting on a /out. Please action it.")


def _ladder_rounds():
    """How many group reminders the ladder allows in total."""
    total = len(CASHOUT_NUDGE_MINUTES)
    return min(total, CASHOUT_MAX_NUDGES) if CASHOUT_MAX_NUDGES else total


def _rungs_due(waited):
    """How many group reminders should have gone out by now."""
    return min(sum(1 for m in CASHOUT_NUDGE_MINUTES if waited >= m),
               _ladder_rounds())


def _mark_acknowledged(request, who=None):
    """Record an acknowledgement, once.

    Only the FIRST one counts. Re-stamping the time on every later reaction or
    message would keep pushing Larry's notice out, which is the recurring chase
    this shape exists to avoid."""
    if request['seen']:
        return
    request['seen'] = True
    request['seen_at'] = datetime.now(timezone.utc)
    request['seen_by'] = who
    # An acknowledgement is new information, so it reopens a chase that had
    # already run out: somebody has just picked up a request everything had
    # given up on, and it is still not paid. chase_acknowledged() fires once
    # and sets this again, so there is no loop.
    request['exhausted'] = False


def _ladder_text():
    """'5, 8 and 12' - the group rungs, for /help and /status.

    Rendered from the live ladder rather than written out, so the help cannot
    drift away from what the bot actually does."""
    rungs = [str(m) for m in CASHOUT_NUDGE_MINUTES[:_ladder_rounds()]]
    if len(rungs) > 1:
        return ', '.join(rungs[:-1]) + ' and ' + rungs[-1]
    return rungs[0] if rungs else ''


async def post_group_nudge(handling, request, waited, round_no):
    """The loud part of the ladder: re-tag the crew in the handling group.

    Ethan and Larry are DMed alongside the FIRST reminder only - chasing is
    their job, so they need it early, and repeating it every round is what
    buried the signal.

    The crew are deliberately NOT DMed here. They are tagged right there in the
    group, and their private word comes only once this ladder has run out - see
    dm_crew_last_resort()."""
    request['nudges'] = round_no
    try:
        # Remembered for the same reason the crew's DM is: a late /out makes
        # this post untrue, and it is taken back rather than left tagging the
        # crew about a cashout that is already paid.
        sent = await send_group(handling, cashout_nudge_text(waited, handling),
                               reply_to_message_id=request['message_id'])
        if getattr(sent, 'message_id', None):
            request['group_notice'].append(sent.message_id)
        print(f"⏰ [CASHOUT] reminder #{round_no} in {chat_name(handling)} "
              f"after {_humanise(waited)}", flush=True)
    except Exception as e:
        print(f"❌ [CASHOUT] reminder failed in {chat_name(handling)}: {e}", flush=True)

    if not request['admin_told']:
        request['admin_told'] = True
        await warn_unreachable(await dm_handles(
            CASHOUT_ADMIN_HANDLES, cashout_admin_dm_text(request, handling, waited)))


async def dm_crew_last_resort(handling, request, waited):
    """The group ladder has run out and there is still no /out.

    This is the ONLY timer that reaches the crew privately, and it fires after
    the group reminders rather than alongside them. Once it has gone the chasing
    stops: the request stays OPEN, so a late /out is still forwarded, booked and
    hearted, and deleting either copy still settles it."""
    request['crew_told'] = True
    print(f"📣 [CASHOUT] nobody acknowledged it after {_humanise(waited)} in "
          f"{chat_name(handling)} - telling the crew privately, once", flush=True)
    # Remember where each copy landed. A late /out makes this message untrue,
    # and it is taken back rather than left sitting in their inbox.
    await warn_unreachable(await dm_handles(crew_handles(handling),
                                            cashout_crew_dm_text(request, waited),
                                            receipts=request['crew_notice']))


async def delete_crew_notice(request):
    """Take back the crew's last-resort chase once the /out has landed.

    That DM says the request is "still waiting on a /out". The moment one
    arrives it is no longer true, and a chase left sitting in their inbox is
    read as work still outstanding - which on this flow means somebody paying a
    second time. Deleting it is the only way to unsay it; Telegram has no edit
    that removes a notification already delivered.

    Only the CREW's copy. Ethan's and Larry's notices are a record of what
    happened rather than an instruction, and they stay put.

    A failure here is not cosmetic - the stale instruction is still live - so
    Ethan is told, with the figure already booked so nobody pays it again."""
    notices, request['crew_notice'] = request['crew_notice'], []
    failed = []
    for kind, where, msg_id in notices:
        try:
            if kind == 'bot':
                await bot.delete_message(where, msg_id)
            elif _active_client is not None:
                await _active_client.delete_messages(where, [msg_id])
            else:
                failed.append(str(where))
                continue
            print(f"🗑️ [CASHOUT] took back the chase DM to {where}", flush=True)
        except Exception as e:
            print(f"⚠️ [CASHOUT] could not take back the chase DM to {where}: {e}",
                  flush=True)
            failed.append(str(where))

    if failed:
        await notify_admin(
            "⚠️ A cashout was completed, but the earlier "
            f"{_humanise(CASHOUT_CREW_DM_MINUTES)} chase DM to the crew could not "
            f"be deleted ({', '.join(failed)}).\n\n"
            "It still reads as waiting on a /out. The money side is already done "
            "and booked — ask them not to send it again.")


async def delete_group_notice(handling, request):
    """Take back the group reminders once the /out has landed.

    Same reasoning as delete_crew_notice(), in the open. Every reminder reads
    "this cashout request is still waiting on a /out" and ends in
    CASHOUT_MENTIONS, so what is left behind is not a stale note but a live tag
    on the crew about a cashout that has already been paid - and on this flow a
    chase somebody acts on twice is somebody paying twice. The request itself
    stays exactly where it is: the forwarded copy, the /out under it and the ❤
    in the chime group are the record, and none of them is touched.

    A reminder that has already been deleted by hand is not a failure. Anything
    else is - the tag is still live - so Ethan is told, with the figure already
    booked so nobody sends it again. The crew are never told: they are the ones
    being chased, and the group is where they can already see the /out."""
    notices, request['group_notice'] = request['group_notice'], []
    failed = []
    for msg_id in notices:
        try:
            await bot.delete_message(handling, msg_id)
            print(f"🗑️ [CASHOUT] took back reminder {msg_id} in "
                  f"{chat_name(handling)}", flush=True)
        except Exception as e:
            if _looks_deleted(e):
                continue
            print(f"⚠️ [CASHOUT] could not take back reminder {msg_id} in "
                  f"{chat_name(handling)}: {e}", flush=True)
            failed.append(str(msg_id))

    if failed:
        await notify_admin(
            f"⚠️ A cashout in {chat_name(handling)} was completed, but "
            f"{len(failed)} earlier reminder(s) could not be deleted from the "
            f"group ({', '.join(failed)}).\n\n"
            "They still tag the crew saying it is waiting on a /out. The money "
            "side is already done and booked — ask them not to send it again.")


async def chase_acknowledged(handling, request, waited):
    """Somebody reacted, or spoke in the handling group.

    The group falls silent from that moment: they have acknowledged it there, so
    re-tagging them is noise. Larry is told once instead, CASHOUT_SEEN_DM_MINUTES
    after the acknowledgement rather than straight away - "still not processed"
    is not worth saying at the instant somebody reacts, and six minutes later it
    is. Measured from the acknowledgement, so reacting late does not shorten it.

    The crew's last-resort DM does NOT fire for an acknowledged request. Once
    they have put their hand up it is Larry's to carry, not theirs to be chased
    over again in private."""
    if request['larry_told']:
        return
    quiet = (datetime.now(timezone.utc) - request['seen_at']).total_seconds() / 60.0
    if quiet < CASHOUT_SEEN_DM_MINUTES:
        return

    request['larry_told'] = True
    request['exhausted'] = True
    print(f"👤 [CASHOUT] acknowledged but still no /out after {_humanise(waited)} in "
          f"{chat_name(handling)} - telling Larry, once. The group is not chased "
          "again; the request stays open.", flush=True)
    await warn_unreachable(await dm_handles(
        CASHOUT_PROGRESS_HANDLES, cashout_acknowledged_text(request, handling, waited)))


async def chase_cashout(handling, request, waited):
    """One pass of the ladder for one open request.

    The rungs are absolute minutes since the request was posted, not gaps
    between reminders, so nothing that happens in between can push them out.

    Two shapes, and which one applies is decided by whether anybody has
    acknowledged it. Silence is chased in the open, on the group ladder;
    an acknowledgement moves it to a single private word to Larry."""
    if request['seen']:
        await chase_acknowledged(handling, request, waited)
        return

    due = _rungs_due(waited)
    if due > request['nudges']:
        await post_group_nudge(handling, request, waited, due)
        if due >= _ladder_rounds():
            print(f"🔕 [CASHOUT] group reminders stopped in {chat_name(handling)} "
                  f"after {due} round(s) ({_humanise(waited)})", flush=True)

    if (CASHOUT_CREW_DM_MINUTES > 0 and not request['crew_told']
            and waited >= CASHOUT_CREW_DM_MINUTES):
        await dm_crew_last_resort(handling, request, waited)

    # Chasing is over only once BOTH channels have run out. The crew DM can
    # fall between two group rungs - at 10 minutes it sits between the 8 and
    # the 12 - and firing it must not cancel the rest of the ladder, nor the
    # ladder finishing cancel it.
    ladder_done = request['nudges'] >= _ladder_rounds()
    crew_done = request['crew_told'] or CASHOUT_CREW_DM_MINUTES <= 0
    if ladder_done and crew_done and not request['exhausted']:
        request['exhausted'] = True
        print(f"🔕 [CASHOUT] {chat_name(handling)} request left alone after "
              f"{_humanise(waited)}. It stays open - a late /out still counts.",
              flush=True)


async def cashout_watchdog():
    """Chase every request that has gone quiet past the timeout."""
    if CASHOUT_TIMEOUT_MINUTES <= 0:
        print("ℹ️ [CASHOUT] escalation disabled (CASHOUT_TIMEOUT_MINUTES=0)", flush=True)
        return

    for source, route in CASHOUT_ROUTES.items():
        line = f"📤 [CASHOUT] {chat_name(source)} -> {chat_name(route['handling'])}"
        if route['out_to'] != source:
            line += f", /out to {chat_name(route['out_to'])}"
        print(line, flush=True)
    ladder = ', '.join(f"{m}m" for m in CASHOUT_NUDGE_MINUTES[:_ladder_rounds()])
    print(f"📤 [CASHOUT] {len(CASHOUT_ROUTES)} route(s), group reminders at "
          f"{ladder}"
          + (f", crew DM at {CASHOUT_CREW_DM_MINUTES}m"
             if CASHOUT_CREW_DM_MINUTES > 0 else ", no crew DM"), flush=True)

    while True:
        await asyncio.sleep(20)
        now = datetime.now(timezone.utc)

        # Out of service: no reminders, no escalation DMs. The open requests are
        # deliberately KEPT rather than dropped - stopping is a pause, and
        # resuming has to find them still there.
        if _cashout_stopped:
            continue

        for handling, queue in list(_pending_cashouts.items()):
            # Nothing is chased into a group that is out of service. The open
            # requests are KEPT rather than dropped, the same way the
            # emergency stop keeps them - a pause is a pause, and resuming has
            # to find them still there.
            if route_paused(handling):
                continue
            for request in list(queue):
                # Before the exhausted guard: a problem raised while the chase
                # was still running must still be reported after it stops.
                await flush_pending_issue(handling, request)

                if request['exhausted']:
                    continue

                # One clock, running from when the request was posted. Which
                # rung is due is a function of that elapsed time alone, so
                # replaying a pass cannot double-post and a reaction cannot
                # push the ladder out.
                waited = (now - request['opened']).total_seconds() / 60.0
                await chase_cashout(handling, request, waited)


async def process_incoming(chat_id, text, origin, from_bot=False, sent_at=None,
                           source_msg_id=None):
    global forwarded_count, today_count

    print(f"📩 [MSG RECEIVED] ({origin}) Chat ID: {chat_id} | Text: '{text}'", flush=True)

    fixed = correct_timestamp(text, sent_at)
    if fixed != text:
        print(f"🕒 [TIME FIXED] -> {_STAMP_RE.search(fixed).group(0)}", flush=True)

    rule_key = resolve_rule_key(chat_id)
    if rule_key is None:
        print(f"⚠️ [SKIP] Chat ID {chat_id} is not in FORWARD_RULES.", flush=True)
        return

    if route_paused(rule_key):
        print(f"⏸️ [PAUSED] {chat_name(rule_key)} is out of service - "
              "nothing forwarded.", flush=True)
        return

    if rule_key in BOT_ONLY_SOURCES and not from_bot:
        print(f"⚠️ [SKIP] Chat {chat_id} forwards bot messages only.", flush=True)
        return

    lowered = text.lower()
    if not any(kw.lower() in lowered for kw in KEYWORDS):
        print("⚠️ [SKIP] Message does not contain keyword 'You received'.", flush=True)
        return

    print("✅ [KEYWORD MATCH] Forwarding message...", flush=True)
    source_key = (rule_key, source_msg_id) if source_msg_id is not None else None
    for target in FORWARD_RULES[rule_key]:
        await deliver_to_target(target, fixed, text, from_bot, source_key)


async def deliver_to_target(target, fixed, text, from_bot, source_key=None):
    """Send one notification to one group and move that group's books.

    Split out of process_incoming so the start-up catch-up can deliver to a
    single target: a message can be missing from one group while another
    already has it, and re-sending to the group that has it would duplicate.

    `source_key` identifies the message this came from, so a later reaction on
    that original can find this copy again and undo it - see retract_payment."""
    global forwarded_count, today_count

    # Checked here as well as in process_incoming: the boot sweep calls this
    # one directly, and a group that has just come back must not be handed the
    # history it deliberately missed.
    if chat_paused(target):
        print(f"⏸️ [PAUSED] {chat_name(target)} is out of service - "
              "not delivering.", flush=True)
        return False

    # Each target keeps its own books, so the totals shown are per target.
    movement = ledger_preview_payment(target, text) if from_bot else None
    body = fixed
    if movement:
        body = rewrite_totals(fixed, movement[1][0], movement[1][1])
    elif target in _ledger:
        # Human-pasted text. Show the ledger so the group never displays two
        # different sets of numbers, but do not book an amount we cannot
        # trust - only the notification bot moves the books.
        body = rewrite_totals(fixed, *ledger_snapshot(target))
    try:
        # Always sent with the bot token, never from the user account.
        sent = await send_group(target, body)
    except Exception as e:
        # Nothing committed: the ledger stays level with the last message
        # the group can actually see.
        print(f"❌ [FORWARD ERROR] Target {target}: {e}", flush=True)
        return False

    if movement:
        ledger_commit(target, movement[1])
        await note_payment(target)  # a real payment landed: reset the clock
    remember_delivery(source_key, target, getattr(sent, 'message_id', None),
                      (movement[1][0] - movement[0][0]) if movement else 0.0)
    print(f"🚀 [FORWARD SUCCESS] Sent to {target}", flush=True)
    forwarded_count += 1
    today_count += 1
    last_messages.append(body[:50])
    if len(last_messages) > 10:
        last_messages.pop(0)

    # After the payment confirmation, so the alert lands beneath it.
    if movement:
        await check_milestones(target, movement[0], movement[1])
    return True


def retract_text(amount, after):
    """Carries BOTH totals, like every other message that moves the ledger.

    recover_ledgers() rebuilds the books from the newest bot message showing
    both, and the message being deleted here was one of them - so without this
    the next redeploy would read an older figure and quietly put the payment
    back on."""
    # Word for word what "/add -10" posts, because that is exactly what this
    # is: Total In adjusted downwards. A retraction the group has never seen
    # before would read as a different kind of event; this one is already
    # familiar, and the payment vanishing at the same moment says the rest.
    heading = (f"✏️ Total In adjusted by -{amount:,.2f}$" if amount > 0
               else "✏️ Forwarded payment retracted")
    return (f"{heading}\n\n"
            f"📊 Group Total:\n"
            f"➕ Total In : {after[0]:,.2f}$\n"
            f"➖ Total Out: {after[1]:,.2f}$")


async def _source_message_text(client, chat_id, message_id):
    """Read the original back, since a reaction carries only ids and no text."""
    try:
        entity = await _resolve_one(client, chat_id)
        if entity is None:
            return None
        found = await client.get_messages(entity, ids=[message_id])
        if not found or found[0] is None:
            return None
        return found[0].raw_text or None
    except Exception as e:
        print(f"⚠️ [RETRACT] could not read the original in {chat_name(chat_id)}: {e}",
              flush=True)
        return None


async def _find_forwarded_copy(client, target, text):
    """Locate the copy we posted in the target, by content.

    Reuses the catch-up sweep's signature, which is the only thing common to
    both sides: the timestamp and the totals are BOTH rewritten on the way out,
    so the name and the amount are all that survive forwarding. Newest match
    wins - if the same amount really was received twice, the one being reacted
    to is the recent one."""
    want = _catchup_signature(text)
    try:
        entity = await _resolve_one(client, target)
        if entity is None:
            return None, None
        async for msg in client.iter_messages(entity, limit=RETRACT_SCAN_LIMIT):
            if msg.raw_text and _catchup_signature(msg.raw_text) == want:
                # The entity travels with the id. A message id only means
                # anything alongside the chat it was read from, and handing the
                # Bot API an id found this way is what produced "message to
                # delete not found" on 2026-08-03.
                return msg.id, entity
    except Exception as e:
        print(f"⚠️ [RETRACT] could not search {chat_name(target)}: {e}", flush=True)
    return None, None


async def retract_from_history(rule_key, message_id):
    """Work out what a forward did by reading the groups, not memory.

    This is what makes retraction survive a redeploy. The amount comes from the
    original's own text - which is exactly what was asked for: take off what
    the "You received" line says."""
    client = _active_client
    if client is None:
        return None
    text = await _source_message_text(client, rule_key, message_id)
    if not text:
        return None
    amount = parse_received_amount(text)
    if amount is None:
        print(f"↩️ [RETRACT] message {message_id} carries no amount - not a payment",
              flush=True)
        return None

    found = []
    for target in FORWARD_RULES.get(rule_key, []):
        copy_id, entity = await _find_forwarded_copy(client, target, text)
        if copy_id is None:
            print(f"↩️ [RETRACT] no copy of that payment found in "
                  f"{chat_name(target)} within {RETRACT_SCAN_LIMIT} messages",
                  flush=True)
            continue
        found.append((target, copy_id, amount, entity))
    if found:
        print(f"↩️ [RETRACT] recovered {amount:,.2f} from the group history "
              "(not in memory - forwarded before the last restart)", flush=True)
    return found or None


async def retract_payment(chat_id, message_id, user_id):
    """Undo a forwarded payment, because its original was reacted to.

    A payment is forwarded and booked within seconds, so undoing it means
    reaching into the target group after the fact: delete the copy, take the
    amount back off Total In.

    The order is deliberate and matches the rest of the ledger. The correction
    is POSTED first, then the books are committed, then the copy is deleted -
    so the figures can never run ahead of what the group can see, and a delete
    that fails still leaves the corrected totals as the newest ones for
    recover_ledgers() to find.

    Returns True if anything was retracted."""
    if chat_id not in RETRACT_SOURCES:
        return False
    if route_paused(chat_id):
        print(f"⏸️ [PAUSED] {chat_name(chat_id)} is out of service - "
              "not retracting.", flush=True)
        return False
    rule_key = resolve_rule_key(chat_id)
    if rule_key is None:
        return False
    if user_id not in LEDGER_ADMINS:
        print(f"⛔ [RETRACT] {user_id} may not retract in {chat_name(chat_id)}",
              flush=True)
        return False

    # Popped, not read: a second reaction on the same message must do nothing.
    entries = _delivered.pop((rule_key, message_id), None)
    if not entries:
        # Nothing in memory. A redeploy wipes it, so fall back to reading the
        # groups - the messages themselves are the durable record here, exactly
        # as they are for the ledger.
        entries = await retract_from_history(rule_key, message_id)
    if not entries:
        print(f"↩️ [RETRACT] nothing to retract for message {message_id} in "
              f"{chat_name(chat_id)} - no copy found, or never forwarded.",
              flush=True)
        return False

    for target, target_msg_id, amount, entity in entries:
        # An empty ledger is "not loaded yet", NOT "zero". Subtracting from a
        # guess wrote 0.00/0.00 into a live group on 2026-08-03 and made that
        # the newest totals message, which is what recover_ledgers() then reads
        # back. The Bot API handles updates on its own schedule and does not
        # wait for the boot sweep, so this really does happen.
        if target not in _ledger and _active_client is not None:
            await recover_one_ledger(_active_client, target)
        if target not in _ledger:
            print(f"⛔ [RETRACT] {chat_name(target)} books are not loaded - refusing "
                  "to subtract from a figure that is only a guess.", flush=True)
            await notify_admin(
                f"⚠️ Could not retract from {chat_name(target)}: its running totals "
                "are not loaded yet, and subtracting from an assumed zero would "
                "wipe the books.\n\nNothing was changed. Try the reaction again in "
                "a minute.")
            continue

        before = ledger_snapshot(target)
        # Refused rather than clamped, which is what /add -N already does with
        # an overshoot. Clamping would silently invent a figure and then delete
        # the evidence; refusing leaves both the books and the message alone
        # and points at the command that CAN set it outright.
        if amount > before[0]:
            print(f"⛔ [RETRACT] {amount:,.2f} is more than {chat_name(target)} has "
                  f"in Total In ({before[0]:,.2f}) - refusing.", flush=True)
            await send_group(
                target,
                f"⛔ Retracting {amount:,.2f}$ would take Total In below zero.\n\n"
                f"➕ Total In : {before[0]:,.2f}$\n"
                f"➖ Total Out: {before[1]:,.2f}$\n\n"
                "Nothing was changed. Use /set in <amount> to correct it outright.")
            continue
        after = (before[0] - amount, before[1])

        try:
            await send_group(target, retract_text(amount, after))
        except Exception as e:
            # Nothing committed and nothing deleted: the copy stays, and so do
            # the books that match it.
            print(f"❌ [RETRACT] correction NOT posted to {chat_name(target)}: {e} "
                  "- leaving the payment as it was.", flush=True)
            continue

        ledger_commit(target, after, f" (retracted {amount:,.2f} by {user_id})")

        if await delete_forwarded_copy(target, target_msg_id, entity):
            print(f"🗑️ [RETRACT] {amount:,.2f} taken back off {chat_name(target)} "
                  "and the forwarded copy deleted", flush=True)
        else:
            await notify_admin(
                f"⚠️ Retracted {amount:,.2f}$ from {chat_name(target)}, but the "
                "forwarded message could not be deleted.\n\n"
                "The totals are correct - the old message is still sitting there "
                "showing the figures from before. Delete it by hand.")
    return True


async def delete_forwarded_copy(target, message_id, entity=None):
    """Remove the copy, by whichever route can actually address it.

    The bot token is tried first: when the id came from our own send it is
    exactly right, and the bot is the message's author. When the id was instead
    FOUND by the userbot, it is only meaningful alongside the entity it was read
    from - so the user account, holding that entity, is the one that can
    reliably address it. Handing that id to the Bot API instead is what produced
    "message to delete not found".

    Telegram also refuses to let a bot delete a group message over 48 hours old,
    which no amount of routing fixes."""
    try:
        await bot.delete_message(target, message_id)
        return True
    except Exception as e:
        print(f"⚠️ [RETRACT] bot could not delete {message_id} in "
              f"{chat_name(target)}: {e}", flush=True)

    if entity is None or _active_client is None:
        return False
    try:
        await _active_client.delete_messages(entity, [message_id])
        print(f"🗑️ [RETRACT] deleted via the user account instead", flush=True)
        return True
    except Exception as e:
        print(f"⚠️ [RETRACT] the user account could not delete it either: {e}",
              flush=True)
        return False


def mention_alert_text(chat_id, who, user_id, when, text):
    """Who, where, when, and what they actually said.

    The message itself is included because the point is to save a trip back
    into a muted group - a notice that says only "you were mentioned" would
    make you open the thing it is standing in for."""
    stamp = (when.astimezone(LOCAL_TZ).strftime('%I:%M %p - %d %b %Y')
             if when else 'time unknown')
    body = ' '.join((text or '').split())
    if len(body) > 400:
        body = body[:400] + '…'
    return ("🔔 YOU WERE MENTIONED\n\n"
            f"Group: {chat_name(chat_id)}\n"
            f"By: {who}\n"
            f"Their id: {user_id}\n"
            f"At: {stamp}\n\n"
            f"{body}")


async def observe_mentions(chat_id, text, message_id, sent_at,
                           user_id=None, username=None, full_name=None):
    """Privately tell Ethan and Larry when somebody @s them in a watched group.

    Telegram never tells a bot whether a chat is muted - that is a per-user
    setting the API does not expose - so this does not try to detect it. These
    four groups are muted as a matter of course, which is the whole reason the
    DM has to exist.

    Skipped: the bot's own posts, since it tags people on every request and
    reminder; and whoever sent the message, who does not need telling about
    their own. Every mention is sent - unlike the cashout escalations, these
    are separate events rather than repeats of one."""
    if _MENTION_RE is None or not text:
        return
    if chat_id not in MENTION_CHATS or chat_paused(chat_id):
        return
    if BOT_ID is not None and user_id == BOT_ID:
        return
    if not _MENTION_RE.search(text):
        return
    if _is_duplicate(('mention', chat_id, message_id)):
        return                          # both input paths see the same message

    author = (username or '').lower()
    targets = [h for h in MENTION_ALERT_HANDLES if h.lower() != author]
    if not targets:
        return

    who = describe_user(username, full_name)
    print(f"🔔 [MENTION] {who} mentioned someone in {chat_name(chat_id)}", flush=True)
    missed = await dm_handles(
        targets, mention_alert_text(chat_id, who, user_id, sent_at, text))
    if missed:
        # Deliberately not warn_unreachable(): that one says "about an overdue
        # cashout", which this is not, and a mention is not worth waking the
        # admin a second time over.
        print(f"❌ [MENTION] could not DM: {', '.join('@' + h for h in missed)}",
              flush=True)


# ---------------------------------------------------------------------------
# Bot API handlers (unchanged behaviour, async syntax)
# ---------------------------------------------------------------------------

@bot.message_handler(commands=['start'])
async def start(message):
    if message.chat.id in LEDGER_ADMINS:
        await bot.reply_to(message, 'Bot is running. Use /help to see all commands.')


@bot.message_handler(commands=['status'])
async def status(message):
    if message.chat.id in LEDGER_ADMINS:
        paused = sorted(t for t in IDLE_ALERT_CHATS if _idle_slot(t)['paused'])
        prompts = ('off in ' + ', '.join(str(t) for t in paused)) if paused else 'on'
        open_requests = sum(len(q) for q in _pending_cashouts.values())
        # First line when it is off, not buried at the bottom. Somebody reading
        # /status during an incident needs to know the flow is out of service
        # before they read anything else - the counts below all look normal
        # while nothing is actually being forwarded.
        heading = ('⏸️ CASHOUT FORWARDING IS STOPPED - /cashout on to start it\n\n'
                   if _cashout_stopped else '')
        # Same reasoning as the line above, and the same trap: with a group out
        # of service every count below still looks perfectly normal while
        # nothing whatsoever is happening in it.
        if PAUSED_CHAT_IDS:
            heading += ('⏸️ OUT OF SERVICE: '
                        + ', '.join(chat_name(c) for c in PAUSED_CHAT_IDS)
                        + ' - /group on to put them back\n\n')
        txt = (f"{heading}Bot Online\nGroups: {len(FORWARD_RULES)}\n"
               f"Forwarded: {forwarded_count}\nUptime: {get_uptime()}\n"
               f"Userbot: {userbot_status}\n"
               f"Payment prompts: {prompts}\n"
               f"Cashout flow: {'STOPPED' if _cashout_stopped else 'running'}\n"
               f"Cashouts awaiting /out: {open_requests}")
        await bot.reply_to(message, txt)


@bot.message_handler(commands=['help'])
async def help_command(message):
    """Private-chat command reference. Values come from the live config so this
    cannot drift out of date when an env var changes."""
    if message.chat.id not in LEDGER_ADMINS:
        return

    groups = ' or '.join(IDLE_NAMES[t] for t in sorted(IDLE_ALERT_CHATS))
    await send_group(message.chat.id,
        "📖 COMMANDS\n"
        "\n"
        f"LEDGER — type these inside {groups}\n"
        "\n"
        "/add 500 — adds 500.00$ to Total In\n"
        "/out 100 — adds 100.00$ to Total Out\n"
        "/add -100 — takes 100.00$ back off Total In\n"
        "/out -50 — takes 50.00$ back off Total Out\n"
        "(the space is optional: /add-100 works too)\n"
        "/set in 800 — sets Total In to 800.00$\n"
        "/set out 300 — sets Total Out to 300.00$\n"
        "\n"
        "PAYMENT PROMPTS — here in private, or silently in a group\n"
        "\n"
        "/pause — stops prompts in both groups\n"
        "/pause gaffer — stops them in CHIME GAFFER only\n"
        "/pause piccaso — stops them in CHIME PICCASO only\n"
        "/resume — starts them again, same three forms\n"
        "\n"
        "TAKING A GROUP OUT OF SERVICE — here in private\n"
        "\n"
        "/group — what is running and what is not\n"
        "/group off piccaso — stops EVERYTHING for that route\n"
        "/group on piccaso — puts it back, working as before\n"
        "\n"
        "Out of service means out of service: nothing is\n"
        "forwarded, chased, booked or answered there, and\n"
        "nothing about it reaches you or the crew. It lasts\n"
        "through a redeploy, and coming back never replays\n"
        "the time it was off. Names: "
        + ', '.join(sorted(set(GROUP_ALIASES))) + "\n"
        "\n"
        "REPORTS\n"
        "\n"
        f"A workbook lands here every day at {REPORT_AT} Nepal\n"
        "time, covering the whole day just finished. If a\n"
        "cashout is still unanswered it waits for the /out,\n"
        f"up to {REPORT_WAIT_MINUTES} min.\n"
        "/report — today so far, as a workbook, privately\n"
        "/report yesterday — the day before\n"
        "/report 2026-08-05 — any past day\n"
        "/report 24h — typed in CHIME PICCASO or CHIME\n"
        "GAFFER, posts THAT group's own figures in the\n"
        "group for everyone there to read. 6h, 3d, any\n"
        "span. It shows that group's money only — never\n"
        "the other group, the crew, or the gap.\n"
        "\n"
        "EMERGENCY STOP — from here, or from any cashout group\n"
        "\n"
        "/cashout off — stops the whole cashout flow: nothing\n"
        "forwarded, nothing chased, no totals moved. Open\n"
        "requests are kept, not cancelled.\n"
        "/cashout on — starts it again\n"
        "/cashout — says which it currently is\n"
        "Nothing is ever posted in a group — the switch is\n"
        "private to you and Larry. It survives a redeploy,\n"
        "and anything real that arrives while stopped is\n"
        "sent to you both.\n"
        "\n"
        "CASHOUTS — automatic, nothing to type\n"
        "\n"
        f"A \"{CASHOUT_KEYWORD}\" in CHIME PICCASO or CHIME GAFFER is\n"
        "posted to its group with the crew tagged.\n"
        f"No /out and the crew are tagged again at\n"
        f"{_ladder_text()} min, with Ethan and Larry getting a\n"
        "private warning on the first one.\n"
        f"Nobody acknowledges it by {CASHOUT_CREW_DM_MINUTES} min and the crew\n"
        "get a private warning too — deleted again as\n"
        "soon as the /out lands. The request stays open.\n"
        "If the CREW react, the group is not chased\n"
        f"again — Larry hears {CASHOUT_SEEN_DM_MINUTES} min later if there is\n"
        "still no /out, and the crew get no DM at all.\n"
        "A /out paying a different figure than the\n"
        "request asked for is DMed to Larry at once.\n"
        "A /out EDITED onto a message counts the same\n"
        f"as one typed fresh; problems wait {CASHOUT_ISSUE_DELAY_SECONDS}s first,\n"
        "in case the edit is on its way.\n"
        "Their /out reply goes back to the chime group, its\n"
        "amount is added to that group's Total Out, and the\n"
        "original request there gets a ❤.\n"
        "A /out with nothing pending is ignored.\n"
        "\n"
        "INFO\n"
        "\n"
        "/status — bot state, userbot, paused groups, and\n"
        "whether the cashout flow is stopped\n"
        "/ping — quick alive check\n"
        "/help — this list\n"
        "\n"
        "GOOD TO KNOW\n"
        "\n"
        "Only Ethan and Larry can use any of these.\n"
        "Payments add to Total In on their own.\n"
        "A payment also lifts a pause by itself.\n"
        f"Milestones fire every {MILESTONE_IN_STEP:,}$ in "
        f"and every {MILESTONE_OUT_STEP:,}$ out.\n"
        f"Prompts ask the group after {_humanise(IDLE_ALERT_MINUTES)} "
        "without a payment.\n"
        "\n"
        "In a group, /pause and /resume post nothing at all — check /status."
        + (("\n\nOUT OF SERVICE: "
            + ', '.join(chat_name(c) for c in PAUSED_CHAT_IDS)
            + "\nThe bot does nothing at all in those groups —\n"
              "no forwarding, no cashouts, no prompts, no replies.")
           if PAUSED_CHAT_IDS else ""))


@bot.message_handler(commands=['ping'])
async def ping(message):
    if message.chat.id in LEDGER_ADMINS:
        await bot.reply_to(message, 'Pong! Bot is alive.')


@bot.message_handler(commands=['add', 'out'])
async def ledger_command(message):
    """/add <amount> and /out <amount>, target groups only, Ethan and Larry only."""
    chat_id = message.chat.id
    user_id = getattr(message.from_user, 'id', None)

    # Nothing at all in a paused group, not even a refusal. A refusal is a
    # reply, and a reply says the bot is still listening in a group it is
    # meant to read as absent from. Every command still works from a DM and
    # from the groups that are in service.
    if chat_paused(chat_id):
        return

    # A /out typed in a handling group may be answering a cashout request, and
    # telebot routes commands here rather than to forward_text. This has to run
    # before the guard below, which would otherwise drop it. It does nothing
    # unless a request we forwarded is actually open in that chat.
    await cashout_from_bot_api(message)

    # Silent in source groups and DMs, so the source side cannot reach the books.
    # LEDGER_CHATS rather than TARGET_CHATS: a group that stopped receiving
    # forwards still keeps its books, and they are still Ethan's and Larry's to
    # correct.
    if chat_id not in LEDGER_CHATS:
        return

    if user_id not in LEDGER_ADMINS:
        print(f"⛔ [LEDGER] {user_id} denied in {chat_id}", flush=True)
        await bot.reply_to(message, "⛔ Not permitted. Only Ethan and Larry can "
                                    "change the totals.")
        return

    name = _CMD_NAME_RE.match(message.text)
    command = name.group(1).lower() if name else ''
    match = _CMD_AMOUNT_RE.match(message.text)
    amount = _to_float(match.group(2)) if match else None
    if amount is None or amount == 0:
        await bot.reply_to(message, f"Usage: /{command} <amount>\n"
                                    f"Example: /{command} 100\n"
                                    f"         /{command} -100  (takes it back off)")
        return
    if match.group(1) == '-':
        amount = -amount

    before = ledger_snapshot(chat_id)
    column = 'Total In' if command == 'add' else 'Total Out'
    if command == 'add':
        after = (before[0] + amount, before[1])
    else:
        after = (before[0], before[1] + amount)

    # A correction that overshoots is a typo, not an instruction. Refuse it
    # rather than publish a negative total, which would then be carried into
    # every forwarded notification by rewrite_totals().
    if after[0] < 0 or after[1] < 0:
        print(f"⛔ [LEDGER] /{command} {amount:,.2f} by {user_id} would take "
              f"{column} below zero in {chat_id}", flush=True)
        await bot.reply_to(message,
                           f"⛔ That would take {column} below zero.\n\n"
                           f"➕ Total In : {before[0]:,.2f}$\n"
                           f"➖ Total Out: {before[1]:,.2f}$\n\n"
                           f"Use /set {'in' if command == 'add' else 'out'} "
                           "<amount> to set it outright.")
        return

    if amount < 0:
        heading = f"✏️ {column} adjusted by {amount:,.2f}$"
    elif command == 'add':
        heading = f"💰 Deposit = +{amount:,.2f}$"
    else:
        heading = f"📤 Out = -{amount:,.2f}$"

    try:
        await send_group(chat_id,
                               f"{heading}\n\n"
                               f"📊 Group Total:\n"
                               f"➕ Total In : {after[0]:,.2f}$\n"
                               f"➖ Total Out: {after[1]:,.2f}$")
    except Exception as e:
        print(f"❌ [LEDGER] /{command} NOT applied, send failed: {e}", flush=True)
        return

    ledger_commit(chat_id, after, f" (/{command} {amount:,.2f} by {user_id})")
    await check_milestones(chat_id, before, after)


@bot.message_handler(func=lambda m: bool(getattr(m, 'text', None))
                     and bool(_JOINED_LEDGER_RE.match(m.text)))
async def ledger_joined_command(message):
    """"/add-100" and "/out-50", written without the space.

    Telegram reads the command as the word right after the slash, so these
    arrive as the commands "add-100" and "out-50" and never reach the handler
    above. Same code, just reached a different way. Registered ahead of the
    catch-all text handler, which telebot would otherwise hand them to."""
    await ledger_command(message)


@bot.message_handler(commands=['pause', 'resume'])
async def idle_pause_command(message):
    """/pause and /resume the payment prompts, per group.

    For when the hold-up is on your side and asking the group whether they have
    a problem would be the wrong message."""
    chat_id = message.chat.id
    user_id = getattr(message.from_user, 'id', None)
    command = message.text.lstrip('/').split('@')[0].split()[0].lower()
    pausing = command == 'pause'

    # ---- private chat: name a group or hit both, and get a reply ----
    if chat_id in LEDGER_ADMINS:
        if user_id not in LEDGER_ADMINS:
            return
        parts = message.text.split()
        argument = parts[1].lower() if len(parts) > 1 else ''
        if not argument:
            targets = sorted(IDLE_ALERT_CHATS)
        elif argument in IDLE_ALIASES:
            targets = [IDLE_ALIASES[argument]]
        else:
            await bot.reply_to(message,
                               f"Usage: /{command}          (both groups)\n"
                               f"       /{command} gaffer   (CHIME GAFFER only)\n"
                               f"       /{command} piccaso  (CHIME PICCASO only)")
            return

        # A group that is out of service has no payment prompts to pause or
        # resume. Saying so is better than appearing to act on it - the reply
        # is where somebody would otherwise read a silent no-op as a yes.
        blocked = [t for t in targets if chat_paused(t)]
        targets = [t for t in targets if not chat_paused(t)]
        if blocked:
            await bot.reply_to(message, "⏸️ Out of service: " + ', '.join(
                IDLE_NAMES.get(t, str(t)) for t in blocked)
                + "\n\nPayment prompts there are off with everything else "
                  "until the group is back.")
        if not targets:
            return

        for target in targets:
            set_idle_paused(target, pausing)
        names = ', '.join(IDLE_NAMES.get(t, str(t)) for t in targets)
        print(f"{'🔇' if pausing else '🔔'} [IDLE] {names} "
              f"{'paused' if pausing else 'resumed'} by {user_id} via DM", flush=True)

        if pausing:
            await bot.reply_to(message,
                               f"🔇 Paused: {names}\n\n"
                               "Nothing will be posted in those groups.\n"
                               "Send /resume when you want them back.")
        else:
            await bot.reply_to(message,
                               f"🔔 Resumed: {names}\n\n"
                               f"Next prompt after {_humanise(IDLE_ALERT_MINUTES)} "
                               "of silence.")
        return

    # ---- inside a target group: deliberately SILENT ----
    # Pausing usually means the hold-up is on our side, and announcing it would
    # be exactly the wrong message. Check the state with /status.
    if chat_id not in IDLE_ALERT_CHATS or chat_paused(chat_id):
        return

    if user_id not in LEDGER_ADMINS:
        print(f"⛔ [IDLE] pause/resume by {user_id} denied in {chat_id}", flush=True)
        await bot.reply_to(message, "⛔ Not permitted. Only Ethan and Larry can "
                                    "change this.")
        return

    set_idle_paused(chat_id, pausing)
    print(f"{'🔇' if pausing else '🔔'} [IDLE] {chat_id} "
          f"{'paused' if pausing else 'resumed'} by {user_id}", flush=True)


@bot.message_handler(commands=['cashout'])
async def cashout_switch_command(message):
    """/cashout off - stop the whole cashout flow. /cashout on - start it again.

    Bare /cashout reports the state. Ethan and Larry only, from a DM or from any
    of the cashout groups, because when this is needed it is needed from
    whatever chat is already open."""
    chat_id = message.chat.id
    user_id = getattr(message.from_user, 'id', None)

    in_dm = chat_id in LEDGER_ADMINS
    in_cashout_chat = (_canonical(chat_id, CASHOUT_ROUTES) is not None
                       or _canonical(chat_id, _CASHOUT_HANDLERS) is not None)
    if not (in_dm or in_cashout_chat):
        return

    # Silent in a paused group, like every other command there. The switch is
    # still reachable from a DM and from the cashout groups still in service,
    # which is what it was written for.
    if chat_paused(chat_id):
        return

    # Every answer this command gives is private. Typed in a group, replying
    # there would put the switch in front of the crew - which is the one thing
    # it must never do, and the reason a refusal is silent there too rather than
    # announcing that a kill switch exists.
    async def answer(text):
        if in_dm:
            await bot.reply_to(message, text)
        else:
            try:
                await send_group(user_id, text)
            except Exception as e:
                print(f"⚠️ [SWITCH] could not confirm privately to {user_id}: {e}",
                      flush=True)

    if user_id not in LEDGER_ADMINS:
        # Logged, never answered. In a group a refusal would announce that a
        # kill switch exists, to exactly the people it is kept from; and the
        # only private chats this command is accepted in are Ethan's and
        # Larry's, so there is nobody else here to answer.
        print(f"⛔ [SWITCH] /cashout by {user_id} denied in {chat_id}", flush=True)
        return

    parts = (message.text or '').split()
    argument = parts[1].lower() if len(parts) > 1 else ''
    who = f"@{getattr(message.from_user, 'username', None) or user_id}"

    if argument in ('off', 'stop'):
        if _cashout_stopped:
            await answer("⏸️ Already stopped. /cashout on to start it.")
            return
        failed = await set_cashout_stopped(True, who)
        open_now = sum(len(q) for q in _pending_cashouts.values())
        # The private message IS the durable record, so failing to deliver it
        # means the stop will not survive the next deploy. Say so plainly.
        warning = ('\n\n⚠️ Could not reach ' + ', '.join(f'@{h}' for h in failed)
                   + ', so this will NOT survive a redeploy. Check /cashout again '
                     'after the next deploy.') if failed else ''
        await answer(
            "⏸️ Cashout forwarding STOPPED.\n\n"
            "No requests forwarded, nothing chased, no totals moved.\n"
            "Nothing has been posted in any group.\n"
            f"{open_now} request(s) already open are kept, not cancelled.\n\n"
            "Anything real that arrives while stopped is sent to you and Larry "
            "instead of being dropped.\n\n"
            "Send /cashout on to start it again." + warning)
        return

    if argument in ('on', 'start'):
        if not _cashout_stopped:
            await answer("▶️ Already running.")
            return
        failed = await set_cashout_stopped(False, who)
        open_now = sum(len(q) for q in _pending_cashouts.values())
        await answer(
            "▶️ Cashout forwarding RESUMED.\n\n"
            "Nothing was posted in any group.\n\n"
            f"{open_now} request(s) still open; chasing continues from here.\n\n"
            "Anything that arrived while it was stopped was NOT queued - if you "
            "were told about one, it still needs handling by hand.")
        return

    open_now = sum(len(q) for q in _pending_cashouts.values())
    if _cashout_stopped:
        await answer(
            "⏸️ Cashout forwarding is STOPPED"
            f"{f' (by {_cashout_stopped_by})' if _cashout_stopped_by else ''}.\n\n"
            f"{open_now} request(s) held open.\n\n"
            "Send /cashout on to start it again.")
    else:
        await answer(
            "▶️ Cashout forwarding is running normally.\n\n"
            f"{open_now} request(s) open.\n\n"
            "Send /cashout off to stop everything.")


@bot.message_handler(commands=['group', 'groups'])
async def group_switch_command(message):
    """/group off piccaso - take a route out of service. /group on - put it back.

    Bare /group reports what is in service and what is not. Ethan and Larry
    only, from a private chat or from any group still in service.

    Naming either end takes the WHOLE route: the chime group, the handling
    group and the payment direction between them. Half a route running is the
    failure this is meant to prevent, and it would be far too easy to create
    from a phone by naming one group and assuming the other followed.

    A group that is out of service ignores everything, this command included -
    so it is put back from a private chat, which is where it is worked from
    anyway."""
    chat_id = message.chat.id
    user_id = getattr(message.from_user, 'id', None)

    in_dm = chat_id in LEDGER_ADMINS
    if not (in_dm or _canonical(chat_id, CASHOUT_ROUTES) is not None
            or _canonical(chat_id, _CASHOUT_HANDLERS) is not None):
        return
    if chat_paused(chat_id):
        return

    # Private, for the same reason the kill switch is: which groups the bot is
    # working is not the crew's business, and a refusal in a group would tell
    # them the switch exists.
    async def answer(text):
        if in_dm:
            await bot.reply_to(message, text)
        else:
            try:
                await send_group(user_id, text)
            except Exception as e:
                print(f"⚠️ [PAUSED] could not confirm privately to {user_id}: {e}",
                      flush=True)

    if user_id not in LEDGER_ADMINS:
        print(f"⛔ [PAUSED] /group by {user_id} denied in {chat_id}", flush=True)
        return

    parts = (message.text or '').split()
    action = parts[1].lower() if len(parts) > 1 else ''
    argument = ' '.join(parts[2:]).strip() if len(parts) > 2 else ''
    who = f"@{getattr(message.from_user, 'username', None) or user_id}"

    if action not in ('off', 'stop', 'on', 'start'):
        await answer(f"{group_switch_state_text()}\n\n"
                     "/group off piccaso — take that route out of service\n"
                     "/group on piccaso — put it back\n"
                     "/group off — everything currently running\n"
                     "/group on — everything currently out of service\n\n"
                     "Names: " + ', '.join(sorted(set(GROUP_ALIASES))) + "\n"
                     "Naming either end takes the whole route with it.")
        return

    pausing = action in ('off', 'stop')

    if argument:
        chats = resolve_group(argument)
        if chats is None:
            await answer(f"❓ I do not know a group called {argument!r}.\n\n"
                         "Try: " + ', '.join(sorted(set(GROUP_ALIASES)))
                         + "\nOr give the chat id.")
            return
    else:
        # No name given: every group on the other side of the switch. Spelled
        # out in the confirmation, because "everything" typed on a phone should
        # never leave somebody guessing which groups just moved.
        chats = {c for c in _known_groups() if chat_paused(c) is not pausing}

    # The whole route moves, not just the end that was named, and only the
    # groups this actually changes are reported as changed.
    changing = {member for c in chats for member in route_members(c)
                if chat_paused(member) is not pausing}
    if not changing:
        await answer("Nothing to do — that is already the state.\n\n"
                     + group_switch_state_text())
        return

    touched = _with_variants(changing)
    open_now = sum(len(queue) for handling, queue in _pending_cashouts.items()
                   if handling in touched)
    failed = await set_groups_paused(changing, pausing, who)
    names = ', '.join(chat_name(c) for c in sorted(changing,
                                                   key=lambda c: chat_name(c)))

    # The private message IS the durable record, exactly as it is for the kill
    # switch, so failing to deliver it means this will not survive a redeploy.
    warning = ('\n\n⚠️ Could not reach ' + ', '.join(f'@{h}' for h in failed)
               + ', so this will NOT survive a redeploy. Send /group again '
                 'after the next deploy to check.') if failed else ''

    if pausing:
        await answer(
            f"⏸️ OUT OF SERVICE: {names}\n\n"
            "Nothing forwarded, no cashout opened, answered or chased, no "
            "totals moved, no prompts, no reply to anything typed there.\n"
            "Nothing was posted in any group, and nobody else was told.\n"
            f"{open_now} request(s) already open there are kept, not "
            "cancelled.\n\n"
            "Anything that happens there while it is out of service reaches "
            "NOBODY — that is what being out of service means.\n\n"
            f"Send /group on {argument or ''}".rstrip() + " to put it back."
            + warning)
    else:
        await answer(
            f"▶️ BACK IN SERVICE: {names}\n\n"
            "Working exactly as before — same routes, same crew, same ladder, "
            "same books.\n"
            "Nothing was posted in any group.\n\n"
            "The time they were out of service is NOT replayed: nothing from "
            "that window will be forwarded or booked, now or on the next "
            "deploy." + warning)


@bot.message_handler(commands=['set'])
async def ledger_set_command(message):
    """/set in <amount> or /set out <amount> - overwrite a column outright.

    This is the correction route: /add and /out can only increase a column, so a
    figure that has gone wrong needs to be set directly."""
    chat_id = message.chat.id
    user_id = getattr(message.from_user, 'id', None)

    if chat_paused(chat_id):
        return

    if chat_id not in LEDGER_CHATS:
        return

    if user_id not in LEDGER_ADMINS:
        print(f"⛔ [LEDGER] /set by {user_id} denied in {chat_id}", flush=True)
        await bot.reply_to(message, "⛔ Not permitted. Only Ethan and Larry can "
                                    "change the totals.")
        return

    match = _SET_RE.match(message.text)
    amount = _to_float(match.group(2)) if match else None
    if amount is None or amount < 0:
        await bot.reply_to(message, "Usage: /set in <amount>   or   /set out <amount>\n"
                                    "Example: /set in 800")
        return

    column = match.group(1).lower()
    before = ledger_snapshot(chat_id)
    after = (amount, before[1]) if column == 'in' else (before[0], amount)

    try:
        await send_group(chat_id,
                               f"✏️ Corrected: Total {'In' if column == 'in' else 'Out'} "
                               f"set to {amount:,.2f}$\n\n"
                               f"📊 Group Total:\n"
                               f"➕ Total In : {after[0]:,.2f}$\n"
                               f"➖ Total Out: {after[1]:,.2f}$")
    except Exception as e:
        print(f"❌ [LEDGER] /set NOT applied, send failed: {e}", flush=True)
        return

    ledger_commit(chat_id, after, f" (/set {column} {amount:,.2f} by {user_id})")
    await check_milestones(chat_id, before, after)


async def cashout_from_bot_api(message, is_edit=False):
    """Bot API side of observe_cashout.

    Telethon normally sees these first and the dedup key discards this copy, but
    it keeps the cashout flow alive if the userbot is down.

    Media carries its words in .caption and leaves .text empty, so the fallback
    is what lets a /out written on a screenshot be read at all. Only
    cashout_caption() below reaches here with one, and only when it holds a
    /out - see is_caption_out()."""
    reply = getattr(message, 'reply_to_message', None)
    sent_at = (datetime.fromtimestamp(message.date, tz=timezone.utc)
               if getattr(message, 'date', None) else None)
    text = getattr(message, 'text', None) or getattr(message, 'caption', None)
    sender = getattr(message, 'from_user', None)
    full_name = ' '.join(part for part in (getattr(sender, 'first_name', None),
                                           getattr(sender, 'last_name', None)) if part)
    await observe_cashout(message.chat.id, text, message.message_id, sent_at,
                          getattr(sender, 'id', None),
                          getattr(sender, 'username', None),
                          getattr(reply, 'message_id', None),
                          full_name or None,
                          has_media=getattr(message, 'content_type', 'text') != 'text',
                          is_edit=is_edit,
                          media_group_id=getattr(message, 'media_group_id', None))

    # Every Bot API handler that can carry a mention already routes through
    # here, so this is the one place it has to be called from. Kept after the
    # cashout work: a mention never changes what a message means, and the
    # money side goes first.
    await observe_mentions(message.chat.id, text, message.message_id, sent_at,
                           getattr(sender, 'id', None),
                           getattr(sender, 'username', None), full_name or None)


@bot.message_reaction_handler(func=lambda r: True)
async def on_request_reaction(reaction):
    """A reaction on a forwarded cashout request buys another timeout.

    Telegram only delivers these to a bot that is an ADMINISTRATOR in the chat,
    which is why run_bot() has to ask for message_reaction explicitly and why
    the bot needs admin rights in the handling groups. A reaction being removed
    is not an answer, and an anonymous admin cannot be attributed to anyone, so
    both are ignored."""
    if not getattr(reaction, 'new_reaction', None):
        return
    user = getattr(reaction, 'user', None)
    if user is None:
        return
    if route_paused(getattr(reaction.chat, 'id', None)):
        return                      # out of service: not even a ❤ or a retraction
    remember_user(user)
    full_name = ' '.join(part for part in (getattr(user, 'first_name', None),
                                           getattr(user, 'last_name', None)) if part)
    await note_cashout_seen(reaction.chat.id, reaction.message_id,
                            getattr(user, 'id', None), getattr(user, 'username', None),
                            full_name or None)

    # The two cannot collide: note_cashout_seen only looks at handling groups,
    # retract_payment only at RETRACT_SOURCES, and no chat is both.
    await retract_payment(reaction.chat.id, reaction.message_id,
                          getattr(user, 'id', None))


# ══════════════════════════ THE DAILY REPORT ══════════════════════════
#
# One workbook a day, built by READING THE GROUPS BACK - never by accumulating
# rows as events happen. Railway wipes the disk on every deploy and there were
# two deploys inside one hour on 2026-08-10, so anything kept in memory or on
# disk would report a fraction of the day and never say which fraction. The
# messages are the durable record here exactly as they are for the ledger, and
# rebuilding from them means any past day can still be produced on demand.
#
# This half of the file READS. It never posts into a group, never moves a
# ledger and never touches an open request: a reporting bug must not be able to
# cost money.

REPORT_ENABLED = os.getenv('REPORT_ENABLED', '1') not in ('0', 'false', 'False')
# Local time. Midnight, so a report covers one whole calendar day.
REPORT_AT = os.getenv('REPORT_AT', '00:00')
# How long midnight will wait for a cashout that is still unanswered. The day is
# not really finished while one is outstanding - the /out belongs to it - but a
# request nobody ever answers must not hold the report for ever.
REPORT_WAIT_MINUTES = int(os.getenv('REPORT_WAIT_MINUTES', '180'))
REPORT_SCAN_LIMIT = int(os.getenv('REPORT_SCAN_LIMIT', '3000'))

# 📤 Out = -200.00$ - what book_cashout_out() posts when it moves Total Out.
_REPORT_OUT_RE = re.compile(r'Out\s*=\s*-?\s*([\d,]+(?:\.\d+)?)\s*\$', re.I)
# ✏️ Total In adjusted by -5.00$ - a retraction, or /add -N by hand.
_REPORT_ADJ_RE = re.compile(
    r'Total\s+(In|Out)\s+adjusted\s+by\s+(-?[\d,]+(?:\.\d+)?)\s*\$', re.I)
# The payer's name off the notification itself.
_REPORT_FROM_RE = re.compile(
    r'You\s+received\s+\$?\s*[\d,]+(?:\.\d+)?\s+from\s+([^\n]+)', re.I)


def report_pairs():
    """The two ends of each route, as one pair.

    A pair is one handling/source group and the chime group it serves - MH X
    LARRY GROUP 2 with CHIME PICCASO, Chime Rev & out no-7 with CHIME GAFFER.
    Payments travel one way and cashouts the other BETWEEN THE SAME TWO GROUPS,
    which is what makes a gap between them meaningful.

    Derived from the live tables rather than written out, so re-pointing a route
    in Railway cannot leave the report reconciling groups that no longer talk to
    each other."""
    pairs = []
    for chime, route in CASHOUT_ROUTES.items():
        handling = route['handling']
        # A pair that is out of service has nothing to reconcile, and a row of
        # zeroes would read as a quiet day rather than as a group the bot is
        # not working. report_summary_text() names them instead.
        if chat_paused(chime) or chat_paused(handling):
            continue
        source = _canonical(handling, FORWARD_RULES)
        targets = FORWARD_RULES.get(source, []) if source is not None else []
        # 'paired' says the two really are two ends of one route: payments come
        # this way and cashouts go back. Only then does a gap between them mean
        # anything - reconciling groups that do not talk would invent one.
        pairs.append({'chime': chime, 'handling': handling,
                      'paired': _canonical(chime, {t: None for t in targets})
                      is not None})
    return pairs


def _report_window(day):
    """The UTC window covering one LOCAL calendar day.

    Nepal is +05:45, so a 'day' is nothing like a UTC day and the difference is
    a whole evening of payments landing in the wrong report."""
    start_local = datetime(day.year, day.month, day.day, tzinfo=LOCAL_TZ)
    return start_local.astimezone(timezone.utc), \
        (start_local + timedelta(days=1)).astimezone(timezone.utc)


async def _report_scan(client, chat_id, start, end):
    """Every message one group saw inside the window, oldest first.

    None means the group could not be read at all, which the caller reports as
    unknown rather than as an empty day - an empty sheet and a silent failure
    look identical otherwise, and one of them means the books are unwatched."""
    entity = await _resolve_one(client, chat_id)
    if entity is None:
        return None
    rows = []
    try:
        async for msg in client.iter_messages(entity, limit=REPORT_SCAN_LIMIT):
            if msg.date < start:
                break
            if msg.date >= end:
                continue
            rows.append(msg)
    except Exception as e:
        print(f"⚠️ [REPORT] could not read {chat_name(chat_id)}: {e}", flush=True)
        return None
    rows.reverse()
    return rows


def _hhmm(when):
    return when.astimezone(LOCAL_TZ).strftime('%H:%M')


def _report_is_ours(msg):
    """Did the BOT post this, as far as Telegram will say?

    Only what the bot posted moved the books. A human pasting a notification
    into a chime group never moved the ledger - deliver_to_target() books an
    amount only `if from_bot` - so counting it would invent money that was
    never on the books and report a gap that is not there.

    UNKNOWN authorship counts as ours, deliberately. Telegram credits a channel
    post to the channel rather than to the bot that sent it, and an anonymous
    admin's post to the group, so demanding a positive match on BOT_ID can
    exclude the bot's own messages entirely - and a report that silently
    counted nothing would look exactly like a quiet day. Only a different,
    identifiable sender is dropped. Same reasoning as _delivered_signatures()."""
    sender_id = getattr(msg, 'sender_id', None)
    if sender_id is None:
        sender_id = getattr(getattr(msg, 'sender', None), 'id', None)
    return sender_id is None or BOT_ID is None or sender_id == BOT_ID


def _read_chime_side(messages):
    """What the chime group's own books did today."""
    payments, booked_out, adjustments = [], [], []
    closing = None
    for msg in messages:
        text = msg.raw_text or ''
        if not text or not _report_is_ours(msg):
            continue
        amount = parse_received_amount(text)
        if amount is not None and 'received' in text.lower():
            name = _REPORT_FROM_RE.search(text)
            payments.append({'at': msg.date, 'amount': amount,
                             'name': (name.group(1).strip() if name else '?')})
        out = _REPORT_OUT_RE.search(text)
        if out and 'total' not in text.split('\n')[0].lower():
            booked_out.append({'at': msg.date, 'amount': _to_float(out.group(1))})
        adj = _REPORT_ADJ_RE.search(text)
        if adj:
            adjustments.append({'at': msg.date, 'column': adj.group(1).title(),
                                'amount': _to_float(adj.group(2))})
        totals = parse_totals(text)
        if totals[0] is not None and totals[1] is not None:
            closing = totals            # newest wins: the day's closing books
    return {'payments': payments, 'booked_out': booked_out,
            'adjustments': adjustments, 'closing': closing}


def _read_handling_side(messages):
    """What the other end of the pair actually saw.

    The notifications as they arrived, the cashout requests the bot posted, and
    the crew's /out answers. This is the independent side of the reconciliation:
    it is read from a different group than the books it is compared against."""
    received, requests, outs = [], [], []
    for msg in messages:
        text = msg.raw_text or getattr(msg, 'message', '') or ''
        caption = getattr(msg, 'caption', None)
        text = text or caption or ''
        if not text:
            continue
        amount = parse_received_amount(text)
        if amount is not None and 'received' in text.lower():
            received.append({'at': msg.date, 'amount': amount})
        if CASHOUT_KEYWORD.lower() in text.lower():
            requests.append({'at': msg.date, 'id': msg.id,
                             'asked': request_amount(text),
                             'text': ' '.join(text.split())[:120],
                             'chases': 0, 'paid_at': None, 'paid': None,
                             'by': None})
        elif _OUT_CMD_RE.search(text):
            outs.append({'at': msg.date, 'amount': out_amount(text),
                         'reply_to': getattr(msg, 'reply_to_msg_id', None)
                         or getattr(getattr(msg, 'reply_to', None),
                                    'reply_to_msg_id', None),
                         'by': _report_sender(msg)})
        elif 'OUT REQUEST HAS CROSSED' in text:
            requests_by_id = {r['id']: r for r in requests}
            replied = (getattr(msg, 'reply_to_msg_id', None)
                       or getattr(getattr(msg, 'reply_to', None),
                                  'reply_to_msg_id', None))
            if replied in requests_by_id:
                requests_by_id[replied]['chases'] += 1
    return {'received': received, 'requests': requests, 'outs': outs}


def _report_sender(msg):
    """Who posted it, in whatever form Telegram cached.

    The numeric id is the fallback rather than '?': it is the one thing that is
    always there, and the escalation DMs already carry ids for the same reason -
    a name nobody recognises is worse than a number that can be looked up."""
    sender = getattr(msg, 'sender', None)
    handle = getattr(sender, 'username', None)
    if handle:
        return '@' + handle
    name = ' '.join(part for part in (getattr(sender, 'first_name', None),
                                      getattr(sender, 'last_name', None)) if part)
    if name:
        return name
    sender_id = getattr(msg, 'sender_id', None) or getattr(sender, 'id', None)
    return f"id:{sender_id}" if sender_id else '?'


def _pair_cashouts(requests, outs):
    """Match each /out to the request it paid - the report's version of
    _match_request(), and deliberately the same rule.

    An explicit reply wins outright. Otherwise the figure decides, because a
    /out settles the request it PAID and not the oldest one; taking queue[0]
    blindly is what mis-marked a live cashout on 2026-08-04. The oldest open
    request is still the fallback for a /out with no figure."""
    for out in sorted(outs, key=lambda o: o['at']):
        open_now = [r for r in requests if r['paid_at'] is None
                    and r['at'] <= out['at']]
        if not open_now:
            continue
        match = next((r for r in open_now if r['id'] == out['reply_to']), None)
        if match is None and out['amount'] is not None:
            match = next((r for r in open_now if r['asked'] == out['amount']), None)
        if match is None:
            match = open_now[0]
        match['paid_at'] = out['at']
        match['paid'] = out['amount']
        match['by'] = out['by']
    return requests


async def build_day_report(client, day):
    """One whole local day. The scheduled report, and /report <date>."""
    start, end = _report_window(day)
    return await build_report(client, start, end, day)


async def build_report(client, start, end, day):
    """Everything both ends of every pair did in the window. Reads only."""
    report = {'day': day, 'pairs': [], 'unreadable': []}

    for pair in report_pairs():
        chime, handling = pair['chime'], pair['handling']
        chime_msgs = await _report_scan(client, chime, start, end)
        handling_msgs = await _report_scan(client, handling, start, end)
        if chime_msgs is None:
            report['unreadable'].append(chat_name(chime))
        if handling_msgs is None:
            report['unreadable'].append(chat_name(handling))

        books = _read_chime_side(chime_msgs or [])
        floor = _read_handling_side(handling_msgs or [])
        cashouts = _pair_cashouts(floor['requests'], floor['outs'])

        booked_in = sum(p['amount'] for p in books['payments'])
        booked_out = sum(o['amount'] for o in books['booked_out'])
        seen_in = sum(r['amount'] for r in floor['received'])
        seen_out = sum(o['amount'] for o in floor['outs']
                       if o['amount'] is not None)

        # THE GAP: what the chime group's books DID, minus what the other end
        # of the pair actually saw. Positive means the books moved more than
        # the money that really arrived - a payment booked twice, or one
        # brought back from the dead by a deploy, which is exactly what
        # happened to CHIME GAFFER on 2026-08-10 (+10.00$). Negative means
        # something the source saw never reached the books at all.
        #
        # Adjustments are deliberately NOT part of this. A retraction is
        # somebody choosing to take money off the books; it is not a
        # disagreement between the two groups, and folding it in here would
        # make every honest retraction read as a hole. They are listed on
        # their own in the Exceptions sheet instead.
        report['pairs'].append({
            'chime': chime, 'handling': handling,
            'payments': books['payments'], 'adjustments': books['adjustments'],
            'cashouts': cashouts, 'closing': books['closing'],
            'booked_in': booked_in, 'booked_out': booked_out,
            'seen_in': seen_in, 'seen_out': seen_out,
            'gap_in': round(booked_in - seen_in, 2),
            'gap_out': round(booked_out - seen_out, 2),
            'readable': chime_msgs is not None and handling_msgs is not None,
        })
    return report


def report_summary_text(report):
    """The same figures as the workbook, for anyone reading on a phone."""
    lines = [f"📊 DAILY REPORT — {report['day']:%d %b %Y}", '']
    for pair in report['pairs']:
        paid = [c for c in pair['cashouts'] if c['paid_at'] is not None]
        open_still = [c for c in pair['cashouts'] if c['paid_at'] is None]
        lines.append(f"{chat_name(pair['chime'])}")
        lines.append(f"  In  {pair['booked_in']:,.2f}$ over {len(pair['payments'])} "
                     f"payment(s)")
        lines.append(f"  Out {pair['booked_out']:,.2f}$ over {len(paid)} cashout(s)")
        if pair['closing']:
            lines.append(f"  Books {pair['closing'][0]:,.2f}$ in / "
                         f"{pair['closing'][1]:,.2f}$ out")
        gap = 'clean' if not (pair['gap_in'] or pair['gap_out']) else (
            f"⚠️ in {pair['gap_in']:+,.2f}$  out {pair['gap_out']:+,.2f}$")
        lines.append(f"  Gap vs {chat_name(pair['handling'])}: {gap}")
        if open_still:
            lines.append(f"  ⏳ {len(open_still)} cashout(s) still unanswered")
        lines.append('')
    if PAUSED_CHAT_IDS:
        lines.append("⏸️ Out of service, not counted here: "
                     + ', '.join(chat_name(c) for c in PAUSED_CHAT_IDS))
    if report['unreadable']:
        lines.append("⚠️ Could not read: " + ', '.join(sorted(set(report['unreadable'])))
                     + " — those figures are missing, not zero.")
    return '\n'.join(lines).strip()


def _report_rows(report):
    """The workbook as plain data, so the sheet layout is testable without a
    spreadsheet library in the room."""
    summary = [('Group', 'Paired with', 'In', 'Out', 'Payments', 'Cashouts paid',
                'Cashouts open', 'Closing In', 'Closing Out', 'Gap In', 'Gap Out')]
    payments = [('Day', 'Time', 'Group', 'From', 'Amount')]
    cashouts = [('Day', 'Asked at', 'Paid at', 'Group', 'Asked', 'Paid',
                 'Waited (min)', 'Chases', 'Paid by', 'Request')]
    exceptions = [('Day', 'Time', 'Group', 'What')]
    stamp = f"{report['day']:%Y-%m-%d}"

    for pair in report['pairs']:
        paid = [c for c in pair['cashouts'] if c['paid_at'] is not None]
        open_still = [c for c in pair['cashouts'] if c['paid_at'] is None]
        closing = pair['closing'] or (None, None)
        summary.append((chat_name(pair['chime']), chat_name(pair['handling']),
                        float(pair['booked_in']), float(pair['booked_out']),
                        len(pair['payments']), len(paid), len(open_still),
                        closing[0], closing[1],
                        float(pair['gap_in']), float(pair['gap_out'])))

        for row in pair['payments']:
            payments.append((stamp, _hhmm(row['at']), chat_name(pair['chime']),
                             row['name'], row['amount']))

        for row in pair['cashouts']:
            waited = (round((row['paid_at'] - row['at']).total_seconds() / 60.0, 1)
                      if row['paid_at'] else None)
            cashouts.append((stamp, _hhmm(row['at']),
                             _hhmm(row['paid_at']) if row['paid_at'] else 'UNPAID',
                             chat_name(pair['handling']), row['asked'], row['paid'],
                             waited, row['chases'], row['by'] or '', row['text']))
            if row['paid_at'] is None:
                asked = (f"{row['asked']:,.2f}$" if row['asked'] is not None
                         else 'an unreadable figure')
                exceptions.append((stamp, _hhmm(row['at']),
                                   chat_name(pair['handling']),
                                   f"Cashout never answered — asked {asked}"))
            elif (row['asked'] is not None and row['paid'] is not None
                    and row['asked'] != row['paid']):
                exceptions.append((stamp, _hhmm(row['paid_at']),
                                   chat_name(pair['handling']),
                                   f"Paid {row['paid']:,.2f} against a request for "
                                   f"{row['asked']:,.2f}"))

        for row in pair['adjustments']:
            exceptions.append((stamp, _hhmm(row['at']), chat_name(pair['chime']),
                               f"{row['column']} adjusted by {row['amount']:,.2f}$ "
                               "(retraction or manual correction)"))

        if pair['gap_in'] or pair['gap_out']:
            exceptions.append((stamp, '', chat_name(pair['chime']),
                               f"GAP against {chat_name(pair['handling'])}: "
                               f"in {pair['gap_in']:+,.2f}$, out {pair['gap_out']:+,.2f}$ "
                               "— the two groups do not agree"))
        if not pair['readable']:
            exceptions.append((stamp, '', chat_name(pair['chime']),
                               "One end of this pair could not be read — these "
                               "figures are incomplete, not zero"))

    # Chronological, not grouped by kind: reading down the sheet should read
    # the day in the order it happened. The header stays put.
    exceptions[1:] = sorted(exceptions[1:], key=lambda row: (row[0], row[1]))
    return {'Summary': summary, 'Payments': payments,
            'Cashouts': cashouts, 'Exceptions': exceptions}


def render_report_file(report):
    """(filename, bytes). A real .xlsx when openpyxl is installed, a CSV when it
    is not - a missing spreadsheet library must not cost the whole report."""
    sheets = _report_rows(report)
    stamp = f"{report['day']:%Y-%m-%d}"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("⚠️ [REPORT] openpyxl missing - sending CSV instead", flush=True)
        lines = []
        for name, rows in sheets.items():
            lines.append(f"# {name}")
            for row in rows:
                lines.append(','.join(
                    '' if cell is None else f'"{cell}"' if isinstance(cell, str)
                    else str(cell) for cell in row))
            lines.append('')
        return f"report-{stamp}.csv", '\n'.join(lines).encode('utf-8')

    book = Workbook()
    book.remove(book.active)
    for name, rows in sheets.items():
        sheet = book.create_sheet(name)
        for row in rows:
            sheet.append(list(row))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        # Money reads as money. Counts are ints and stay plain, so this needs
        # no column map that could drift as sheets change.
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'
        for column in sheet.columns:
            longest = max((len(str(c.value)) for c in column if c.value is not None),
                          default=8)
            sheet.column_dimensions[column[0].column_letter].width = min(longest + 2, 60)
        sheet.freeze_panes = 'A2'

    buffer = io.BytesIO()
    book.save(buffer)
    return f"report-{stamp}.xlsx", buffer.getvalue()


async def send_day_report(client, day, to=None):
    """Build one day and deliver it privately. Never posted in a group."""
    report = await build_day_report(client, day)
    filename, blob = render_report_file(report)
    caption = report_summary_text(report)

    delivered = []
    for chat_id in (to if to is not None else sorted(LEDGER_ADMINS)):
        try:
            await bot.send_document(chat_id, InputFile(io.BytesIO(blob), filename),
                                    caption=caption, visible_file_name=filename)
            delivered.append(chat_id)
        except Exception as e:
            print(f"⚠️ [REPORT] could not send to {chat_id}: {e}", flush=True)
            try:
                await bot.send_message(chat_id, caption)
                delivered.append(chat_id)
            except Exception as inner:
                print(f"❌ [REPORT] {chat_id} got nothing at all: {inner}", flush=True)
    print(f"📊 [REPORT] {day} delivered to {len(delivered)} recipient(s)", flush=True)
    return report, delivered


_REPORT_SPAN_RE = re.compile(r'^(\d+)\s*(h|hr|hrs|hour|hours|d|day|days)$', re.I)


def parse_report_span(argument):
    """'6h', '24h', '3d' -> a timedelta. None when it is not a span at all."""
    match = _REPORT_SPAN_RE.match((argument or '').strip())
    if not match:
        return None
    size = int(match.group(1))
    if size <= 0:
        return None
    unit = match.group(2).lower()
    span = timedelta(days=size) if unit.startswith('d') else timedelta(hours=size)
    # A window nobody can read back is not worth building.
    return span if span <= timedelta(days=31) else None


def group_report_text(books, span_label):
    """Two figures, and nothing else.

    What moved in and what moved out over the window, in the same shape the
    group already sees on every payment. No net, no counts, no list of who
    paid what, no running books - asked for explicitly on 2026-08-10, and the
    reason is sound: this is read at a glance in a busy group, and every extra
    line is something to read past.

    Deliberately NARROW, and narrow by construction rather than by filtering:
    it is handed nothing but this group's own messages, so there is nothing
    else in the room to leak. A chime group is never told who handles its
    cashouts or that a handling group exists at all - that is the oldest rule
    in the flow, and a summary posted in the open is the last place to break
    it. The private workbook is where the routing, the crew and the
    reconciliation live.

    Adjustments are folded into the figures rather than shown on their own
    line. A retraction really did take money back off this group's books, so
    leaving it out would print a Total In the group's own ledger disagrees
    with - and reconciling the two would then need the very detail this shape
    exists to leave out."""
    adjust_in = sum(row['amount'] for row in books['adjustments']
                    if row['column'] == 'In')
    adjust_out = sum(row['amount'] for row in books['adjustments']
                     if row['column'] == 'Out')
    total_in = sum(row['amount'] for row in books['payments']) + adjust_in
    total_out = sum(row['amount'] for row in books['booked_out']) + adjust_out
    return (f"📊 REPORT — last {span_label}\n\n"
            f"➕ Total In : {total_in:,.2f}$\n"
            f"➖ Total Out: {total_out:,.2f}$")


async def send_group_report(client, chime, span, span_label):
    """Post one chime group's own report into that group.

    Only this group is read. Not the handling group, not the other route, not
    the ledger of anywhere else - so what goes out cannot contain them."""
    end = datetime.now(timezone.utc)
    messages = await _report_scan(client, chime, end - span, end)
    if messages is None:
        return None
    books = _read_chime_side(messages)
    await send_group(chime, group_report_text(books, span_label))
    print(f"📊 [REPORT] posted a {span_label} report in {chat_name(chime)}",
          flush=True)
    return books


def _seconds_until_report(now=None):
    """Seconds to the next REPORT_AT in LOCAL time."""
    now = now or datetime.now(timezone.utc)
    local = now.astimezone(LOCAL_TZ)
    try:
        hour, minute = (int(part) for part in REPORT_AT.split(':'))
    except ValueError:
        hour, minute = 0, 0
    target = local.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if target <= local:
        target += timedelta(days=1)
    return (target - local).total_seconds()


async def _wait_for_pending_cashouts():
    """Hold the report while a cashout is still unanswered.

    A day is not finished while somebody still owes a /out: the request was
    asked today and the money is today's, even if the answer lands after
    midnight. So midnight waits for it - but a request nobody EVER answers must
    not hold the report for ever, so the wait is capped and what was still open
    goes into the sheet as an exception rather than delaying it further."""
    waited = 0
    while REPORT_WAIT_MINUTES > 0 and waited < REPORT_WAIT_MINUTES:
        # A request in a group that is out of service is never going to be
        # answered, and holding the whole report for it every night would be
        # the pause quietly costing the groups that ARE being worked.
        open_now = sum(len(queue) for handling, queue in _pending_cashouts.items()
                       if not route_paused(handling))
        if not open_now:
            return waited
        if waited == 0:
            print(f"📊 [REPORT] holding: {open_now} cashout(s) still unanswered "
                  f"(up to {REPORT_WAIT_MINUTES}m)", flush=True)
        await asyncio.sleep(60)
        waited += 1
    return waited


async def daily_report_loop():
    """One workbook a day, at REPORT_AT local time."""
    if not REPORT_ENABLED:
        print("📊 [REPORT] disabled", flush=True)
        return
    print(f"📊 [REPORT] daily workbook at {REPORT_AT} Nepal time, waiting up to "
          f"{REPORT_WAIT_MINUTES}m for an unanswered /out", flush=True)
    while True:
        await asyncio.sleep(_seconds_until_report())
        # The day that has just ENDED. Read before the wait below, so holding
        # for a late /out cannot roll the report onto the next day.
        day = (datetime.now(timezone.utc).astimezone(LOCAL_TZ) - timedelta(minutes=5)).date()
        try:
            await _wait_for_pending_cashouts()
            if _active_client is None:
                print("⚠️ [REPORT] no user account connected - cannot read the "
                      "groups, skipping today", flush=True)
                await notify_admin("⚠️ The daily report could not be built: the user "
                                   "account is not connected, so the groups cannot "
                                   "be read. Use /report once it is back.")
            else:
                await send_day_report(_active_client, day)
        except Exception as e:
            print(f"❌ [REPORT] {day} failed: {e}", flush=True)
            await notify_admin(f"⚠️ The daily report for {day} failed to build: {e}\n\n"
                               "Nothing else is affected — this half of the bot only "
                               "reads. Try /report to rebuild it by hand.")
        await asyncio.sleep(90)         # never fire twice for one midnight


@bot.message_handler(commands=['report'])
async def report_command(message):
    """Two reports, and WHERE it is typed decides which.

    In CHIME PICCASO or CHIME GAFFER, `/report 24h` posts that group's own
    figures into the group, for everyone there to read. It covers a rolling
    window because that is what a group asks - "what have we done today" - and
    it contains nothing but that group's own money.

    Anywhere else it is private, whole days, and the full workbook: every
    group, the cashout turnarounds, who paid them, and the gap between each
    pair. That version names the handling groups and the crew, so it goes to
    Ethan and Larry and never into a chat anyone else can read.

    Ethan and Larry only, in both shapes - the same rule every command that
    shows figures already keeps."""
    chat_id = message.chat.id
    user_id = getattr(message.from_user, 'id', None)
    if chat_paused(chat_id):
        return
    if user_id not in LEDGER_ADMINS:
        # Silent, like the other refusals: in a group, an explanation would
        # advertise a command to people who cannot use it.
        print(f"⛔ [REPORT] /report by {user_id} in {chat_id} denied", flush=True)
        return

    parts = (message.text or '').split()
    argument = parts[1].lower() if len(parts) > 1 else ''

    if _active_client is None:
        await send_group(user_id, "⚠️ The user account is not connected, so the "
                                  "groups cannot be read right now.")
        return

    # In a chime group: that group's own report, posted right there.
    chime = _canonical(chat_id, CASHOUT_ROUTES)
    if chime is not None:
        span = parse_report_span(argument) or timedelta(hours=24)
        label = argument if parse_report_span(argument) else '24h'
        try:
            if await send_group_report(_active_client, chime, span, label) is None:
                await send_group(user_id, f"⚠️ Could not read "
                                          f"{chat_name(chime)} to build that report.")
        except Exception as e:
            print(f"❌ [REPORT] group report in {chat_id} failed: {e}", flush=True)
            # Privately: a failure notice in the group would tell the room the
            # bot is having trouble, which is Ethan's business and nobody else's.
            await send_group(user_id, f"⚠️ Could not post the report in "
                                      f"{chat_name(chime)}: {e}")
        return

    # Anywhere else: the private workbook.
    today = datetime.now(timezone.utc).astimezone(LOCAL_TZ).date()
    span = parse_report_span(argument)
    if argument in ('yesterday', 'y'):
        day = today - timedelta(days=1)
    elif span is not None:
        day = today          # a span in a DM still reports whole days
    elif argument:
        try:
            day = datetime.strptime(argument, '%Y-%m-%d').date()
        except ValueError:
            await send_group(user_id, "Use /report, /report yesterday, "
                                      "/report 2026-08-05, or /report 24h in a "
                                      "chime group.")
            return
    else:
        day = today

    await send_group(user_id, f"📊 Building the report for {day:%d %b %Y}…")
    try:
        await send_day_report(_active_client, day, to=[user_id])
    except Exception as e:
        print(f"❌ [REPORT] /report {day} failed: {e}", flush=True)
        await send_group(user_id, f"⚠️ Could not build that report: {e}")


# Everything the Bot API can hang a caption on. Kept explicit rather than
# "not text": a caption is only ever read for a /out, so widening this list is a
# deliberate act, not something a new Telegram content type does on its own.
CAPTIONED_TYPES = ['photo', 'document', 'video', 'animation']


def media_concerns_cashout(message):
    """Two reasons the cashout flow needs to see a piece of media.

    A /out written as a caption, anywhere - that answers a request. And
    ANYTHING posted in a handling group, caption or not: a screenshot sent
    instead of the /out is how the crew signal they have hit a problem, and
    that has to reach Ethan and Larry rather than sit there looking like
    silence."""
    if is_caption_out(getattr(message, 'caption', None)):
        return True
    return _canonical(getattr(message.chat, 'id', None), _CASHOUT_HANDLERS) is not None


@bot.message_handler(content_types=CAPTIONED_TYPES, func=media_concerns_cashout)
async def cashout_caption(message):
    """Media the cashout flow cares about - see media_concerns_cashout().

    forward_text below is text-only on purpose and never sees this message.
    Only the cashout path runs: process_incoming is deliberately NOT called
    here, so a screenshot can never be booked as a payment notification."""
    await cashout_from_bot_api(message)


# An edit reaches the CASHOUT path and nothing else.
#
# The crew write the message and edit the /out onto it, so without this the
# answer to a cashout is simply never seen: the request stays open, the ❤ never
# lands, the money is never booked and the chase runs on. That was live until
# 2026-08-09 - `edited_message` was not even in allowed_updates.
#
# process_incoming() is deliberately NOT called here, and must never be. A
# payment notification that gets edited would otherwise be forwarded and booked
# a SECOND time, inventing a deposit. Same reasoning as cashout_caption above.
@bot.edited_message_handler(content_types=CAPTIONED_TYPES,
                            func=media_concerns_cashout)
async def cashout_caption_edited(message):
    await cashout_from_bot_api(message, is_edit=True)


@bot.edited_message_handler(content_types=['text'])
async def cashout_text_edited(message):
    await cashout_from_bot_api(message, is_edit=True)


@bot.message_handler(content_types=['text'])
async def forward_text(message):
    sender_is_bot = bool(getattr(message.from_user, 'is_bot', False))
    sent_at = (datetime.fromtimestamp(message.date, tz=timezone.utc)
               if getattr(message, 'date', None) else None)
    await cashout_from_bot_api(message)
    await process_incoming(message.chat.id, message.text, 'bot-api',
                           from_bot=sender_is_bot, sent_at=sent_at,
                           source_msg_id=message.message_id)


# ---------------------------------------------------------------------------
# Telethon listener
# ---------------------------------------------------------------------------

def _parse_source_bots():
    ids, usernames = set(), set()
    for item in SOURCE_BOTS.split(','):
        item = item.strip().lstrip('@')
        if not item:
            continue
        if item.lstrip('-').isdigit():
            ids.add(int(item))
        else:
            usernames.add(item.lower())
    return ids, usernames


def _configured_source_chats():
    if TELETHON_SOURCE_CHATS.strip():
        return [int(c.strip()) for c in TELETHON_SOURCE_CHATS.split(',') if c.strip()]
    # Both ends of a cashout route are watched too: the chime groups because the
    # requests start there, and the handling groups because the /out answering
    # one is a human message the Bot API may never deliver.
    chats = list(FORWARD_RULES.keys())
    for source, route in CASHOUT_ROUTES.items():
        for chat_id in (source, route['handling']):
            if chat_id not in chats:
                chats.append(chat_id)
    return chats


def _is_plain_text(message):
    """Mirrors the Bot API handler's content_types=['text']: text only, no
    photos, videos, documents, stickers or their captions. A message carrying
    only a link preview is still plain text."""
    if message.media is None:
        return True
    return isinstance(message.media, MessageMediaWebPage)


def _is_target_sender(sender, bot_ids, bot_usernames):
    if sender is None:
        return False
    if bot_ids or bot_usernames:
        username = (getattr(sender, 'username', '') or '').lower()
        return sender.id in bot_ids or username in bot_usernames
    # Nothing configured: accept any bot. Humans already arrive via the Bot API.
    return bool(getattr(sender, 'bot', False))


async def _resolve_one(client, chat_id):
    """Try both spellings of a chat id and return whichever resolves."""
    for candidate in [chat_id] + _id_variants(chat_id):
        try:
            return await client.get_entity(candidate)
        except Exception:
            continue
    return None


async def _resolve_source_chats(client, wanted):
    """Resolve each configured chat separately so one bad id cannot stop the
    listener from starting on the good ones."""
    resolved = []
    for chat_id in wanted:
        entity = await _resolve_one(client, chat_id)
        if entity is None:
            print(f"⚠️ [TELETHON] Could not resolve source chat {chat_id} - skipping it.", flush=True)
            continue
        title = getattr(entity, 'title', None) or getattr(entity, 'username', '')
        print(f"👁️ [TELETHON] Watching {chat_id} ({title})", flush=True)
        resolved.append(entity)
    return resolved


def _catchup_signature(text):
    """Content key for one notification that survives forwarding.

    TWO things are rewritten on the way out, and neither can be part of the key.
    The timestamp becomes the real send time. The totals become this target's
    own ledger figures, which is the whole point of the ledger - so they match
    the source only until the books first diverge, and from then on every
    delivered copy looks like a different message. That made the sweep re-send
    the entire window on each connect, and it re-runs on every reconnect, not
    just on boot.

    What is left - the name and the amount - is identical in the source and in
    the copy sitting in the target group."""
    stripped = _STAMP_RE.sub('', text or '')
    stripped = _TOTAL_IN_SUB.sub(r'\1', stripped)
    stripped = _TOTAL_OUT_SUB.sub(r'\1', stripped)
    return ' '.join(stripped.split())


def _retraction_mark(msg):
    """Was this payment taken back by a reaction on it?

    retract_payment() DELETES the forwarded copy out of the target group, and
    the copy being absent is the only thing catch_up() has to go on. A retracted
    payment therefore reads exactly like one that was never delivered, so the
    next deploy re-sent it and re-booked the money - undoing the retraction and
    putting a payment the group had taken back into its books a second time.

    Seen live in CHIME GAFFER on 2026-08-10: two retracted $5 payments both back
    on the books within a minute of a restart, +10.00$ against a ledger that was
    correct before the deploy.

    The reaction that retracted it is still sitting on the original, and that is
    the durable record - the same role the ❤ plays for a cashout request.
    Returns 'admin' when somebody who can retract reacted, 'unconfirmed' when
    Telegram will not say who did, and None when this is not a retraction.

    Only Ethan and Larry can retract, so a stranger's reaction is not one and
    must not hold a real payment back. When Telegram gives no peers at all the
    reaction IS treated as a retraction: a payment held back can be sent by hand
    afterwards, and a double-booking cannot be taken back. Same direction as
    _delivered_signatures()."""
    reactions = getattr(msg, 'reactions', None)
    if reactions is None or not getattr(reactions, 'results', None):
        return None
    peers = []
    for recent in (getattr(reactions, 'recent_reactions', None) or []):
        user_id = getattr(getattr(recent, 'peer_id', None), 'user_id', None)
        if user_id is not None:
            peers.append(user_id)
    if not peers:
        return 'unconfirmed'
    return 'admin' if any(uid in LEDGER_ADMINS for uid in peers) else None


async def _delivered_signatures(client, target):
    """What is already sitting in a target group, and how far back that is known.

    Returns (counts, reached). `reached` is the oldest moment the scan actually
    accounted for; anything older than it cannot be judged and must not be
    guessed at.

    Sender attribution is deliberately NOT used. Telegram credits a post in a
    channel to the channel itself rather than to the bot that sent it, and a
    post by an anonymous admin to the group - so filtering on the bot's own id
    can match nothing at all and make a fully delivered window look entirely
    missing. Content is the honest test: if the text is already in the group it
    has been delivered, whoever Telegram says put it there. A human pasting the
    same text would count too, which errs towards staying quiet - the right
    direction, because a missed payment can still be backfilled by hand and a
    duplicate cannot be taken back."""
    counts = Counter()
    entity = await _resolve_one(client, target)
    if entity is None:
        return None, None               # unreadable: caller must not guess
    cutoff = datetime.now(timezone.utc) - timedelta(minutes=CATCHUP_LOOKBACK_MINUTES)

    scanned = 0
    last_date = None
    covered_window = False
    async for msg in client.iter_messages(entity, limit=CATCHUP_TARGET_SCAN_LIMIT):
        scanned += 1
        last_date = msg.date
        if msg.date < cutoff:
            covered_window = True       # read right past the far edge
            break
        if msg.raw_text:
            counts[_catchup_signature(msg.raw_text)] += 1

    if covered_window or scanned < CATCHUP_TARGET_SCAN_LIMIT:
        # Either we read past the window, or the group simply has no more
        # history - either way the whole window is accounted for.
        reached = cutoff
    else:
        # The scan hit its limit first. Everything older than the last message
        # we saw is unknown, not absent.
        reached = last_date
        print(f"⚠️ [CATCHUP] {target} history is deeper than {CATCHUP_TARGET_SCAN_LIMIT} "
              f"messages; only judging back to {last_date:%m-%d %H:%M}.", flush=True)
    return counts, reached


async def catch_up(client):
    """Forward notifications that arrived while this listener was not running.

    There is a deaf window on every deploy - the start-up hold alone is ~45s -
    and after an outage it can be hours. A payment landing in it would
    otherwise be lost in silence, which is the failure mode that started all
    this.

    Railway wipes the disk on deploy, so there is no local record of what was
    forwarded. The target group's own history is the record: a notification
    whose text is not already sitting in the target was never delivered.

    Copies are counted rather than tested for membership. Timestamps are
    stripped before comparing, so two identical payments minutes apart look
    alike - counting means it forwards the difference instead of deciding one
    of each was enough."""
    if not CATCHUP_LOOKBACK_MINUTES:
        return

    cutoff = datetime.now(timezone.utc) - timedelta(minutes=CATCHUP_LOOKBACK_MINUTES)
    total = 0

    for rule_key, targets in FORWARD_RULES.items():
        if route_paused(rule_key):
            print(f"⏸️ [CATCHUP] {chat_name(rule_key)} is out of service - "
                  "not sweeping it.", flush=True)
            continue
        source = await _resolve_one(client, rule_key)
        if source is None:
            continue

        # A group that has just come back must not be handed the history it
        # deliberately missed. The paused window is not a deaf window: it is
        # time somebody decided the bot would not act on, and sweeping it now
        # would post and BOOK hours of stale payments the moment the group is
        # turned back on. Self-expiring, so nobody has to remember to undo it.
        since = resumed_at(rule_key)
        window_start = max(cutoff, since) if since else cutoff
        if since and since > cutoff:
            print(f"⏪ [CATCHUP] {chat_name(rule_key)} came back at "
                  f"{since:%H:%M} UTC - sweeping only from there.", flush=True)
        # The live handler keys dedup off event.chat_id, so use the same
        # canonical form here or the two would not recognise each other.
        source_id = utils.get_peer_id(source)

        # A reaction only means "retracted" where a reaction retracts. Elsewhere
        # it is somebody acknowledging a payment, and holding that back would
        # lose a real notification.
        retract_here = (rule_key in RETRACT_SOURCES or source_id in RETRACT_SOURCES)

        pending, retracted, unconfirmed = [], 0, 0
        async for msg in client.iter_messages(source, limit=CATCHUP_SCAN_LIMIT):
            if msg.date < window_start:
                break
            if not _is_plain_text(msg) or not msg.raw_text:
                continue
            if not any(k.lower() in msg.raw_text.lower() for k in KEYWORDS):
                continue
            sender = await msg.get_sender()
            if not bool(getattr(sender, 'bot', False)):
                continue        # the live path forwards bot notifications only
            # Checked BEFORE the copies are counted, not after. With the
            # signature counting, a retracted payment and an identical live one
            # are indistinguishable once both are in the list - whichever came
            # first would consume the single remaining copy and the other would
            # be re-sent. Dropping it at the source keeps the count honest.
            mark = _retraction_mark(msg) if retract_here else None
            if mark:
                retracted += 1
                unconfirmed += (mark == 'unconfirmed')
                continue
            pending.append(msg)

        if retracted:
            print(f"↩️ [CATCHUP] {retracted} payment(s) in {chat_name(rule_key)} carry "
                  "a reaction - already retracted, not re-sending them", flush=True)
        if unconfirmed:
            # Telegram would not say who reacted, so this was assumed to be a
            # retraction. If it was not, a real payment is sitting unforwarded.
            await notify_admin(
                f"⚠️ Catch-up left {unconfirmed} payment(s) in {chat_name(rule_key)} "
                "alone: they carry a reaction, but Telegram did not say whose.\n\n"
                "Treated as retracted rather than risk booking them twice. If one "
                "of them was NOT retracted it has not been forwarded — backfill.py "
                "can send it.")
        if not pending:
            continue
        pending.reverse()                       # oldest first, so groups read in order

        for target in targets:
            delivered, reached = await _delivered_signatures(client, target)
            if delivered is None:
                print(f"⚠️ [CATCHUP] {target} unreadable - skipping, will not guess.",
                      flush=True)
                continue

            missing = []
            unjudged = 0
            for msg in pending:
                if reached is not None and msg.date < reached:
                    # Older than the target scan could see. Unknown, not absent.
                    unjudged += 1
                    continue
                sig = _catchup_signature(msg.raw_text)
                if delivered[sig] > 0:
                    delivered[sig] -= 1         # this copy is already in the group
                else:
                    missing.append(msg)
            if unjudged:
                print(f"⚠️ [CATCHUP] {target}: {unjudged} message(s) older than the "
                      "readable history - left alone rather than risk a duplicate.",
                      flush=True)
            if not missing:
                continue

            if len(missing) > CATCHUP_MAX:
                # Something is off - a wrong id, a cleared group. Deliver the
                # newest and say so rather than flooding the group with hours
                # of history.
                print(f"⚠️ [CATCHUP] {target} looks {len(missing)} behind; sending the "
                      f"newest {CATCHUP_MAX}.", flush=True)
                await notify_admin(
                    f"⚠️ Catch-up found {len(missing)} unforwarded notifications for "
                    f"{target}, more than the {CATCHUP_MAX} cap. Sent the newest "
                    f"{CATCHUP_MAX}. Use backfill.py if the rest are wanted.")
                missing = missing[-CATCHUP_MAX:]

            print(f"⏪ [CATCHUP] {target}: {len(missing)} missed notification(s).",
                  flush=True)
            for msg in missing:
                # Mark before sending: the live handler is already queued behind
                # this and must not send the same message a second time.
                _is_duplicate((source_id, msg.id))
                fixed = correct_timestamp(msg.raw_text, msg.date)
                if await deliver_to_target(target, fixed, msg.raw_text, True,
                                           (source_id, msg.id)):
                    total += 1
                await asyncio.sleep(1.5)        # stay under the send rate limit

    if total:
        print(f"⏪ [CATCHUP] delivered {total} missed notification(s).", flush=True)
        await notify_admin(f"⏪ Caught up {total} notification(s) that arrived while "
                           f"the forwarder was not listening.")
    else:
        print("⏪ [CATCHUP] nothing missed.", flush=True)


async def recover_one_ledger(client, target):
    """Restore one target's books from the last totals this bot published there.

    Split out of recover_ledgers so anything that needs the books BEFORE the
    boot sweep has reached them can ask for them. That is not hypothetical: the
    Bot API starts handling updates on its own schedule, with no dependency on
    the userbot having connected, so a reaction can arrive while _ledger is
    still empty."""
    try:
        entity = await client.get_entity(target)
    except Exception as e:
        print(f"⚠️ [LEDGER] {target} unreachable, will open on first message: {e}",
              flush=True)
        return False

    async for message in client.iter_messages(entity, limit=200):
        sender = await message.get_sender()
        if BOT_ID is None or getattr(sender, 'id', None) != BOT_ID:
            continue
        total_in, total_out = parse_totals(message.raw_text or '')
        if total_in is None or total_out is None:
            continue
        _ledger[target] = {'in': total_in, 'out': total_out}
        print(f"📒 [LEDGER] {target} recovered: in={total_in:,.2f} "
              f"out={total_out:,.2f} (from {message.date:%m-%d %H:%M})", flush=True)
        return True

    print(f"📒 [LEDGER] {target} has no prior totals - opens on first message",
          flush=True)
    return False


async def recover_cashout_switch(client):
    """Restore the emergency stop from the last marker the bot published.

    The same principle as recover_ledgers(): Railway wipes the disk on every
    deploy, so a message is the durable record. Without this, /cashout off would
    quietly undo itself on the next push - and the incident that made somebody
    stop the flow is very often the reason a push is coming.

    The marker is private, so it cannot be read the way the ledger is. A bot
    cannot read its own history at all; only a user account can. This works
    because the userbot IS Ethan's or Larry's account, so the bot's DM to them
    is simply one of the userbot's own conversations - opened here by asking for
    the BOT as an entity, which from the user account's side is that chat.

    Failing to read is NOT treated as "running": it leaves the flag alone and
    says so, because guessing wrong in that direction starts moving real
    money."""
    global _cashout_stopped, _cashout_stopped_by

    newest = None
    reachable = False
    # The WHOLE scan is guarded. This runs inside run_userbot()'s reconnect
    # loop, so anything escaping it is retried forever and the userbot never
    # finishes starting - the boot equivalent of the SIGTERM handler that caused
    # the exact failure it was written to prevent.
    try:
        entity = await client.get_entity(BOT_ID)
        async for message in client.iter_messages(entity, limit=100):
            sender = await message.get_sender()
            if BOT_ID is None or getattr(sender, 'id', None) != BOT_ID:
                continue               # our own side of the conversation
            text = message.raw_text or ''
            if CASHOUT_STOP_MARK in text:
                newest = (message.date, True)
            elif CASHOUT_RESUME_MARK in text:
                newest = (message.date, False)
            else:
                continue
            break                      # newest marker wins; history is newest first
        reachable = True
    except Exception as e:
        print(f"⚠️ [SWITCH] could not read the switch back from the private "
              f"chat: {e}", flush=True)

    if not reachable:
        # Neither direction is safe to assume. Guessing "running" could start
        # moving money that was deliberately stopped; guessing "stopped" would
        # take the flow out of service with nobody knowing why. So change
        # nothing and make sure a person is told.
        print("⚠️ [SWITCH] the private chat could not be read - leaving the switch "
              f"as it is ({'STOPPED' if _cashout_stopped else 'running'}).",
              flush=True)
        await warn_unreachable(await dm_handles(
            CASHOUT_ADMIN_HANDLES,
            "⚠️ On starting up, the bot could not read back whether cashout "
            "forwarding was stopped.\n\n"
            f"It is currently {'STOPPED' if _cashout_stopped else 'RUNNING'}, "
            "which is the default and may not be what you left it as.\n\n"
            "Send /cashout to check, and /cashout off if it should be stopped."))
        return

    if newest is None:
        print("▶️ [SWITCH] no marker found - cashout forwarding is running.",
              flush=True)
        return

    _cashout_stopped = newest[1]
    _cashout_stopped_by = None
    print(f"{'⏸️' if newest[1] else '▶️'} [SWITCH] recovered from the groups: "
          f"cashout forwarding is {'STOPPED' if newest[1] else 'running'} "
          f"(marker from {newest[0]:%m-%d %H:%M})", flush=True)
    if newest[1]:
        await warn_unreachable(await dm_handles(
            CASHOUT_ADMIN_HANDLES,
            "⏸️ This bot has just restarted and cashout forwarding is still "
            "STOPPED, as it was before the restart.\n\n"
            "Nothing is being forwarded or chased. Send /cashout on when you "
            "want it back."))


async def recover_group_switch(client):
    """Restore which groups are out of service from the last marker.

    Same principle, same mechanism and same failure rule as
    recover_cashout_switch(): the disk is wiped on every deploy, so the record
    is a private message, and only the userbot can read a private chat.

    Without this, a group taken out of service from a phone would quietly come
    back on the next push - and start forwarding and BOOKING again with nobody
    having asked for it. That is the whole reason this is not a variable
    somebody has to remember to edit.

    The marker also carries the resume marks, which is what stops catch_up()
    below replaying a window nobody was working; see _live_resume_marks()."""
    marker = None
    reachable = False
    # Guarded whole, like the switch above: this runs inside run_userbot()'s
    # reconnect loop, so an exception escaping it is retried for ever and the
    # userbot never finishes starting.
    try:
        entity = await client.get_entity(BOT_ID)
        async for message in client.iter_messages(entity, limit=100):
            sender = await message.get_sender()
            if BOT_ID is None or getattr(sender, 'id', None) != BOT_ID:
                continue                       # our own side of the conversation
            text = message.raw_text or ''
            if GROUP_PAUSE_MARK not in text and GROUP_RESUME_MARK not in text:
                continue
            state = _GROUP_STATE_RE.search(text)
            if state is None:
                continue                       # not a marker we can read
            marker = (message.date, state.group(1),
                      (_GROUP_RESUMED_RE.search(text) or None))
            break                              # newest marker wins
        reachable = True
    except Exception as e:
        print(f"⚠️ [PAUSED] could not read the group switch back from the "
              f"private chat: {e}", flush=True)

    if not reachable:
        # Neither direction is safe to guess. Assuming "in service" could start
        # forwarding money into a group somebody deliberately silenced;
        # assuming "out of service" would silently stop a group that should be
        # running. Change nothing, and make sure a person is told.
        current = (', '.join(chat_name(c) for c in PAUSED_CHAT_IDS)
                   or 'nothing')
        print("⚠️ [PAUSED] the private chat could not be read - leaving the "
              f"switch as it is (out of service: {current}).", flush=True)
        await warn_unreachable(await dm_handles(
            CASHOUT_ADMIN_HANDLES,
            "⚠️ On starting up, the bot could not read back which groups are "
            "out of service.\n\n"
            f"It is currently working everything except: {current}. That is "
            "the default and may not be what you left it as.\n\n"
            "Send /group to check, and /group off to take one out again."))
        return

    if marker is None:
        print("▶️ [PAUSED] no group marker found - keeping the boot state "
              f"({', '.join(chat_name(c) for c in PAUSED_CHAT_IDS) or 'nothing'} "
              "out of service).", flush=True)
        return

    when, state, resumed = marker
    set_paused_chats(_parse_chat_list(state))
    if resumed is not None:
        RESUMED_CHATS.update(_parse_resume_marks(resumed.group(1)))

    out = ', '.join(chat_name(c) for c in PAUSED_CHAT_IDS)
    print(f"{'⏸️' if out else '▶️'} [PAUSED] recovered from the private chat: "
          f"{out or 'nothing'} out of service (marker from {when:%m-%d %H:%M})",
          flush=True)
    if out:
        await warn_unreachable(await dm_handles(
            CASHOUT_ADMIN_HANDLES,
            f"⏸️ This bot has just restarted and {out} "
            f"{'is' if len(PAUSED_CHAT_IDS) == 1 else 'are'} still out of "
            "service, as before the restart.\n\n"
            "Nothing is being forwarded, chased or booked there, and nothing "
            "from that time will be replayed when it comes back.\n\n"
            "Send /group on when you want it working again."))


async def recover_ledgers(client):
    """Rebuild every target's ledger from the last totals this bot published.

    Railway wipes the filesystem on every deploy, so the books cannot live on
    disk. They live in the messages themselves: the bot's own posts carry the
    figures, so reading the newest one back restores the running totals.

    LEDGER_CHATS, not TARGET_CHATS: a group that no longer receives forwards
    keeps its books, and books that are not read back are books that revert on
    the next deploy."""
    for target in sorted(LEDGER_CHATS):
        await recover_one_ledger(client, target)


async def run_userbot():
    global userbot_status

    if not TELETHON_ENABLED:
        userbot_status = 'disabled (no TELEGRAM_API_ID / TELEGRAM_API_HASH)'
        print("ℹ️ [TELETHON] Disabled - set TELEGRAM_API_ID and TELEGRAM_API_HASH to "
              "receive messages from other bots.", flush=True)
        return

    bot_ids, bot_usernames = _parse_source_bots()
    session = StringSession(TELETHON_SESSION) if TELETHON_SESSION else TELETHON_SESSION_NAME
    warned_unauthorised = False

    if TELETHON_START_DELAY:
        # Only before the first connect. Reconnects inside the loop below are
        # replacing a dropped connection of our own, not racing another
        # container, so they must not be delayed.
        userbot_status = f'waiting {TELETHON_START_DELAY}s for the previous deploy to exit'
        print(f"⏸️ [TELETHON] Holding {TELETHON_START_DELAY}s so the outgoing container "
              "releases the session first (set TELETHON_START_DELAY=0 to disable).",
              flush=True)
        await asyncio.sleep(TELETHON_START_DELAY)

    global _active_client

    while True:
        client = TelegramClient(session, int(API_ID), API_HASH)
        _active_client = client
        try:
            # connect() instead of start() so a missing session fails loudly
            # rather than blocking on an interactive login prompt in production.
            await client.connect()
            if not await client.is_user_authorized():
                userbot_status = 'not authorised'
                print("❌ [TELETHON] Session is not authorised. Regenerate it with: "
                      "python3 telethon_login.py --deploy", flush=True)
                if not warned_unauthorised:
                    warned_unauthorised = True
                    await notify_admin(
                        "⚠️ Telethon session is NOT authorised - messages from bots are "
                        "not being forwarded. Regenerate TELETHON_SESSION with:\n"
                        "python3 telethon_login.py --deploy\n\n"
                        "Note: the deployment needs its own session. Sharing one auth key "
                        "between Railway and your Mac makes Telegram revoke it.")
                await asyncio.sleep(300)
                continue

            me = await client.get_me()
            print(f"🔐 [TELETHON] Logged in as {me.first_name} (@{me.username})", flush=True)
            global _userbot_username
            _userbot_username = (me.username or '').lower()

            # Warms the entity cache so numeric chat ids resolve reliably.
            await client.get_dialogs()

            chats = await _resolve_source_chats(client, _configured_source_chats())
            if not chats:
                # Retry rather than return: a transient resolve failure must not
                # silently retire the listener for the life of the process.
                userbot_status = 'no source chats resolved'
                print("❌ [TELETHON] No source chats could be resolved - retrying in 60s.", flush=True)
                await asyncio.sleep(60)
                continue

            # Register the handler BEFORE the ledger rebuild and the catch-up
            # sweep, so a payment arriving during them is queued rather than
            # missed. Each event waits on this gate, which opens once the books
            # are straight and the sweep has run - that ordering keeps the
            # groups reading chronologically and lets dedup do its job.
            ready = asyncio.Event()

            @client.on(events.NewMessage(chats=chats))
            async def on_source_message(event):
                await ready.wait()
                sender = await event.get_sender()
                text = event.raw_text or ''
                plain = _is_plain_text(event.message)
                in_handling = _canonical(event.chat_id, _CASHOUT_HANDLERS) is not None

                # Media is dropped as it always was, with two exceptions: a /out
                # written on the payment screenshot answers a request, and
                # anything at all in a handling group may be the crew saying
                # they are stuck. Both reach the cashout path below and stop
                # there - the payment side stays strictly plain text.
                if not plain and not (is_caption_out(text) or in_handling):
                    return
                if not text and plain:
                    return
                remember_user(sender)

                sender_id = getattr(sender, 'id', None)
                if BOT_ID is not None and sender_id == BOT_ID:
                    # Our own forwards now land in watched chats - the handling
                    # groups are also payment sources. Acting on them would loop.
                    return

                # Cashout first, and it accepts humans as well as bots: a request
                # or the /out answering one is as likely to be typed by a person.
                # It never consumes the message - the two keyword filters cannot
                # both match, so the payment path still gets its turn below.
                full_name = ' '.join(
                    part for part in (getattr(sender, 'first_name', None),
                                      getattr(sender, 'last_name', None)) if part)
                await observe_cashout(event.chat_id, text, event.id,
                                      event.message.date, sender_id,
                                      getattr(sender, 'username', None),
                                      getattr(event.message, 'reply_to_msg_id', None),
                                      full_name or None,
                                      has_media=not plain,
                                      media_group_id=getattr(event.message,
                                                             'grouped_id', None))
                await observe_mentions(event.chat_id, text, event.id,
                                       event.message.date, sender_id,
                                       getattr(sender, 'username', None),
                                       full_name or None)

                if not plain:
                    return          # a caption is never a payment notification
                if not _is_target_sender(sender, bot_ids, bot_usernames):
                    return
                if _is_duplicate((event.chat_id, event.id)):
                    return
                # _is_target_sender above already guaranteed a bot sender.
                await process_incoming(event.chat_id, text, 'telethon',
                                       from_bot=True, sent_at=event.message.date,
                                       source_msg_id=event.id)

            @client.on(events.MessageEdited(chats=chats))
            async def on_source_edited(event):
                """An edit, on the cashout path ONLY.

                The crew write the message and edit the /out onto it, so this is
                how a great many cashouts are actually answered. The payment path
                is deliberately not run: re-reading an edited "You received"
                would forward and book the same deposit twice."""
                await ready.wait()
                sender = await event.get_sender()
                text = event.raw_text or ''
                plain = _is_plain_text(event.message)
                in_handling = _canonical(event.chat_id, _CASHOUT_HANDLERS) is not None

                if not plain and not (is_caption_out(text) or in_handling):
                    return
                if not text and plain:
                    return
                remember_user(sender)

                sender_id = getattr(sender, 'id', None)
                if BOT_ID is not None and sender_id == BOT_ID:
                    return
                full_name = ' '.join(
                    part for part in (getattr(sender, 'first_name', None),
                                      getattr(sender, 'last_name', None)) if part)
                await observe_cashout(event.chat_id, text, event.id,
                                      event.message.date, sender_id,
                                      getattr(sender, 'username', None),
                                      getattr(event.message, 'reply_to_msg_id', None),
                                      full_name or None,
                                      has_media=not plain, is_edit=True,
                                      media_group_id=getattr(event.message,
                                                             'grouped_id', None))

            @client.on(events.MessageDeleted)
            async def on_message_deleted(event):
                await ready.wait()
                await close_deleted_cashouts(client, event.chat_id, event.deleted_ids)

            # Before the ledgers: if the flow was stopped, that has to be true
            # again the moment this container starts handling anything.
            await recover_cashout_switch(client)
            # Before the sweep, and for the same reason twice over: a group
            # left out of service must not be swept, and one that came back
            # must not have the window it missed replayed into it.
            await recover_group_switch(client)
            await recover_ledgers(client)
            try:
                await catch_up(client)
            except Exception as e:
                # A failed sweep must not cost us the live listener - being late
                # on a few messages beats being deaf to all of them.
                print(f"⚠️ [CATCHUP] sweep failed: {e} - continuing to listen.",
                      flush=True)
            ready.set()

            target = SOURCE_BOTS if SOURCE_BOTS.strip() else 'any bot'
            userbot_status = f'listening ({len(chats)} chats, from: {target})'
            print(f"✅ [TELETHON] Listening for messages from: {target}", flush=True)

            await client.run_until_disconnected()
            print("⚠️ [TELETHON] Disconnected - reconnecting in 15s.", flush=True)
        except (AuthKeyDuplicatedError, AuthKeyUnregisteredError,
                SessionRevokedError, UserDeactivatedBanError) as e:
            # Telegram has destroyed the auth key - almost always because the
            # same session ran from two IPs at once (deployment + laptop, or two
            # overlapping Railway containers). Reconnecting can NEVER fix this:
            # only a fresh login can. Retrying silently every 15s is how this
            # went unnoticed for hours, so shout once and stop.
            userbot_status = f'session dead: {type(e).__name__}'
            print(f"💀 [TELETHON DEAD] {type(e).__name__}: {e}\n"
                  "   Forwarding of bot messages is STOPPED until TELETHON_SESSION "
                  "is replaced. Run: python3 telethon_login.py --deploy", flush=True)
            await notify_admin(
                f"🚨 FORWARDING IS DOWN - Telethon session destroyed "
                f"({type(e).__name__}).\n\n"
                "Messages posted by the notification bot are NOT being forwarded "
                "to any group, and this cannot self-heal.\n\n"
                "Fix:\n"
                "1. python3 telethon_login.py --deploy\n"
                "2. Paste the new string into Railway's TELETHON_SESSION\n"
                "3. Redeploy\n\n"
                "Cause: that session key was used from two IPs at once. Never run "
                "local scripts with the deployment's session, and keep Railway at "
                "a single replica.")
            return
        except Exception as e:
            userbot_status = f'error: {e}'
            print(f"❌ [TELETHON ERROR] {e} - retrying in 15s.", flush=True)
        finally:
            try:
                await client.disconnect()
            except Exception:
                pass

        await asyncio.sleep(15)


# ---------------------------------------------------------------------------
# Entry point - one event loop, both listeners
# ---------------------------------------------------------------------------

async def run_bot():
    # Before the first poll only. Nothing is missed - see BOT_START_DELAY.
    if BOT_START_DELAY:
        print(f"⏸️ [BOT] Holding {BOT_START_DELAY}s before polling so the outgoing "
              "container releases the token first. Updates queue at Telegram and "
              "arrive when polling starts (set BOT_START_DELAY=0 to disable).",
              flush=True)
        await asyncio.sleep(BOT_START_DELAY)

    while True:
        try:
            # message_reaction is NOT included by default and has to be asked
            # for by name, or a crew reaction never reaches on_request_reaction.
            await bot.polling(none_stop=True,
                              allowed_updates=['message', 'edited_message',
                                               'message_reaction'])
        except Exception as e:
            print(f"Polling Error: {e}", flush=True)
            await asyncio.sleep(5)


def _schedule_disconnect(client):
    """Start releasing the Telethon session, whatever disconnect() hands back.

    Telethon returns a Future here rather than a coroutine, and
    `loop.create_task()` accepts ONLY a coroutine - so the old call raised
    TypeError, took the whole SIGTERM callback down with it, and the session was
    never released. `ensure_future()` takes either. Some versions return None
    when already disconnected, which is not an error.

    Everything is caught: this runs on the way out, and nothing it can fail at
    is worth staying alive for."""
    if client is None:
        return None
    try:
        pending = client.disconnect()
    except Exception as e:
        print(f"⚠️ [SHUTDOWN] disconnect() failed: {e}", flush=True)
        return None
    if pending is None:
        return None                     # already down; nothing to wait on
    try:
        return asyncio.ensure_future(pending)
    except Exception as e:
        print(f"⚠️ [SHUTDOWN] could not await the disconnect: {e}", flush=True)
        return None


def _install_shutdown_handler(loop):
    """Leave promptly and quietly when Railway says we are going away.

    This matters more than it looks. In a container this process is PID 1, and
    PID 1 does not get the kernel's default signal dispositions: an unhandled
    SIGTERM is simply IGNORED. So the outgoing container sat there holding the
    Telethon connection until Railway lost patience and sent SIGKILL, which is
    the overlap that destroyed the session. Handling SIGTERM explicitly is what
    makes the old container actually leave.

    Exit via os._exit rather than loop.stop(): stopping the loop out from under
    asyncio.run() raises RuntimeError and exits non-zero, which Railway reads as
    a crash and restarts - turning a clean shutdown into another instance."""
    def _on_term():
        print("🛑 [SHUTDOWN] SIGTERM - releasing the Telethon session before exit.",
              flush=True)
        _schedule_disconnect(_active_client)
        # Scheduled LAST, and deliberately outside everything that can raise. If
        # releasing the session fails we still have to leave: a SIGTERM callback
        # that dies before reaching this line leaves the container running until
        # Railway loses patience and SIGKILLs it, which IS the overlap this
        # handler exists to prevent. That is exactly what happened on 2026-08-03.
        loop.call_later(2, lambda: os._exit(0))

    for sig in (signal.SIGTERM, signal.SIGINT):
        try:
            loop.add_signal_handler(sig, _on_term)
        except (NotImplementedError, RuntimeError):
            pass        # not supported on this platform; default behaviour applies

    return _on_term         # returned so the shutdown path can be tested


async def main():
    print('Bot running...', flush=True)
    _install_shutdown_handler(asyncio.get_running_loop())
    # The log only. Which groups are out of service is not settled yet - the
    # marker in the private chat is, and recover_group_switch() is what reads
    # it back and tells Ethan and Larry. Saying it here as well would announce
    # the boot default, which is exactly the thing that may be out of date.
    out_of_service = ', '.join(chat_name(c) for c in PAUSED_CHAT_IDS)
    if out_of_service:
        print(f'⏸️ [PAUSED] booting with {out_of_service} out of service, '
              'pending the marker', flush=True)
    await notify_admin('Bot is ONLINE. Use /help to see all commands.')
    await asyncio.gather(run_bot(), run_userbot(), idle_watchdog(),
                         cashout_watchdog(), daily_report_loop(),
                         return_exceptions=True)


if __name__ == '__main__':
    asyncio.run(main())
