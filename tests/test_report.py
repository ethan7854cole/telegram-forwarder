"""The daily report: rebuilt from the groups, reconciling both ends of a pair.

Two shapes are tested, and the difference between them is the point. The
private workbook carries every group, the crew and the routing. The one posted
INSIDE a chime group carries that group's own money and nothing else.

Stubbed bot and Telethon client - nothing here touches Telegram or the network.
"""
import asyncio
import os
import sys
from datetime import datetime, timedelta, timezone

os.environ['TELEGRAM_BOT_TOKEN'] = '111222:FAKE'
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import forwarder as f

PICCASO, GAFFER = -5350880041, -5580596463
MHLARRY, CHIMEREV = -1003894781195, -1002335630148
LARRY, ETHAN, CREW = 7418675217, 7578145913, 555

sent, documents, failures = [], [], []
DAY = datetime(2026, 8, 10, tzinfo=f.LOCAL_TZ).date()


def at(hour, minute=0):
    """A moment on the reported day, given in NEPAL time."""
    return datetime(2026, 8, 10, hour, minute, tzinfo=f.LOCAL_TZ).astimezone(timezone.utc)


class Sender:
    def __init__(self, username=None, first_name=None):
        self.username, self.first_name, self.last_name = username, first_name, None


class Msg:
    def __init__(self, mid, text, when, reply_to=None, sender=None):
        self.id, self.raw_text, self.date = mid, text, when
        self.reply_to_msg_id, self.sender = reply_to, sender
        self.media = None


class FakeEntity:
    def __init__(self, cid): self.id, self.title = cid, str(cid)


class FakeClient:
    def __init__(self, history): self.history = history
    async def get_entity(self, cid):
        if cid in self.history: return FakeEntity(cid)
        raise ValueError('no entity %s' % cid)
    def iter_messages(self, entity, limit=None):
        async def gen():
            for m in sorted(self.history.get(entity.id, []),
                            key=lambda m: m.date, reverse=True):   # newest first
                yield m
        return gen()


class FakeBot:
    async def send_message(self, chat_id, text, **kwargs):
        sent.append((chat_id, text))
        return type('M', (), {'message_id': 1})()
    async def send_document(self, chat_id, document, caption=None, **kwargs):
        documents.append((chat_id, kwargs.get('visible_file_name'), caption))
        return type('M', (), {'message_id': 2})()


f.bot = FakeBot()
f.USERBOT_SEND = False


def check(label, cond, detail=''):
    print(('  ok   ' if cond else '  FAIL ') + label
          + (f'  <- {detail}' if detail and not cond else ''))
    if not cond:
        failures.append(label)


def reset():
    sent.clear(); documents.clear()


def payment(amount, name, total_in, total_out):
    """What the bot posts in a chime group when it forwards a payment."""
    return (f"🟢 Hi $Oscar-Gonzalez-472,\n\nYou received ${amount} from {name}\n\n"
            f"10:45 AM - 10 Aug 2026\n"
            f"➕ Total In : {total_in:,.2f}$\n➖ Total Out: {total_out:,.2f}$")


def booked_out(amount, total_in, total_out):
    """What book_cashout_out() posts."""
    return (f"📤 Out = -{amount:,.2f}$\n\n📊 Group Total:\n"
            f"➕ Total In : {total_in:,.2f}$\n➖ Total Out: {total_out:,.2f}$")


def adjusted(amount, total_in, total_out):
    return (f"✏️ Total In adjusted by {amount:,.2f}$\n\n📊 Group Total:\n"
            f"➕ Total In : {total_in:,.2f}$\n➖ Total Out: {total_out:,.2f}$")


async def main():
    # -- the pairs come off the live tables --------------------------------
    pairs = {p['chime']: p for p in f.report_pairs()}
    check('both routes produce a pair', len(pairs) == 2, str(pairs))
    check('CHIME GAFFER is paired with Chime Rev & out no-7',
          pairs[GAFFER]['handling'] == CHIMEREV, str(pairs.get(GAFFER)))
    check('CHIME PICCASO is paired with MH X LARRY GROUP 2',
          pairs[PICCASO]['handling'] == MHLARRY, str(pairs.get(PICCASO)))
    check('both pairs really are two ends of one route',
          all(p['paired'] for p in pairs.values()), str(pairs))

    # -- a day is a NEPAL day, not a UTC one -------------------------------
    start, end = f._report_window(DAY)
    check('the window is exactly 24 hours', end - start == timedelta(days=1))
    check('it opens at midnight Nepal time',
          start.astimezone(f.LOCAL_TZ).hour == 0
          and start.astimezone(f.LOCAL_TZ).minute == 0, str(start))
    check('a 23:50 Nepal payment falls inside the day',
          start <= at(23, 50) < end, str(at(23, 50)))
    check('...and 00:10 the next morning does not',
          not (start <= at(23, 50) + timedelta(minutes=20) < end))

    # -- a clean day, both ends agreeing ------------------------------------
    # GAFFER's books: two payments in, one cashout out.
    gaffer = [Msg(901, payment(5.0, 'Clarence B.', 12151.64, 7026.0), at(10, 45)),
              Msg(902, payment(5.0, 'Clarence B.', 12156.64, 7026.0), at(11, 13)),
              Msg(903, booked_out(200.0, 12156.64, 7226.0), at(15, 20))]
    # Chime Rev saw the same money: the two notifications and the /out.
    chimerev = [Msg(801, payment(5.0, 'Clarence B.', 0, 0), at(10, 45)),
                Msg(802, payment(5.0, 'Clarence B.', 0, 0), at(11, 13)),
                Msg(803, f"{f.CASHOUT_KEYWORD} $200 for Gabriel W.", at(15, 4)),
                Msg(804, '/out 200', at(15, 20), reply_to=803,
                    sender=Sender(username='maynuddin233'))]
    client = FakeClient({GAFFER: gaffer, CHIMEREV: chimerev,
                         PICCASO: [], MHLARRY: []})

    before_ledger = dict(f._ledger)
    report = await f.build_day_report(client, DAY)
    pair = next(p for p in report['pairs'] if p['chime'] == GAFFER)

    check('the payments are read back', len(pair['payments']) == 2,
          str(pair['payments']))
    check('and their total', pair['booked_in'] == 10.0, str(pair['booked_in']))
    check('the cashout is read back', pair['booked_out'] == 200.0,
          str(pair['booked_out']))
    check('the closing books are the newest totals in the group',
          pair['closing'] == (12156.64, 7226.0), str(pair['closing']))
    check('a clean day has no gap in',
          pair['gap_in'] == 0.0, str(pair['gap_in']))
    check('a clean day has no gap out',
          pair['gap_out'] == 0.0, str(pair['gap_out']))
    check('building a report never touches the ledger', f._ledger == before_ledger)
    check('and never posts anything anywhere', sent == [], str(sent))

    # -- a human pasting a notification never moved the books ---------------
    # deliver_to_target() books an amount only `if from_bot`, so counting a
    # paste would invent money and report a gap that is not there.
    pasted = gaffer + [Msg(910, payment(500.0, 'Nobody', 12156.64, 7226.0),
                           at(16, 0), sender=Sender(username='someone'))]
    pasted[-1].sender_id = 999
    client_pasted = FakeClient({GAFFER: pasted, CHIMEREV: chimerev,
                                PICCASO: [], MHLARRY: []})
    pasted_report = await f.build_day_report(client_pasted, DAY)
    pasted_pair = next(p for p in pasted_report['pairs'] if p['chime'] == GAFFER)
    check('a human paste is not counted as booked money',
          pasted_pair['booked_in'] == 10.0, str(pasted_pair['booked_in']))
    check('and so it raises no false gap', pasted_pair['gap_in'] == 0.0,
          str(pasted_pair['gap_in']))

    # ...but the bot's own messages still count when Telegram credits the
    # channel instead of the bot, which is why unknown authorship counts.
    anonymous = [Msg(911, payment(5.0, 'Clarence B.', 12151.64, 7026.0), at(10, 45))]
    anonymous[0].sender_id = None
    check('a post Telegram credits to nobody still counts',
          f._read_chime_side(anonymous)['payments'] != [], 'dropped')

    # -- the cashout is matched to the request it paid ----------------------
    cashout = pair['cashouts'][0]
    check('the request is paired with its /out', cashout['paid'] == 200.0,
          str(cashout))
    check('the wait is measured', cashout['paid_at'] == at(15, 20), str(cashout))
    check('and who paid it is recorded', cashout['by'] == '@maynuddin233',
          str(cashout))

    # -- a retraction is not a gap ------------------------------------------
    # The money was deliberately taken back, so both ends still agree.
    gaffer_retracted = gaffer + [Msg(904, adjusted(-5.0, 12151.64, 7226.0), at(12, 15))]
    client = FakeClient({GAFFER: gaffer_retracted, CHIMEREV: chimerev,
                         PICCASO: [], MHLARRY: []})
    report = await f.build_day_report(client, DAY)
    pair = next(p for p in report['pairs'] if p['chime'] == GAFFER)
    check('an honest retraction leaves no gap', pair['gap_in'] == 0.0,
          str(pair['gap_in']))
    check('but it is still listed as an exception',
          any('adjusted' in str(row) for row in f._report_rows(report)['Exceptions']),
          str(f._report_rows(report)['Exceptions']))

    # -- THE 2026-08-10 INCIDENT: a resurrected payment shows as a gap ------
    # Two $5 retracted, then a deploy re-sent and re-booked both. Chime Rev
    # saw 10.00$ of real money; GAFFER's books moved by 20.00$ minus the
    # 10.00$ taken back. The pair no longer agrees, and the report says so.
    resurrected = gaffer + [
        Msg(904, adjusted(-5.0, 12151.64, 7026.0), at(12, 15)),
        Msg(905, adjusted(-5.0, 12146.64, 7026.0), at(12, 15)),
        Msg(906, payment(5.0, 'Clarence B.', 12151.64, 7026.0), at(12, 25)),
        Msg(907, payment(5.0, 'Clarence B.', 12156.64, 7026.0), at(12, 25))]
    client = FakeClient({GAFFER: resurrected, CHIMEREV: chimerev,
                         PICCASO: [], MHLARRY: []})
    report = await f.build_day_report(client, DAY)
    pair = next(p for p in report['pairs'] if p['chime'] == GAFFER)
    # Positive: the books moved MORE than the money that really arrived.
    check('a resurrected payment shows up as a gap', pair['gap_in'] == 10.0,
          str(pair['gap_in']))
    check('the gap is called out in the exceptions',
          any('GAP' in str(row) for row in f._report_rows(report)['Exceptions']),
          str(f._report_rows(report)['Exceptions']))
    check('and it is in the summary text',
          '⚠️' in f.report_summary_text(report), f.report_summary_text(report))

    # -- a missed payment shows the other way round -------------------------
    # Chime Rev saw it; GAFFER never booked it.
    client = FakeClient({GAFFER: gaffer[:1], CHIMEREV: chimerev,
                         PICCASO: [], MHLARRY: []})
    report = await f.build_day_report(client, DAY)
    pair = next(p for p in report['pairs'] if p['chime'] == GAFFER)
    # Negative: the source saw money the books never took.
    check('a payment that never reached the books is a negative gap',
          pair['gap_in'] == -5.0, str(pair['gap_in']))
    check('and an unbooked /out is a negative gap out',
          pair['gap_out'] == -200.0, str(pair['gap_out']))

    # -- an unanswered cashout is an exception, not a silence ---------------
    unanswered = [Msg(803, f"{f.CASHOUT_KEYWORD} $200 for Gabriel W.", at(23, 40))]
    client = FakeClient({GAFFER: [], CHIMEREV: unanswered,
                         PICCASO: [], MHLARRY: []})
    report = await f.build_day_report(client, DAY)
    rows = f._report_rows(report)
    check('an unanswered cashout is flagged',
          any('never answered' in str(row) for row in rows['Exceptions']),
          str(rows['Exceptions']))
    check('and shows as UNPAID in the cashout sheet',
          any('UNPAID' in str(row) for row in rows['Cashouts']), str(rows['Cashouts']))

    # -- an unreadable group is unknown, never zero -------------------------
    client = FakeClient({CHIMEREV: chimerev})          # GAFFER missing entirely
    report = await f.build_day_report(client, DAY)
    check('an unreadable group is named', report['unreadable'] != [],
          str(report['unreadable']))
    check('and the summary says the figures are missing, not zero',
          'not zero' in f.report_summary_text(report), f.report_summary_text(report))

    # ------------------------------------------------------------------
    # The report posted INSIDE a chime group
    # ------------------------------------------------------------------
    check('6h parses', f.parse_report_span('6h') == timedelta(hours=6))
    check('24h parses', f.parse_report_span('24h') == timedelta(hours=24))
    check('3d parses', f.parse_report_span('3d') == timedelta(days=3))
    check('a date is not a span', f.parse_report_span('2026-08-05') is None)
    check('nonsense is not a span', f.parse_report_span('soon') is None)
    check('an absurd window is refused', f.parse_report_span('900h') is None)

    reset()
    now = datetime.now(timezone.utc)
    recent = [Msg(901, payment(5.0, 'Clarence B.', 12151.64, 7026.0),
                  now - timedelta(hours=2)),
              Msg(903, booked_out(200.0, 12151.64, 7226.0), now - timedelta(hours=1))]
    client = FakeClient({GAFFER: recent, CHIMEREV: chimerev,
                         PICCASO: [], MHLARRY: []})
    await f.send_group_report(client, GAFFER, timedelta(hours=24), '24h')
    posted = [t for c, t in sent if c == GAFFER]
    check('the report is posted in the group itself', len(posted) == 1, str(sent))
    body = posted[0] if posted else ''
    check('it carries Total In', 'Total In : 5.00$' in body, body)
    check('it carries Total Out', 'Total Out: 200.00$' in body, body)
    check('it says which window it covers', 'last 24h' in body, body)

    # Two figures and nothing else - asked for on 2026-08-10.
    check('no net line', 'Net' not in body, body)
    check('no payment-by-payment list', 'Clarence' not in body, body)
    check('no counts', 'payment(s)' not in body, body)
    check('it is four lines long', len(body.strip().split(chr(10))) == 4, body)

    # A retraction really did take money back off this group's books, so the
    # figure has to agree with the group's own ledger rather than ignore it.
    retracted_books = f._read_chime_side(
        recent + [Msg(904, adjusted(-2.0, 12149.64, 7226.0), now - timedelta(minutes=30))])
    folded = f.group_report_text(retracted_books, '24h')
    check('a retraction is folded into Total In', 'Total In : 3.00$' in folded, folded)
    check('and gets no line of its own', 'Adjusted' not in folded, folded)

    # The whole point of the narrow shape: nothing about anywhere else.
    check('it never names the handling group',
          f.chat_name(CHIMEREV) not in body and 'Chime Rev' not in body, body)
    check('it never names the other route',
          f.chat_name(PICCASO) not in body and 'PICCASO' not in body.upper(), body)
    check('it shows no gap in the open', 'GAP' not in body.upper(), body)
    check('and no crew handle',
          not any(h.lower() in body.lower() for h in f.CASHOUT_CREW_HANDLES), body)
    check('nothing was posted in any other group',
          [c for c, _ in sent if c != GAFFER] == [], str(sent))

    # -- /report in a chime group answers in the group ----------------------
    def fake_message(chat_id, user_id, text):
        return type('Message', (), {
            'chat': type('C', (), {'id': chat_id})(),
            'from_user': type('U', (), {'id': user_id, 'username': 'ethannxxxx'})(),
            'text': text})()

    f._active_client = client
    try:
        reset()
        await f.report_command(fake_message(GAFFER, ETHAN, '/report 6h'))
        check('/report in a chime group posts there',
              [c for c, _ in sent if c == GAFFER] != [], str(sent))
        check('and nothing is DMed for it',
              [c for c, _ in sent if c > 0] == [], str(sent))

        # -- a crew member cannot run it ------------------------------------
        reset()
        await f.report_command(fake_message(GAFFER, CREW, '/report 6h'))
        check('the crew cannot run /report', sent == [], str(sent))

        # -- in a HANDLING group it stays private ---------------------------
        # The crew are in there, and the workbook names groups and people.
        reset()
        await f.report_command(fake_message(CHIMEREV, ETHAN, '/report'))
        check('/report in a handling group posts nothing there',
              [c for c, _ in sent if c == CHIMEREV] == [], str(sent))
        check('it answers privately instead',
              [c for c, _ in sent if c == ETHAN] != [], str(sent))
        check('and the workbook goes to that admin only',
              all(c == ETHAN for c, _, _ in documents), str(documents))

        # -- a DM gets the full workbook ------------------------------------
        reset()
        await f.report_command(fake_message(ETHAN, ETHAN, '/report'))
        check('a DM gets a document', len(documents) == 1, str(documents))
        check('named for the day',
              documents and documents[0][1].startswith('report-'), str(documents))
        caption = documents[0][2] if documents else ''
        check('the caption carries the gap line', 'Gap vs' in caption, caption)
    finally:
        f._active_client = None

    # -- the workbook really is a workbook ----------------------------------
    client = FakeClient({GAFFER: gaffer, CHIMEREV: chimerev,
                         PICCASO: [], MHLARRY: []})
    report = await f.build_day_report(client, DAY)
    name, blob = f.render_report_file(report)
    check('an .xlsx is produced when openpyxl is installed',
          name.endswith('.xlsx'), name)
    check('and it is a real zip container', blob[:2] == b'PK', str(blob[:8]))
    try:
        import openpyxl
        book = openpyxl.load_workbook(io_bytes := __import__('io').BytesIO(blob))
        check('it has the four sheets',
              book.sheetnames == ['Summary', 'Payments', 'Cashouts', 'Exceptions'],
              str(book.sheetnames))
        check('the summary has a row per group',
              book['Summary'].max_row == 3, str(book['Summary'].max_row))
    except ImportError:
        print('  ..   openpyxl not installed here, skipping the workbook read')

    # -- and it degrades rather than failing --------------------------------
    real_import = __builtins__.__import__ if hasattr(__builtins__, '__import__') \
        else __builtins__['__import__']

    def no_openpyxl(name, *args, **kwargs):
        if name.startswith('openpyxl'):
            raise ImportError('no openpyxl here')
        return real_import(name, *args, **kwargs)

    if hasattr(__builtins__, '__import__'):
        __builtins__.__import__ = no_openpyxl
    else:
        __builtins__['__import__'] = no_openpyxl
    try:
        name, blob = f.render_report_file(report)
        check('without openpyxl it falls back to CSV', name.endswith('.csv'), name)
        check('and the CSV still carries the figures', b'Clarence B.' in blob,
              str(blob[:120]))
    finally:
        if hasattr(__builtins__, '__import__'):
            __builtins__.__import__ = real_import
        else:
            __builtins__['__import__'] = real_import

    # -- midnight waits for an unanswered /out ------------------------------
    f._pending_cashouts.clear()
    f._pending_cashouts[CHIMEREV] = [{'text': 'CASHOUT REQUEST $200'}]
    real_sleep = asyncio.sleep

    async def instant(_):
        # The wait is a minute a round; drain the queue on the third.
        instant.rounds += 1
        if instant.rounds >= 3:
            f._pending_cashouts.clear()
        await real_sleep(0)
    instant.rounds = 0
    f.asyncio.sleep = instant
    try:
        waited = await f._wait_for_pending_cashouts()
        check('midnight holds while a cashout is unanswered', waited >= 3, str(waited))
        check('and stops as soon as the /out lands', not f._pending_cashouts)

        # ...but not for ever. A fresh stub: this one never drains the queue.
        async def never_drains(_):
            await real_sleep(0)
        f.asyncio.sleep = never_drains
        f._pending_cashouts[CHIMEREV] = [{'text': 'never answered'}]
        original, f.REPORT_WAIT_MINUTES = f.REPORT_WAIT_MINUTES, 5
        try:
            waited = await f._wait_for_pending_cashouts()
            check('a request nobody answers does not hold it for ever', waited == 5,
                  str(waited))
        finally:
            f.REPORT_WAIT_MINUTES = original
    finally:
        f.asyncio.sleep = real_sleep
        f._pending_cashouts.clear()

    # -- the clock lands on the configured local time -----------------------
    seconds = f._seconds_until_report(
        datetime(2026, 8, 10, 18, 0, tzinfo=f.LOCAL_TZ).astimezone(timezone.utc))
    check('midnight is 6 hours after 18:00 Nepal', seconds == 6 * 3600, str(seconds))
    check('it is always in the future', f._seconds_until_report() > 0)

    print()
    if failures:
        print(f"{len(failures)} FAILED: " + '; '.join(failures))
        sys.exit(1)
    print(f"all report checks passed")


asyncio.run(main())
